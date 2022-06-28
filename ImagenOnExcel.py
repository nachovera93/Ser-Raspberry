from openpyxl import Workbook
from datetime import date, timedelta
from datetime import datetime
from openpyxl.styles import Alignment  
from openpyxl.chart import (
    PieChart,
    BarChart,
    ProjectedPieChart,
    Reference
)
import datetime
import openpyxl
from openpyxl.chart.series import DataPoint
import os
import numpy as np


workbook=openpyxl.load_workbook(filename = '2022-06-16.xlsx')
sheet1 = workbook[f"REDCompañia-Fase-1"]
sheet2 = workbook[f"REDCompañia-Fase-2"]
sheet3 = workbook[f"REDCompañia-Fase-3"]
sheet4 = workbook[f"CentralFotovoltaica-Fase-1"]
sheet5 = workbook[f"CentralFotovoltaica-Fase-2"]
sheet6 = workbook[f"CentralFotovoltaica-Fase-3"]
sheet7 = workbook[f"ConsumoCliente-Fase-1"]
sheet8 = workbook[f"ConsumoCliente-Fase-2"]
sheet9 = workbook[f"ConsumoCliente-Fase-3"]
LargeSheet11=len(sheet1["FP"])
LargeSheet21=len(sheet2["FP"])
LargeSheet31=len(sheet3["FP"])
LargeSheet41=len(sheet4["FP"])
LargeSheet51=len(sheet5["FP"])
LargeSheet61=len(sheet6["FP"])
LargeSheet71=len(sheet7["FP"])
LargeSheet81=len(sheet8["FP"])
LargeSheet91=len(sheet9["FP"])
Energy_1 = float(sheet1[f'm{LargeSheet11}'].value)
Energy_2 = float(sheet2[f'm{LargeSheet21}'].value)
Energy_3 = float(sheet3[f'm{LargeSheet31}'].value)
Energy_4 = float(sheet4[f'm{LargeSheet41}'].value)
Energy_5 = float(sheet5[f'm{LargeSheet51}'].value)
Energy_6 = float(sheet6[f'm{LargeSheet61}'].value)
Energy_7 = float(sheet7[f'm{LargeSheet71}'].value)
Energy_8 = float(sheet8[f'm{LargeSheet81}'].value)
Energy_9 = float(sheet9[f'm{LargeSheet91}'].value)

Pos=len(sheet1['A'])


headings0 = ['Fecha y Hora'] + list(['Energia1','Energia2','Energia3'])
headings1 = ['Fecha y Hora'] + list(['Energia4','Energia5','Energia6'])
headings2 = ['Fecha y Hora'] + list(['Energia7','Energia8','Energia9'])
ceros=list([0,0,0,0,0,0,0,0,0])
datetim=datetime.datetime.now()-datetime.timedelta(minutes=5)

book = Workbook()
sheet22  = book.create_sheet("Sheet1")
sheet22.append(headings0)
sheet23  = book.create_sheet("Sheet2")
sheet23.append(headings1)
sheet24  = book.create_sheet("Sheet3")
sheet24.append(headings2)
book.save(filename = 'pie.xlsx')


workbook2=openpyxl.load_workbook(filename = 'pie.xlsx')
sheet22 = workbook2[f"Sheet1"]
sheet23 = workbook2[f"Sheet2"]
sheet24 = workbook2[f"Sheet3"]

DataHourFase1=[f'{datetim.day}-{datetim.month}-{datetim.year}',round(Energy_1,5),round(Energy_2,5),round(Energy_3,5)]
DataHourFase2=[f'{datetim.day}-{datetim.month}-{datetim.year}',round(Energy_4,5),round(Energy_5,5),round(Energy_6,5)]
DataHourFase3=[f'{datetim.day}-{datetim.month}-{datetim.year}',round(Energy_7,5),round(Energy_8,5),round(Energy_9,5)]

sheet22.append(list(DataHourFase1))
sheet23.append(list(DataHourFase2))
sheet24.append(list(DataHourFase3))


chart = BarChart()
chart.title = "Energias Totales Diarias"
chart.style = 12
chart.x_axis.title = 'Fecha'
chart.y_axis.title = 'KWh'
chart.type = "col"
chart.height = 10
chart.width = 30

cats = Reference(sheet22, min_col=1, min_row=2, max_row=Pos+1)
data = Reference(sheet22, min_col=2, min_row=1, max_col=10, max_row=Pos+1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

sheet22.add_chart(chart, "K1")



today = date.today()
print("El mes actual es {}".format(today.month))
mes=today.month
dia=date.today()-timedelta(1)
dia=str(dia)
print(dia)
Energy_1_TotalMes = []
Energy_2_TotalMes = []
Energy_3_TotalMes = []
Energy_4_TotalMes = []
Energy_5_TotalMes = []
Energy_6_TotalMes = []
Energy_7_TotalMes = []
Energy_8_TotalMes = []
Energy_9_TotalMes = []
#sheet22.delete_rows(11)
ejemplo_dir = '/Users/ignaciovera/Desktop/Codigos/SER-Raspberry/Ser-Raspberry/'
x=0
for f in os.listdir('/Users/ignaciovera/Desktop/Codigos/SER-Raspberry/Ser-Raspberry/'):
    if(dia[5:7]==f[5:7]):
        x = x+1
        #print(f)  
        workbook=openpyxl.load_workbook(filename = f'{f}')
        try:
            
            sheet1 = workbook[f"REDCompañia-Fase-1"]
            sheet2 = workbook[f"REDCompañia-Fase-2"]
            sheet3 = workbook[f"REDCompañia-Fase-3"]
            sheet4 = workbook[f"CentralFotovoltaica-Fase-1"]
            sheet5 = workbook[f"CentralFotovoltaica-Fase-2"]
            sheet6 = workbook[f"CentralFotovoltaica-Fase-3"]
            sheet7 = workbook[f"ConsumoCliente-Fase-1"]
            sheet8 = workbook[f"ConsumoCliente-Fase-2"]
            sheet9 = workbook[f"ConsumoCliente-Fase-3"]
            LargeSheet11=len(sheet1["FP"])
            LargeSheet21=len(sheet2["FP"])
            LargeSheet31=len(sheet3["FP"])
            LargeSheet41=len(sheet4["FP"])
            LargeSheet51=len(sheet5["FP"])
            LargeSheet61=len(sheet6["FP"])
            LargeSheet71=len(sheet7["FP"])
            LargeSheet81=len(sheet8["FP"])
            LargeSheet91=len(sheet9["FP"])
            Energy_1 = float(sheet1[f'm{LargeSheet11}'].value)
            Energy_2 = float(sheet2[f'm{LargeSheet21}'].value)
            Energy_3 = float(sheet3[f'm{LargeSheet31}'].value)
            Energy_4 = float(sheet4[f'm{LargeSheet41}'].value)
            Energy_5 = float(sheet5[f'm{LargeSheet51}'].value)
            Energy_6 = float(sheet6[f'm{LargeSheet61}'].value)
            Energy_7 = float(sheet7[f'm{LargeSheet71}'].value)
            Energy_8 = float(sheet8[f'm{LargeSheet81}'].value)
            Energy_9 = float(sheet9[f'm{LargeSheet91}'].value)   
            Energy_1_TotalMes.append(Energy_1)
            Energy_2_TotalMes.append(Energy_2)
            Energy_3_TotalMes.append(Energy_3)
            Energy_4_TotalMes.append(Energy_4)
            Energy_5_TotalMes.append(Energy_5)
            Energy_6_TotalMes.append(Energy_6)
            Energy_7_TotalMes.append(Energy_7)
            Energy_8_TotalMes.append(Energy_8)
            Energy_9_TotalMes.append(Energy_9)
            print(f)
        except:
            continue
print(Energy_1_TotalMes)
print(np.sum(Energy_1_TotalMes))
print(np.sum(Energy_2_TotalMes))
print(np.sum(Energy_3_TotalMes))
print(np.sum(Energy_4_TotalMes))
print(np.sum(Energy_5_TotalMes))
print(np.sum(Energy_6_TotalMes))
print(np.sum(Energy_7_TotalMes))
print(np.sum(Energy_8_TotalMes))
print(np.sum(Energy_9_TotalMes))
Suma_Mes_1=np.sum(Energy_1_TotalMes)
Suma_Mes_2=np.sum(Energy_2_TotalMes)
Suma_Mes_3=np.sum(Energy_3_TotalMes)
Suma_Mes_4=np.sum(Energy_4_TotalMes)
Suma_Mes_5=np.sum(Energy_5_TotalMes)
Suma_Mes_6=np.sum(Energy_6_TotalMes)
Suma_Mes_7=np.sum(Energy_7_TotalMes)
Suma_Mes_8=np.sum(Energy_8_TotalMes)
Suma_Mes_9=np.sum(Energy_9_TotalMes)
 
sheet22['E1'] = 'Acumulado Energia Mes REDCompañia-Fase-1'  
sheet22['F1'] = 'Acumulado Energia REDCompañia-Fase-2'
sheet22['G1'] = 'Acumulado Energia REDCompañia-Fase-3'
sheet23['E1'] = 'Acumulado Energia CentralFotovoltaica-Fase-1'
sheet23['F1'] = 'Acumulado Energia CentralFotovoltaica-Fase-2'
sheet23['G1'] = 'Acumulado Energia CentralFotovoltaica-Fase-3'
sheet24['E1'] = 'Acumulado Energia ConsumoCliente-Fase-1'
sheet24['F1'] = 'Acumulado Energia ConsumoCliente-Fase-2'
sheet24['G1'] = 'Acumulado Energia ConsumoCliente-Fase-3'

sheet22['E2'] = Suma_Mes_1
sheet22['F2'] = Suma_Mes_2
sheet22['G2'] = Suma_Mes_3
sheet23['E2'] = Suma_Mes_4
sheet23['F2'] = Suma_Mes_5
sheet23['G2'] = Suma_Mes_6
sheet24['E2'] = Suma_Mes_7
sheet24['F2'] = Suma_Mes_8
sheet24['G2'] = Suma_Mes_9
workbook2.save("pie.xlsx")
"""
from ast import For
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import (
    AreaChart,
    Reference,
    Series,
)
import datetime
from datetime import date
from datetime import time
import openpyxl
dest_filename = '2022-06-16.xlsx'
workbook=openpyxl.load_workbook(filename = dest_filename)
sheet11 = workbook[f'MaxHora Fase 1']
sheet12 = workbook[f"MaxHora Fase 2"]
sheet13 = workbook[f"MaxHora Fase 3"]

#print("Ultima columna", len(sheet11['A']))    
codigos0 = [celda[0].value for celda in sheet11['A2':f"A{len(sheet11['A']) }"]] #{sheet11.max_column}
print(codigos0)
#horas=[]
#for hour in codigos0:
#    print(hour)
#    hora=hour.hour
#    my_time = time(hora, 0, 0)
    #horas.append(f'{datetim.hour}:{datetim.minute}')
#print(len(horas))  
#Separar fases en excel
codigos1 = [celda[0].value for celda in sheet11['B2':f"B{len(sheet11['A']) }"]]
codigos2 = [celda[0].value for celda in sheet11['C2':f"C{len(sheet12['A']) }"]]
codigos3 = [celda[0].value for celda in sheet11['D2':f"D{len(sheet13['A']) }"]]
codigos4 = [celda[0].value for celda in sheet12['B2':f"B{len(sheet11['A']) }"]]
codigos5 = [celda[0].value for celda in sheet12['C2':f"C{len(sheet12['A']) }"]]
codigos6 = [celda[0].value for celda in sheet12['D2':f"D{len(sheet13['A']) }"]]
codigos7 = [celda[0].value for celda in sheet13['B2':f"B{len(sheet11['A']) }"]]
codigos8 = [celda[0].value for celda in sheet13['C2':f"C{len(sheet12['A']) }"]]
codigos9 = [celda[0].value for celda in sheet13['D2':f"D{len(sheet13['A']) }"]]

import pandas as pd 
df1 = pd.DataFrame({'Hora':horas,'Energia1':codigos1,
                                 'Energia2':codigos2,
                                 'Energia3':codigos3
                                 })
df2 = pd.DataFrame({'Hora':horas,'Energia4':codigos4,
                                 'Energia5':codigos5,
                                 'Energia6':codigos6
                                 })
df3 = pd.DataFrame({'Hora':horas,'Energia7':codigos7,
                                 'Energia8':codigos8,
                                 'Energia9':codigos9
                                 })

#new_df=df.assign(Profit=6)
#print(df1)
#print(df2)
#print(df3)
#arr = np.array(df)


#wb = Workbook()
#ws = wb.active
#headings1=list(['Hora','Energia-Fase-1-REDCompañia', 'Energia-Fase-1-CentralFotovoltaica','Energia-Fase-1-ConsumoCliente'])
#headings2=list(['Hora','Energia-Fase-1-REDCompañia', 'Energia-Fase-1-CentralFotovoltaica','Energia-Fase-1-ConsumoCliente'])
#headings3=list(['Hora','Energia-Fase-1-REDCompañia', 'Energia-Fase-1-CentralFotovoltaica','Energia-Fase-1-ConsumoCliente'])#,'Energia-Fase-2-REDCompañia','Energia-Fase-2-CentralFotovoltaica','Energia-Fase-2-ConsumoCliente','Energia-Fase-3-REDCompañia','Energia-Fase-3-CentralFotovoltaica','Energia-Fase-3-ConsumoCliente'])
#sheet11.append(headings1)
#sheet12.append(headings2)
#sheet13.append(headings3)
df_list1=df1.values.tolist()
df_list2=df2.values.tolist()
df_list3=df3.values.tolist()
arr = np.array(df1)
print(np.shape(arr))

for row in df_list1:
    sheet11.append(row)
for row in df_list2:
    sheet12.append(row)
for row in df_list3:
    sheet13.append(row)

chart = AreaChart()
chart.title = "Energias Fase 1"
chart.style = 13
chart.x_axis.title = 'Hora'
chart.y_axis.title = 'KWh'

cats = Reference(sheet11, min_col=1, min_row=2, max_row=f'{sheet11.max_column+1}')
data = Reference(sheet11, min_col=2, min_row=1, max_col=f'{sheet11.max_column+1}', max_row=f'{sheet11.max_column+1}')
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

sheet11.add_chart(chart, f"A{sheet11.max_column+1}")

workbook.save("2022-06-16.xlsx")

"""
