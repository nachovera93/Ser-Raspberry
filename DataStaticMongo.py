from connector import iot_ser_db
from connector import local_db
import json
import pymongo
import datetime
import os
import openpyxl
import pandas as pd

userId="63d87d13afeae04cb7827725"
dId=121212
db = iot_ser_db()

def list_data_db_insert_dispositive_historic(df):
    df_with_id=df.assign(userId=userId)
    list_filtered = df_with_id.to_dict('records')
    col_use = db.Dispositive_Variables
    col_use.insert_many(list_filtered)


def list_data_db_insert_max_day(df):
    df_with_id=df.assign(userId=userId)
    list_filtered = df_with_id.to_dict('records')
    col_use = db.monthmaxs
    col_use.insert_many(list_filtered)

def list_data_db_insert_max_month(df,i):
    print("Paso2")
    df_with_id=df.assign(userId=userId,dId=dId)
    #print(df_with_id.columns)
    df=pd.DataFrame(df_with_id,columns=['time',f'energia_fase_{i}_redcompañia_mensual',f'energia_fase_{i}_consumocliente_mensual',f'energia_fase_{i}_centralfotovoltaica_mensual','userId','dId'])
    print(df)
    col_use = db.monthmaxs 
    col_use.insert_many(df.to_dict('records'))
    

def main():
    #DiasExcel = [filecsv for filecsv in os.listdir('Reportes/') if filecsv[10:15]!="_save"]                   
    #DiasExcel=sorted(DiasExcel)
    #for archivo_csv in DiasExcel:
    archivo_csv=f'Reportes/2022-09-30.xlsx'
    try:
        #VariablesDispositivos(archivo_csv)
        MaximosDiariosMesFase1(archivo_csv) 
        MaximosDiariosMesFase2(archivo_csv)  
        MaximosDiariosMesFase3(archivo_csv)  
    except Exception as e:
         print(f"El formato del archivo no es el correspondiente: {archivo_csv} -> {str(e)} ") 

def MaximosDiariosMesFase2(archivo_csv):
    print("Paso")
    df=pd.read_excel(open(archivo_csv, 'rb'),sheet_name='Energia Fase 2 Mensual')
    try:
        df.columns = ['time','energia_fase_2_redcompañia_mensual','energia_fase_2_consumocliente_mensual','energia_fase_2_centralfotovoltaica_mensual','save','max1','max2','max3']
        #print(df[0:1])
        df=TimetoTimestampdf(df)
        workbook=openpyxl.load_workbook(filename = archivo_csv)
        sheet = workbook[f"Energia Fase 2 Mensual"]
        LargeSheet=sheet.max_row
        #if(sheet[f'E{LargeSheet-1}']==1):
        #print(df)
        list_data_db_insert_max_month(df,2)
        print("Insertando save = 1")
        for row in range(1, LargeSheet):
          sheet[f'E{row+1}']=1
        workbook.save(filename = archivo_csv)
            #os.rename(f'Reportes/{f}', f'Reportes/{f[0:10]}_save.xlsx')
            
    except Exception as e:
        print(f"Excepcion en Reporte: {archivo_csv} -> {str(e)} ")

def MaximosDiariosMesFase1(archivo_csv):
    print("Paso")
    df=pd.read_excel(open(archivo_csv, 'rb'),sheet_name='Energia Fase 1 Mensual')
    try:
        df.columns = ['time','energia_fase_1_redcompañia_mensual','energia_fase_1_consumocliente_mensual','energia_fase_1_centralfotovoltaica_mensual','save','max1','max2','max3']
        #print(df[0:1])
        df=TimetoTimestampdf(df)
        workbook=openpyxl.load_workbook(filename = archivo_csv)
        sheet = workbook[f"Energia Fase 1 Mensual"]
        LargeSheet=sheet.max_row
        #if(sheet[f'E{LargeSheet-1}']==1):
        #print(df)
        list_data_db_insert_max_month(df,1)
        print("Insertando save = 1")
        for row in range(1, LargeSheet):
          sheet[f'E{row+1}']=1
        workbook.save(filename = archivo_csv)
            #os.rename(f'Reportes/{f}', f'Reportes/{f[0:10]}_save.xlsx')
            
    except Exception as e:
        print(f"Excepcion en Reporte: {archivo_csv} -> {str(e)} ")

def MaximosDiariosMesFase3(archivo_csv):
    print("Paso")
    df=pd.read_excel(open(archivo_csv, 'rb'),sheet_name='Energia Fase 1 Mensual')
    try:
        df.columns = ['time','energia_fase_3_redcompañia_mensual','energia_fase_3_consumocliente_mensual','energia_fase_3_centralfotovoltaica_mensual','save','max1','max2','max3']
        #print(df[0:1])
        df=TimetoTimestampdf(df)
        workbook=openpyxl.load_workbook(filename = archivo_csv)
        sheet = workbook[f"Energia Fase 3 Mensual"]
        LargeSheet=sheet.max_row
        #if(sheet[f'E{LargeSheet-1}']==1):
        #print(df)
        list_data_db_insert_max_month(df,3)
        print("Insertando save = 1")
        for row in range(1, LargeSheet):
          sheet[f'E{row+1}']=1
        workbook.save(filename = archivo_csv)
            #os.rename(f'Reportes/{f}', f'Reportes/{f[0:10]}_save.xlsx')
            
    except Exception as e:
        print(f"Excepcion en Reporte: {archivo_csv} -> {str(e)} ")

def MaximosDiarios(archivo_csv):
    df=pd.read_excel(open(f'Reportes/{archivo_csv}', 'rb'),sheet_name='MaxHora Fase 1 Diario')
    try:
        df.columns = ['Fecha_y_Hora','Energia_Fase_1_REDCompañia_Diario','Energia_Fase_1_ConsumoCliente_Diario','Energia_Fase_1_CentralFotovoltaica_Diario']
        #print(df[0:2])
        c=-1
        for i in df['Fecha_y_Hora'].values:
            c=c+1
            df['Fecha_y_Hora'].values[c]=f"{archivo_csv[0:10]} {df['Fecha_y_Hora'].values[c]}:00"
        print(df[0:1])
        df=TimetoTimestampdf(df)
        workbook=openpyxl.load_workbook(filename = f'Reportes/{archivo_csv}')
        sheet = workbook[f"MaxHora Fase 1 Diario"]
        #DA = workbook.get_sheet_by_name('MaxHora Fase 1 Diario')
        LargeSheet=sheet.max_row
        if(sheet[f'E{LargeSheet-1}']!=1):
            list_data_db_insert_max_day(df)
            print("Insertando save = 1")
            for row in range(1, LargeSheet):
              sheet[f'E{row+1}']=1
        workbook.save(filename = f'Reportes/{archivo_csv}')
            #os.rename(f'Reportes/{f}', f'Reportes/{f[0:10]}_save.xlsx')
            
    except Exception as e:
        print(f"Excepcion en Reporte: {archivo_csv} -> {str(e)} ")

def VariablesDispositivos(archivo_csv):
    df=pd.read_excel(open(f'Reportes/{archivo_csv}', 'rb'),sheet_name='Var Dispositivos')
    try:
        df.columns = ['Fecha_y_Hora','T_Raspberry','Uso_CPU','RAM','T_ESP32']
        df=TimetoTimestampdf(df)
        workbook=openpyxl.load_workbook(filename = f'Reportes/{archivo_csv}')
        sheet = workbook[f"Var Dispositivos"]
        LargeSheet=len(sheet["RAM"])
        if(sheet[f'F{LargeSheet-1}']!=1):
            list_data_db_insert_dispositive_historic(df)
            print("Insertando save = 1")
            for row in range(1, LargeSheet):
              sheet[f'F{row+1}']=1
            #os.rename(f'Reportes/{f}', f'Reportes/{f[0:10]}_save.xlsx')
        workbook.save(filename = f'Reportes/{archivo_csv}')
    except Exception as e:
                    print(f"Excepcion en Reporte: {archivo_csv} -> {str(e)} ")

def TimetoTimestampdf(df):
    try:
        print("TimetoTimestampdf")
        c=-1
        for i in df['time'].values:
           c=c+1
           date_time_str = str(df['time'].values[c])
           date_time_str=date_time_str[0:19]
           date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d')
           #date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d %H:%M:%S')
           #print('Time:', date_time_obj.time())##print('Date:', date_time_obj.date())#print('Date-time:', (date_time_obj))
           df['time'].values[c]=int(date_time_obj.timestamp())
        return df
    except Exception as e:
        print(f"Excepcion en TimetoTimestampdf -> {str(e)} ") 


#mycol = db["monthmaxs"]
#mycol.drop()#
main()


#mycol = db.maximos_mensual
#query={"Energia_Fase_3_REDCompañia_Mensual": float("NaN")}
#d = mycol.delete_many(query)
#print(d.deleted_count, " documents deleted !!")

#wb = openpyxl.load_workbook('Reportes/2022-10-06.xlsx')

# Primero vemos los nombres de las hojas que tiene el libro
# Devuelve una lista de python
#hojas = wb.get_sheet_names()
#print(hojas)
# Escojo la primera por su nombre
# La variable es un objeto <sheet>
#DA = wb.get_sheet_by_name('MaxHora Fase 1 Diario')
#sheet = wb[f"MaxHora Fase 1 Diario"]
#print("Ultima columna", sheet.max_row)
# Una columna
#A = DA['A']
#
#ultima_fila = 0
#for i, celda in enumerate(A):
#    if celda.value is None:
#        ultima_fila = i
#        break
#
#print(ultima_fila)