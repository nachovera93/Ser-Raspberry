import requests
from datetime import date
from datetime import datetime
import json
import xlsxwriter
import random
import time
import paho.mqtt.client as mqtt
import os
import threading
import datetime
from scipy import interpolate
from scipy.fft import fft, fftfreq
from scipy.interpolate import lagrange
from scipy import signal
from scipy.signal import savgol_filter
import numpy as np
import subprocess
from time import sleep
import sys
import socket
import RPi.GPIO as GPIO
import time
import glob
import os
import serial
import math
#import board
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.chart import LineChart,Reference
import openpyxl
import smtplib, ssl
import getpass
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.mime.base import MIMEBase
import datetime
import matplotlib.pyplot as plt
import collections
import psutil
import gzip
"""
    0: connection succeeded
    1: connection failed - incorrect protocol version
    2: connection failed - invalid client identifier
    3: connection failed - the broker is not available
    4: connection failed - wrong username or password
    5: connection failed - unauthorized
    6-255: undefined
    """
"""
from board import SCL, SDA
import busio
from PIL import Image, ImageDraw, ImageFont
import adafruit_ssd1306
import time

i2c = busio.I2C(SCL, SDA)
disp = adafruit_ssd1306.SSD1306_I2C(128, 64, i2c)

disp.fill(0)
disp.show()

image = Image.new('1', (128, 64))

draw = ImageDraw.Draw(image)

font = ImageFont.load_default()

#  Escribe 2 lineas texto
draw.text((50, 16),    'Iniciando ..',  font=font, fill=255)
disp.image(image)
disp.show()
#import pywhatkit
#pywhatkit.sendwhatmsg("+56945959125", "Hi",15,36)
"""


def BorrarArchivos():

      current_time = time.time()
      for f in os.listdir():
            creation_time = os.path.getctime(f)
            if (((current_time - creation_time) // (24 * 3600)) >= 7):
                os.unlink(f)

#BorrarArchivos()
try:                           
     esp32 = serial.Serial('/dev/ttyUSB0', 230400, timeout=0.5)
     esp32.flushInput()
except:
     esp32 = serial.Serial('/dev/ttyUSB1', 230400, timeout=0.5)
     esp32.flushInput()
     
horasetup=datetime.datetime.now()
print("Hora de comienzo:", horasetup)

broker = '18.228.175.193'    #mqtt server
port = 1883
dId = '123454321'
passw = '25ebiBqWgR'
webhook_endpoint = 'http://18.228.175.193:3001/api/getdevicecredentials'


def get_mqtt_credentials():
    global usernamemqtt
    global passwordmqtt
    global mqttopic
    global str_client_id
    global topicmqtt
    global data
    print("Getting MQTT Credentials from WebHook")
    time.sleep(2)
    toSend = {"dId": dId, "password": passw}
    respuesta = requests.post(webhook_endpoint, data=toSend)

    if(respuesta.status_code < 0):
          print("Error Sending Post Request ", respuesta.status_code)
          respuesta.close()
          return False
    if(respuesta.status_code != 200):
          print("Error in response ", respuesta.status_code)
          respuesta.close()
          return False
    if(respuesta.status_code == 200):
          print("Mqtt Credentials Obtained Successfully :)   ")
          #print("json: " ,resp.content)
          #print('- ' * 20)
          my_bytes_value = respuesta.content      #Contenido entero del json
          my_new_string = my_bytes_value.decode("utf-8").replace("'", '"')
          data = json.loads(my_new_string)
          s = json.dumps(data, indent=4, sort_keys=True)
          print(s)
          usernamemqtt = data["username"]
          #print("username:",usernamemqtt)
          passwordmqtt = data["password"]
          topicmqtt = data["topic"]   #topico al que nos vamos a suscribir
          mqttopic = f"{topicmqtt}+/actdata"
          str_client_id = f'device_{dId}_{random.randint(0, 9999)}'
          #print(mqttopic)
          respuesta.close()
          print("Ends mqtt credentials")
    return True

   
def on_disconnect(client, userdata, rc):
    if (rc != 0 and rc != 5):
        print("Unexpected disconnection, will auto-reconnect")
    elif(rc==5):
        print("Getting new credentials!")
        get_mqtt_credentials()
        client.username_pw_set(usernamemqtt, passwordmqtt)
                      
def on_connected(client, userdata, flags, rc):
  
    if rc==0:
        client.connected_flag=True #set flag
        client.subscribe(mqttopic)
        print("connected OK")
        print("rc =",client.connected_flag)
    else:
        print("Bad connection Returned code=",rc)
        client.bad_connection_flag=False

   

get_mqtt_credentials()     
client = mqtt.Client(str_client_id)   #Creación cliente
client.connect(broker, port)     #Conexión al broker
client.on_disconnect = on_disconnect
client.username_pw_set(usernamemqtt, passwordmqtt)
client.on_connect = on_connected
client.loop_start()

def reconnectmqtt():
    get_mqtt_credentials()     
    client = mqtt.Client(str_client_id)   #Creación cliente
    client.connect(broker, port)     #Conexión al broker
    client.on_disconnect = on_disconnect
    client.username_pw_set(usernamemqtt, passwordmqtt)
    client.on_connect = on_connected
    client.loop_start()
  

def setup():
    GPIO.setmode(GPIO.BCM) 
    GPIO.setup(4,GPIO.OUT)
    GPIO.setup(24,GPIO.IN)
    GPIO.setup(23, GPIO.OUT)
    GPIO.setup(8, GPIO.OUT)
    GPIO.output(8, GPIO.LOW)
    GPIO.setwarnings(False)
    return()


setup()

def get_cpuload():
    cpuload = psutil.cpu_percent(interval=1, percpu=False)
    return str(cpuload)

def cpu_temp():
	thermal_zone = subprocess.Popen(
	    ['cat', '/sys/class/thermal/thermal_zone0/temp'], stdout=subprocess.PIPE)
	out, err = thermal_zone.communicate()
	cpu_temp = int(out.decode())/1000
	return cpu_temp


CPU_temp = 0.0
EstateVentilador="OFF"

def Ventilador():
    global EstateVentilador
    CPU_temp = round(cpu_temp(),0)
    if CPU_temp > 53:
        GPIO.output(23, True)
        str_num = {"value":"ON","save":0}
        EstateVentilador = json.dumps(str_num)
    elif CPU_temp <= 38:
        GPIO.output(23, False)
        str_num = {"value":"OFF","save":0}
        EstateVentilador = json.dumps(str_num)

def getMaxValues(myList, quantity):
        return(sorted(list(set(myList)), reverse=True)[:quantity]) 
        #print(f'max : {max(myList)}')


def getMinValues(myList, quantity):
        return(sorted(list(set(myList)))[:quantity]) 
        #print(f'max : {max(myList)}')


vrms=0.0
def VoltRms(maximovoltaje2):
    if(maximovoltaje2<13):
        vrms=0
        #print(f'vrms : {vrms}')
        return vrms
    vrms=maximovoltaje2*0.65
    #vrms=(-1.16 + 0.179*(maximovoltaje2) +  -0.00718*(maximovoltaje2**2) + 0.000155*(maximovoltaje2**3) + -0.00000203*(maximovoltaje2**4) + 0.0000000168*(maximovoltaje2**5) + -0.0000000000912*(maximovoltaje2**6) + 0.00000000000032*(maximovoltaje2**7) + -0.000000000000000703*(maximovoltaje2**8) + 0.000000000000000000875*(maximovoltaje2**9) + -0.000000000000000000000472*(maximovoltaje2**10))*maximovoltaje2
    #print(f'vrms : {vrms}')
    return vrms

irms=0.0
def CurrentRms(maximocorriente2):
    #irms=(-0.0248 + 0.00402*maximocorriente2 - 0.000176*(maximocorriente2**2) + 0.00000392*(maximocorriente2**3) - 0.000000046*(maximocorriente2**4) + 0.000000000284*(maximocorriente2**5) - 0.00000000000069*(maximocorriente2**6))*maximocorriente2
    #print(f'irms : {irms}')
    if(maximocorriente2>430):
         irms = maximocorriente2*0.0133
    else:
         irms=(0.0046 + 0.000282*maximocorriente2 - 0.00000328*(maximocorriente2**2) + 0.0000000167*(maximocorriente2**3) - 0.0000000000382*(maximocorriente2**4) + 0.0000000000000322*(maximocorriente2**5))*maximocorriente2
    
    return irms



vrms0=0.0
def VoltajeRms(listVoltage):
    global vrms0
    N = len(listVoltage)
    Squares = []

    for i in range(0,N,1):    #elevamos al cuadrado cada termino y lo amacenamos
         listsquare = listVoltage[i]*listVoltage[i]
         Squares.append(listsquare)
    
    SumSquares=0
    for i in range(0,N,1):    #Sumatoria de todos los terminos al cuadrado
         SumSquares = SumSquares + Squares[i]

    MeanSquares = (1/N)*SumSquares #Dividimos por N la sumatoria

    vrms0=np.sqrt(MeanSquares)
   
    #print(f'Voltaje RMS : {vrms0}')
    return vrms0


irms0=0.0
def CorrienteRms(listCurrent):
    global irms0
    
    N = len(listCurrent)
    Squares = []

    for i in range(0,N,1):    #elevamos al cuadrado cada termino y lo amacenamos
         listsquare = listCurrent[i]*listCurrent[i]
         Squares.append(listsquare)
    
    SumSquares=0
    for i in range(0,N,1):    #Sumatoria de todos los terminos al cuadrado
         SumSquares = SumSquares + Squares[i]

    MeanSquares = (1/N)*SumSquares #Dividimos por N la sumatoria

    irms0=np.sqrt(MeanSquares)
    
    #print(f'Corriente RMS : {irms0}')
    return irms0

#potrms=0.0
def PotenciaRms(listCurrent,listVoltage):
    global Squares
    
    N = len(listCurrent)
    Squares = []

    for i in range(0,N,1):    #elevamos al cuadrado cada termino y lo amacenamos
         listsquare = listCurrent[i]*listVoltage[i]
         Squares.append(listsquare)
    
    SumSquares=0
    for i in range(0,N,1):    #Sumatoria de todos los terminos al cuadrado
         SumSquares = SumSquares + Squares[i]

    MeanSquares = (1/N)*SumSquares #Dividimos por N la sumatoria

    #potrms=np.sqrt(MeanSquares)
    
    
    return MeanSquares

    

DATVoltajeCGE=0.0
phasevoltajeCGE=0.0
FDVoltajeCGE=0.0
DATVoltajePaneles=0.0
phasevoltajePaneles=0.0
FDVoltajePaneles=0.0
DATVoltajeCarga=0.0
phasevoltajeCarga=0.0
FDVoltajeCarga=0.0
FDVoltajeCGE1=0.0

def VoltageFFT(list_fftVoltages, samplings,i):
    global j
    j = str(i)
    global DATVoltajeCGE
    global ejeyabsolutv
    global DATVoltajeCGE1
    global phasevoltajeCGE
    global FDVoltajeCGE
    global DATVoltajePaneles
    global phasevoltajePaneles
    global FDVoltajePaneles
    global DATVoltajeCarga
    global phasevoltajeCarga
    global FDVoltajeCarga
    global xnewv
    global FDVoltajeCGE1
    N = len(list_fftVoltages)
    T = 1 / samplings
    list_fftVoltages -= np.mean(list_fftVoltages)
    datosfft = list_fftVoltages * np.hamming(4096)
    
    yf = np.fft.rfft(datosfft)
    xf = fftfreq(N, T)[:N//2] 
    if (samplings > 5100):
           f = interpolate.interp1d(xf, yf[:N//2] )
           xnewv = np.arange(0, 2575, 1)  # 2550
           ynew = f(xnewv)
           ejeyabsolutv =  2.0/4096 * np.abs(ynew)
           FD = []
           complejo = []
           real=[]
           imag=[]
           z=0
           for i in range(45, 2575, 50):
                 a2 = max(ynew[i:i+10])
                 arra = max(ejeyabsolutv[i:i+10])
                 complejo.append(a2)
                 real1 = a2.real
                 real.append(real1)
                 imag1 = a2.imag
                 imag.append(imag1)
                 z=z+1
                 FD.append(arra)
           FD2=[]       
           for i in range(0,len(FD)):
               if(FD[i]>(FD[0]/10)):
                   FD2.append(FD[i])
                   
           SumMagnitudEficaz = (np.sum([FD[0:len(FD)]]))
           Magnitud1 = FD[0]
           if(j=="1"):
                 global sincvoltaje1            
                 phasevoltajeCGE = np.arctan(real[0]/(imag[0]))
                 FDVoltajeCGE1 = Magnitud1/SumMagnitudEficaz
                 str_num = {"value":FDVoltajeCGE1,"save":1}
                 FDVoltajeCGE = json.dumps(str_num)
                 DATVoltajeCGE1= np.sqrt(((SumMagnitudEficaz**2)-(Magnitud1**2))/(Magnitud1**2))
                 str_num = {"value":DATVoltajeCGE1,"save":1}
                 DATVoltajeCGE = json.dumps(str_num)
                 sincvoltaje1 = 1
           if(j=="3"):
                 global sincvoltaje2              
                 phasevoltajePaneles = np.arctan(real[0]/(imag[0]))
                 FDVoltajePaneles = Magnitud1/SumMagnitudEficaz
                 str_num = {"value":FDVoltajePaneles,"save":1}
                 FDVoltajePaneles = json.dumps(str_num)
                 DATVoltajePaneles1 = np.sqrt(((SumMagnitudEficaz**2)-(Magnitud1**2))/(Magnitud1**2))
                 str_num = {"value":DATVoltajePaneles1,"save":1}
                 DATVoltajePaneles = json.dumps(str_num)
                 sincvoltaje2 = 1
           if(j=="2"):
                 global sincvoltaje3
                 phasevoltajeCarga = np.arctan(real[0]/(imag[0]))
                 #FaseArmonicoFundamentalVoltaje1=round(np.angle(complejo[0]),2)
                 FDVoltajeCarga = Magnitud1/SumMagnitudEficaz
                 str_num = {"value":FDVoltajeCarga,"save":1}
                 FDVoltajeCarga = json.dumps(str_num)
                 DATVoltajeCarga = np.sqrt(((SumMagnitudEficaz**2)-(Magnitud1**2))/(Magnitud1**2))
                 str_num = {"value":DATVoltajeCarga,"save":1}
                 DATVoltajeCarga = json.dumps(str_num)
                 sincvoltaje3 = 1


DATCorrienteCGE = 0.0
DATCorrientePaneles= 0.0
DATCorrienteCarga= 0.0
FDCorrienteCGE= 0.0
FDCorrienteCarga=0.0
DATCorrienteCarga=0.0
FDCorrientePaneles= 0.0
FDCorrienteCarga= 0.0
FDCorrienteCGE= 0.0
phasecorrienteCGE= 0.0
phasecorrientePaneles= 0.0
phasecorrienteCarga= 0.0
FPCGE= 0.99
FPCGE0= 0.99
cosphiCGE= 0.0
FPPaneles= 0.0
cosphiPaneles= 0.0
FPCarga= 0.0
cosphiCarga= 0.0
FDCorrienteCGE1=0.0
DATCorrienteCGE1=0.0
FPCarga1=0.99
FPPaneles1 = 0.99
FDCorrientePaneles1 = 0.0
DATCorrientePaneles1 = 0.0 
FDCorrienteCarga1=0.0 
DATCorrienteCarga1=0.0

def CurrentFFT(list_fftVoltages, samplings, i,irms):
    global DATCorrienteCGE
    global DATCorrienteCGE1
    global a2
    global FDCorrienteCGE 
    global FDCorrienteCGE1
    global phasecorrienteCGE
    global FDCorrientePaneles
    global DATCorrientePaneles
    global phasecorrientePaneles
    global FDCorrienteCarga
    global DATCorrienteCarga
    global FDCorrienteCarga1
    global DATCorrienteCarga1
    global phasecorrienteCarga
    global FPCGE
    global FPCarga1
    global FPCGE0
    global cosphiCGE
    global FPPaneles
    global cosphiPaneles
    global FPCarga
    global cosphiCarga
    global q
    global FPPaneles1
    global FDCorrientePaneles1
    global DATCorrientePaneles1      
    global real
    global imag
    global xnew
    global ejeyabsolut
    global desfaseCGE
    global desfaseCarga
    global desfasePaneles
   
    q = str(i)
    N = len(list_fftVoltages)
    T = 1 / samplings
    list_fftVoltages -= np.mean(list_fftVoltages)
    datosfft = list_fftVoltages * np.hamming(4096)
    yf = np.fft.rfft(datosfft)
    
    xf = fftfreq(N, T)[:N//2]
    if (samplings > 5100):
         f = interpolate.interp1d(xf,yf[:N//2])
         xnew = np.arange(0, 2575, 1)
         ynew = f(xnew)
         ejeyabsolut =  2.0/N * np.abs(ynew)
         p = int(i)
         z=0
         FD= []
         complejo = []
         real=[]
         imag=[]
         for i in range(45, 2575, 50):
               a2 = max(ynew[i:i+10])
               arra = max(ejeyabsolut[i:i+10])
               complejo.append(a2)
               real1 = a2.real
               real.append(real1)
               imag1 = a2.imag
               imag.append(imag1)
               FD.append(arra)
         repite = list(ejeyabsolut)
         #SumMagnitudEficaz = (np.sum([FD[0:len(FD)]]))*0.01
         SumMagnitudEficaz2 = (np.sum([FD[0:len(FD)]]))
         Magnitud1 = FD[0]#*0.01
         ArmonicosRestantes=SumMagnitudEficaz2-Magnitud1
         proporcion = irms/(np.sqrt(Magnitud1**2+ArmonicosRestantes**2))
         irmsarmonico1prop=Magnitud1*proporcion
         irmstotalproporcionado=np.sqrt((irmsarmonico1prop**2)+(ArmonicosRestantes*proporcion)**2)
         if(q=="1"):
             global sincvoltaje1
             FDCorrienteCGE1 = irmsarmonico1prop/irms
             str_num = {"value":FDCorrienteCGE1,"save":0}
             FDCorrienteCGE = json.dumps(str_num)
             DATCorrienteCGE1 = np.sqrt((SumMagnitudEficaz2**2-Magnitud1**2)/(Magnitud1**2))
             str_num2 = {"value":DATCorrienteCGE1,"save":0}
             DATCorrienteCGE = json.dumps(str_num2)
             phasecorrienteCGE = np.arctan(real[0]/(imag[0]))
             if (sincvoltaje1 == 1):
                 if(phasevoltajeCGE-(phasecorrienteCGE)>=0):
                     desfaseCGE = "Corriente Adelantada a Voltaje"
                 else:
                     desfaseCGE = "Voltaje Adelantado a Corriente"
                 FPCGE0=np.cos(phasevoltajeCGE-phasecorrienteCGE)*FDCorrienteCGE1#+0.05
                 cosphiCGE=np.cos(phasevoltajeCGE-phasecorrienteCGE)
                 if(FPCGE0>0.0):
                     FPCGE0=FPCGE0#+0.05
                 else:
                     FPCGE0=FPCGE0#-0.05
                 
                 str_num3 = {"value":FPCGE0,"save":0}
                 FPCGE = json.dumps(str_num3)
                 sincvoltaje1=0  
                 
         if(q=="3"):
             global sincvoltaje2
             FDCorrientePaneles1 = irmsarmonico1prop/irms
             str_num = {"value":FDCorrientePaneles1,"save":0}
             FDCorrientePaneles = json.dumps(str_num)
             DATCorrientePaneles1 = np.sqrt((SumMagnitudEficaz2**2-Magnitud1**2)/(Magnitud1**2))
             str_num2 = {"value":DATCorrientePaneles1,"save":0}
             DATCorrientePaneles = json.dumps(str_num2)
             phasecorrientePaneles = np.arctan(real[0]/(imag[0]))
             if (sincvoltaje2 == 1):
                 if(phasevoltajePaneles-(phasecorrientePaneles)>=0):
                     desfasePaneles = "Corriente Adelantada a Voltaje"
                 else:
                     desfasePaneles = "Voltaje Adelantado a Corriente"
                 FPPaneles1=np.cos(phasevoltajePaneles-(phasecorrientePaneles))*FDCorrientePaneles1
                 if (FPPaneles1>0.0):
                     FPPaneles1 = FPPaneles1#+0.05
                 else:
                     FPPaneles1 = FPPaneles1#-0.05
                
                 cosphiPaneles=np.cos(phasevoltajePaneles-phasecorrientePaneles)
                 str_num = {"value":FPPaneles1,"save":0}
                 FPPaneles = json.dumps(str_num)
                 sincvoltaje2=0  
         if(q=="2"):
             global sincvoltaje3
             FDCorrienteCarga1=irmsarmonico1prop/irms
             str_num = {"value":FDCorrienteCarga1,"save":0}
             FDCorrienteCarga = json.dumps(str_num)
             DATCorrienteCarga1 = np.sqrt((SumMagnitudEficaz2**2-Magnitud1**2)/(Magnitud1**2))
             str_num2 = {"value":DATCorrienteCarga1,"save":0}
             DATCorrienteCarga = json.dumps(str_num2)
             phasecorrienteCarga = np.arctan(real[0]/(imag[0]))
             if (sincvoltaje3 == 1):
                 if(phasevoltajeCarga-(phasecorrienteCarga)>=0):
                     desfaseCarga = "Corriente Adelantada a Voltaje"
                 else:
                     desfaseCarga = "Voltaje Adelantado a Corriente"
                 FPCarga1 = np.cos(phasevoltajeCarga-phasecorrienteCarga)*FDCorrienteCarga1# + 0.05
                 cosphiCarga=np.cos(phasevoltajeCarga-phasecorrienteCarga)
                 if(FPCarga1>0.0):
                     FPCarga1=FPCarga1#+0.05
                 else:
                     FPCarga1=FPCarga1#-0.05
                 str_num = {"value":FPCarga1,"save":0}
                 FPCarga = json.dumps(str_num)
                 sincvoltaje3=0


a = datetime.datetime.now()
b = datetime.datetime.now() 
c = datetime.datetime.now()  
#energyCGEFase11 = 0.0
energyCGEFase11 = 0.0
energyCGEFase11Hour = 0.0
energyCargaFase13 = 0.0
energyCargaFase13Hour = 0.0
AparenteCGEFase11 = 0.0
ActivaCGEFase11 = 0.0
ReactivaCGEFase11 = 0.0
AparentePanelesFase12 = 0.0
ActivaPanelesFase12 = 0.0
ReactivaPanelesFase12 = 0.0
AparenteCargaFase13 = 0.0
ActivaCargaFase13 = 0.0
ReactivaCargaFase13 = 0.0
energyCGEFase1 = 0.0
energyPanelesFase1 = 0.0
energyPanelesFase12Hour = 0.0
energyPanelesFase12 = 0.0
energyCargaFase1 = 0.0
AparenteCGEFase1 = 0.0
ActivaCGEFase1 = 0.0
ReactivaCGEFase1 = 0.0
AparentePanelesFase1 = 0.0
ActivaPanelesFase1 = 0.0
ReactivaPanelesFase1 = 0.0
AparenteCargaFase1 = 0.0
ActivaCargaFase1 = 0.0
ReactivaCargaFase1 = 0.0



def Potencias(i,irms,vrms):
    i = str(i)
    
    if(i=="1"):
          global a
          global energyCGEFase1
          global ActivaCGEFase1
          global AparenteCGEFase1
          global ReactivaCGEFase1
          global energyCGEFase11
          global energyCGEFase11Hour
          global AparenteCGEFase11
          global ActivaCGEFase11
          global ReactivaCGEFase11
          AparenteCGEFase11 = vrms*irms
          if (potrmsCGE>=0):
                ActivaCGEFase11 = vrms*irms*cosphiCGE
                ActivaCGEFase11 = np.abs(ActivaCGEFase11)
          else:
                ActivaCGEFase11 = vrms*irms*cosphiCGE
                ActivaCGEFase11 = np.abs(ActivaCGEFase11)
                ActivaCGEFase11 = ActivaCGEFase11*(-1)
          ReactivaCGEFase11 = vrms*irms*np.sin(phasevoltajeCGE-phasecorrienteCGE)
          a2 = datetime.datetime.now()
          delta=(((a2 - a).microseconds)/1000+((a2 - a).seconds)*1000)/10000000000
          energyCGEFase11 += np.abs(ActivaCGEFase11*delta*2.9)
          energyCGEFase11Hour += np.abs(ActivaCGEFase11*delta*2.9)
          a = datetime.datetime.now()
          if(a2.minute==0):
              energyCGEFase11Hour=0
          if(a2.hour==0 and a2.minute==0):
              energyCGEFase11=0
          if(a2.hour==0 and a2.minute==1):
              energyCGEFase11=0
          str_num = {"value":ActivaCGEFase11,"save":1}
          str_num2 = {"value":ReactivaCGEFase11,"save":0}
          str_num3 = {"value":AparenteCGEFase11,"save":1}
          str_num4 = {"value":energyCGEFase11 ,"save":0}
          ActivaCGEFase1 = json.dumps(str_num)
          AparenteCGEFase1 = json.dumps(str_num3)
          ReactivaCGEFase1 = json.dumps(str_num2)
          energyCGEFase1 = json.dumps(str_num4)
    if(i=="3"):
          global b
          global energyPanelesFase1
          global AparentePanelesFase1
          global ActivaPanelesFase1
          global ReactivaPanelesFase1
          global energyPanelesFase12
          global energyPanelesFase12Hour
          global AparentePanelesFase12
          global ActivaPanelesFase12
          global ReactivaPanelesFase12
          AparentePanelesFase12 = vrms*irms
          if(potrmsPaneles>=0):
                ActivaPanelesFase12= vrms*irms*cosphiPaneles
                ActivaPanelesFase12=np.abs(ActivaPanelesFase12)
          else:
                ActivaPanelesFase12= vrms*irms*cosphiPaneles
                ActivaPanelesFase12= np.abs(ActivaPanelesFase12)
                ActivaPanelesFase12= ActivaPanelesFase12*(-1)
          ReactivaPanelesFase12 = vrms*irms*np.sin(phasevoltajePaneles-phasecorrientePaneles)
          b2 = datetime.datetime.now()
          delta=(((b2 - b).microseconds)/1000+((b2 - b).seconds)*1000)/10000000000
          energyPanelesFase12Hour += np.abs(ActivaPanelesFase12*delta*2.9)
          energyPanelesFase12 += np.abs(ActivaPanelesFase12*delta*2.9)
          b = datetime.datetime.now()
          if(b2.minute==0):
              energyPanelesFase12Hour=0
          if(b2.hour==0 and b2.minute==0):
              energyPanelesFase12=0
          if(b2.hour==0 and b2.minute==1):
              energyPanelesFase12=0
          str_num = {"value":ActivaPanelesFase12,"save":1}
          str_num2 = {"value":ReactivaPanelesFase12,"save":0}
          str_num3 = {"value":AparentePanelesFase12,"save":1}
          str_num4 = {"value":energyPanelesFase12,"save":0}
          ActivaPanelesFase1 = json.dumps(str_num)
          AparentePanelesFase1 = json.dumps(str_num3)
          ReactivaPanelesFase1 = json.dumps(str_num2)
          energyPanelesFase1 = json.dumps(str_num4)
    if(i=="2"):
          global c
          global energyCargaFase1 
          global AparenteCargaFase1
          global ActivaCargaFase1
          global ReactivaCargaFase1
          global energyCargaFase13
          global energyCargaFase13Hour
          global AparenteCargaFase13
          global ActivaCargaFase13
          global ReactivaCargaFase13
          AparenteCargaFase13 = vrms*irms
          #print(f'Aparente Carga: {AparenteCargaFase13}')
          if(potrmsCarga>=0):
              ActivaCargaFase13= vrms*irms*cosphiCarga
              ActivaCargaFase13=np.abs(ActivaCargaFase13)
          else:
              ActivaCargaFase13= vrms*irms*cosphiCarga
              ActivaCargaFase13=np.abs(ActivaCargaFase13)
              ActivaCargaFase13=ActivaCargaFase13*(-1)
           
          #print(f'Activa Carga: {ActivaCargaFase13}')
          ReactivaCargaFase13 = vrms*irms*np.sin(phasevoltajeCarga-phasecorrienteCarga)
          c2 = datetime.datetime.now()
          #print(f'Reactiva Carga: {ReactivaCargaFase13}')
          delta=(((c2 - c).microseconds)/1000+((c2 - c).seconds)*1000)/10000000000
          energyCargaFase13 += np.abs(ActivaCargaFase13*delta*2.9)
          energyCargaFase13Hour += np.abs(ActivaCargaFase13*delta*2.9)
          c = datetime.datetime.now()
          if(c2.minute==0):
              energyCargaFase13Hour=0
          if(c2.hour==0 and c2.minute==0):
              energyCargaFase13=0
          if(c2.hour==0 and c2.minute==1):
              energyCargaFase13=0
          str_num = {"value":ActivaCargaFase13,"save":1}
          str_num2 = {"value":ReactivaCargaFase13,"save":0}
          str_num3 = {"value":AparenteCargaFase13,"save":1}
          str_num4 = {"value":energyCargaFase13,"save":0}
          ActivaCargaFase1 = json.dumps(str_num)
          AparenteCargaFase1 = json.dumps(str_num3)
          ReactivaCargaFase1 = json.dumps(str_num2)
          energyCargaFase1 = json.dumps(str_num4)
    
font = {'family': 'serif',
        'color':  'darkred',
        'weight': 'normal',
        'size': 8,
        }

def graphVoltage(list_fftVoltage,list_FPCurrent,samplings,i):
        global ax
        global imagenVoltaje
        i = str(i)
        global render
        fig=plt.figure(figsize=(8,6))
        tiempo = 1/(samplings*(0.001/4200))
        tiempoms = np.arange(0,tiempo,tiempo/4096)


        ax = fig.add_subplot(9,1,1)
        ax.plot(tiempoms,list_FPCurrent,color="green", label="Corriente")
        if(i=="1"):
             plt.title(f'Corriente | I: {round(irms1,2)}  |  P-Activa: {round(ActivaCGEFase11,2)} | P-Aparente: {round(AparenteCGEFase11,2)}  |  P-Reactiva:{round(ReactivaCGEFase11,2)}  ',fontdict=font)
        if(i=="2"):
             plt.title(f'Corriente | I: {round(irms2,2)}  |  P-Activa: {round(ActivaCargaFase13,2)} | P-Aparente: {round(AparenteCargaFase13,2)}  |  P-Reactiva:{round(ReactivaCargaFase13,2)}  ',fontdict=font)
        if(i=="3"):
             plt.title(f'Corriente | I: {round(irms3,2)}  |  P-Activa: {round(ActivaPanelesFase12,2)} | P-Aparente: {round(AparentePanelesFase12,2)}  |  P-Reactiva:{round(ReactivaPanelesFase12,2)}  ',fontdict=font)
             
        ax.legend(loc='upper left')
        ax.set_xlabel('Tiempo (mS)',fontdict=font)
        ax = fig.add_subplot(9,1,3)
        ax.plot(tiempoms,list_fftVoltage,color="blue", label="Voltaje")
        if(i=="1"):
             plt.title(f'Voltaje | V: {round(vrms1,2)} |  FP: {round(FPCGE0,2)}',fontdict=font)
        if(i=="2"):
             plt.title(f'Voltaje | V: {round(vrms2,2)} |  FP: {round(FPCarga1,2)}',fontdict=font)
        if(i=="3"):
             plt.title(f'Voltaje | V: {round(vrms3,2)} |  FP: {round(FPPaneles1,2)}',fontdict=font)
        
        ax = fig.add_subplot(9,1,5)
        ax.plot(tiempoms,Squares,color="red", label="Pot-Activa")

        ax = fig.add_subplot(9,1,7)
        if(i=="1"):
            plt.title(f'FFT Corriente | DAT: {round(DATCorrienteCGE1,2)}, FD: {round(FDCorrienteCGE1,2)} |   cos phi: {round(cosphiCGE,2)} | phase voltaje CGE : {round(phasevoltajeCGE,2)} | phase Corriente CGE : {round(phasecorrienteCGE,2)}',fontdict=font)
        if(i=="2"):
            plt.title(f'FFT Corriente | DAT: {round(DATCorrienteCarga1,2)}, FD: {round(FDCorrienteCarga1,2)} |   cos phi: {round(cosphiCarga,2)} | phase voltaje Carga : {round(phasevoltajeCarga,2)} | phase Corriente Carga : {round(phasecorrienteCarga,2)} ',fontdict=font)
        if(i=="3"):
            plt.title(f'FFT Corriente | DAT: {round(DATCorrientePaneles1,2)}, FD: {round(FDCorrientePaneles1,2)}  |   cos phi: {round(cosphiPaneles,2)} | phase voltaje Paneles: {round(phasevoltajePaneles,2)} | phase Corriente Paneles : {round(phasecorrientePaneles,2)}',fontdict=font)

        ax.plot(xnew,ejeyabsolut)
        rangex = np.zeros(28)
        n=0
        for h in range(50, 2600, 100):
           rangex[n]=h
           n = n+1
        ax.xaxis.set_ticks(rangex)  
        ax.grid(True)
        ax.set_xlabel('Frecuencia (Hz)',fontdict=font)
        
        ax = fig.add_subplot(9,1,9)
        if(i=="1"):
            plt.title(f'{desfaseCGE}',fontdict=font)
        if(i=="2"):
            plt.title(f'{desfaseCarga}',fontdict=font)
        if(i=="3"):
            plt.title(f'{desfasePaneles} ',fontdict=font)

        ax.plot(xnewv,ejeyabsolutv)
        rangex = np.zeros(28)
        n=0
        for h in range(50, 2600, 100):
           rangex[n]=h
           n = n+1
        ax.xaxis.set_ticks(rangex)  
        ax.grid(True)
        ax.set_xlabel('Frecuencia (Hz)',fontdict=font)
        #ax.set_ylabel('|dB|',fontdict=font) 

        oldepoch = time.time()
        st = datetime.datetime.fromtimestamp(oldepoch).strftime('%Y-%m-%d-%H:%M:%S') 
        #plt.xlabel("Tiempo(ms)",fontdict=font)
        #plt.title(f'FFT Voltaje |    DAT: {DATVoltajeCGE1}     |     FD: {FDVoltajeCGE1}    |   cos phi: {round(cosphiCGE,2)}',fontdict=font)
        ax.set_xlabel(f'Frecuencia (Hz)  |    fecha: {st} ',fontdict=font)
        #ax.set_ylabel('Pk-Pk',fontdict=font) 
        imagenVoltaje = f'images{i}/{st}.png'
        plt.savefig(imagenVoltaje)
        
        
"""        
def EnviarImagenes():
    #oldepoch = time.time()
    current_time = time.time()
    #st = datetime.datetime.fromtimestamp(oldepoch).strftime('%Y-%m-%d')
    #print(f' st Enviar Imagenes: {st}')
    for f in os.listdir('/home/pi/Desktop/IOTSER/images1/'):
            creation_time = os.path.getctime(f)
            if (((current_time - creation_time) // (24 * 3600)) >= 7):
                os.unlink(f)

EnviarImagenes()
"""
folder = "/home/pi/Desktop/IOTSER/images1/"
compress_older_days = 5

now = time.time()
 
 
def sanitize_files():
    # Loop through all the folder
    for file in os.listdir(folder): 
        f = os.path.join(folder,file)
        if not f.endswith('.gz'):
            if os.stat(f).st_mtime < now - (60*60*24*compress_older_days) and os.path.isfile(f):
                print ("...Compressing file "+f)
                out_filename = f + ".gz"
                 
                f_in = open(f, 'rb')
                s = f_in.read()
                f_in.close()
 
                f_out = gzip.GzipFile(out_filename, 'wb')
                f_out.write(s)
                f_out.close()
                # Remove original uncompressed file
                os.remove(f)     
        # We ensure that the file we are going to delete has been compressed before
        

def sanitize_files2():
    delete_older_days = 10
    # Loop through all the folder
    for file in os.listdir(folder): 
        f = os.path.join(folder,file)
        print ("...estrando"+f)
        if os.stat(f).st_mtime < now - (60*60*24*delete_older_days) and os.path.isfile(f):
                print ("...eliminando file "+f)
                # Remove original uncompressed file
                #os.remove(f) 
                os.system(f"sudo rm {f}")    

def sanitize_files3():
     
     path = f"/home/pi/Desktop/MedidorMonofasico/Ser-Raspberry/images1/"
     now = time.time()
     for i in path:
         for f in os.listdir(path):
            f = os.path.join(path, f)
            print(f"paso for {f}")
            if os.stat(os.path.join(path,f)).st_mtime < now - 4 * 86400:
                 #print("paso if")
                 if os.path.isfile(f):
                     print("paso if 2")
                     os.remove(f)

     path = f"/home/pi/Desktop/MedidorMonofasico/Ser-Raspberry/images2/"
     now = time.time()
     for i in path:
         for f in os.listdir(path):
            f = os.path.join(path, f)
            print(f"paso for {f}")
            if os.stat(os.path.join(path,f)).st_mtime < now - 4 * 86400:
                 #print("paso if")
                 if os.path.isfile(f):
                     print("paso if 2")
                     os.remove(f)
     path = f"/home/pi/Desktop/MedidorMonofasico/Ser-Raspberry/images3/"
     now = time.time()
     for i in path:
         for f in os.listdir(path):
            f = os.path.join(path, f)
            print(f"paso for {f}")
            if os.stat(os.path.join(path,f)).st_mtime < now - 4 * 86400:
                 #print("paso if")
                 if os.path.isfile(f):
                     print("paso if 2")
                     os.remove(f)
    

#sanitize_files()
#sanitize_files2()
#sanitize_files3()

vrms1=0.0
vrms2=0.0
vrms3=0.0
irms1=0.0
irms2=0.0
irms3=0.0
vrms11=0.0
vrms22=0.0
vrms33=0.0
irms11=0.0
irms22=0.0
irms33=0.0
VoltajeBaterias=0.0
CorrienteBaterias=0.0
tiempo2Bateria = datetime.datetime.now()  
tiempo2Paneles = datetime.datetime.now()  
energyBaterias = 0.0
energyBateriaHora = 0.0
energyPanelesDC=0.0
energyPanelesHoraDC=0.0
VoltajePanelesDC=0.0
CorrientePanelesDC=0.0
PotenciaBaterias=0.0
PotenciaPanelesDC=0.0
energyPanelesDCSend=0
VoltajePanelesDCSend=0
CorrientePanelesDCSend=0
PotenciaPanelesDCSend=0
energyBateriaSend=0          
PotenciaBateriaSend=0
VoltajeBateriaSend=0
CorrienteBateriaSend=0
modamaximovoltaje11=[]
modamaximocorriente11=[]
modamaximovoltaje22=[]
modamaximocorriente22=[]
modamaximovoltaje33=[]
modamaximocorriente33=[]
global accesoemail
accesoemail=0
global accesoemail2
accesoemail2=0
global accesoemail3
accesoemail3=0
global accesoexcel
accesoxcel=0
modamaximovoltajeBateria=[]
modamaximocorrientebateria=[]
modamaximovoltajePaneles=[]
modamaximocorrientePaneles=[]
countbroker=0
global mediadccorrienteCarga
global mediadccorrienteCGE


def received():
    while True:
                 try:
                     esp32_bytes = esp32.readline()
                     decoded_bytes = str(esp32_bytes[0:len(esp32_bytes)-2].decode("utf-8"))#utf-8
                 except:
                     print("Error en la codificación")
                     continue
                  
                 np_array = np.fromstring(decoded_bytes, dtype=float, sep=',')
                   #print(f'largo array inicial: {len(np_array)}')
        #try:       #        
                 if (len(np_array) == 8402):
                       if (np_array[0] == 11):
                           global modamaximovoltaje11
                           global modamaximocorriente11
                           global vrms1
                           global vrms11
                           global irms1
                           global irms11
                           global modavoltaje
                           global modacorriente
                           global samplings1
                           global NoVoltageoffset1
                           global NoCurrentoffset1
                           global ListaIrmsPeak1
                           global potrmsCGE
                           
                           samplings1 = np_array[-1]
                           list_FPVoltage3 = (np_array[0:4200]) #/1.65
                           list_FPCurrent3 = np_array[4201:8400]
                           sos = signal.butter(10, 2500, 'low', fs=samplings1, output='sos')
                           list_FPVoltage2 = signal.sosfilt(sos, list_FPVoltage3)
                           list_FPCurrent2 = signal.sosfilt(sos, list_FPCurrent3)
                           list_FPVoltage = list_FPVoltage2[104:4200]
                           list_FPCurrent = list_FPCurrent2 [103:4200]

                           #Valor dc de Voltaje
                           valoresmaximovoltajesinmedia=getMaxValues(list_FPVoltage, 50)
                           valoresminimovoltajesinmedia=getMinValues(list_FPVoltage, 50)
                           maximovoltaje = np.median(valoresmaximovoltajesinmedia)
                           minimovoltaje = np.median(valoresminimovoltajesinmedia)
                           mediadcvoltaje = (maximovoltaje+minimovoltaje)/2
                           NoVoltageoffset1=(list_FPVoltage-mediadcvoltaje)
                           
                           vrms1=VoltajeRms(NoVoltageoffset1)*0.94
                          
                           if (len(modamaximovoltaje11)>=5):
                               modavoltaje=np.median(modamaximovoltaje11)
                               modavoltaje1=modavoltaje
                               vrms1=VoltRms(modavoltaje1)
                               #vrms1=modavoltaje
                               #print(f'Vrms CGE: {vrms1}')
                               str_num = {"value":vrms1,"save":1}
                               vrms11 = json.dumps(str_num)
                               VoltageFFT(NoVoltageoffset1,samplings1,1)
                               modamaximovoltaje11=[]
                           else:
                               modamaximovoltaje11.append(vrms1)
                        

                           #Valor dc de corriente
                           valoresmaxcorriente=getMaxValues(list_FPCurrent, 50)
                           valoresmincorriente=getMinValues(list_FPCurrent, 50)
                           maximocorriente = np.median(valoresmaxcorriente)
                           minimocorriente = np.median(valoresmincorriente)
       
                           mediadccorrienteCGE = (maximocorriente+minimocorriente)/2
                           #print("mediadccorriente : ",mediadccorrienteCGE)
                           # Valores maximo y minimos de corriente
                           NoCurrentoffset1=list_FPCurrent-mediadccorrienteCGE
                           maximocorriente2sinmedia=getMaxValues(NoCurrentoffset1, 50)
                           maximocorriente2 = np.median(maximocorriente2sinmedia)
                           irms1=CorrienteRms(NoCurrentoffset1)
                           

                           if (len(modamaximocorriente11)>=5):
                               modacorriente=np.median(modamaximocorriente11)
                               irms1=CurrentRms(modacorriente)
                               #print(f'Irms CGE: {irms1}')
                               str_num = {"value":irms1,"save":1}
                               irms11 = json.dumps(str_num)
                               proporción=maximocorriente2/(irms1*np.sqrt(2))
                               ListaIrmsPeak1 = NoCurrentoffset1/proporción
                               maximocorr=getMaxValues(ListaIrmsPeak1, 10)
                               maximocorrCGE = np.median(maximocorr)
                               CurrentFFT(ListaIrmsPeak1,samplings1,1,irms1)
                               potrmsCGE = PotenciaRms(ListaIrmsPeak1,NoVoltageoffset1)
                               Potencias(1,irms1,vrms1)
                               ExcelAllInsertCGE()
                               ExcelDataCGE()
                               try:
                                    Maximo15minCGE()
                               except OSError as err:
                                    print("OS error: {0}".format(err))
                                    continue
                               except ValueError:
                                    print("Could not convert data to an integer.")
                                    continue
                               #except BaseException as err:
                               #     print(f"Unexpected {err=}, {type(err)=}")
                               #     raise
                                    #continue
                               modamaximocorriente11=[]
                               #CalculoDesfase(list_FPVoltage,list_FPCurrent,samplings1)
                           else:
                               modamaximocorriente11.append(irms1)
                           #    print(f'array corriente: {modamaximocorriente2}')

                       if (np_array[0] == 22):
                           global modamaximovoltaje22
                           global modamaximocorriente22
                           global vrms2
                           global vrms22
                           global irms2
                           global irms22
                           global modavoltaje22
                           global modacorriente22
                           global samplings2
                           global NoVoltageoffset2
                           global NoCurrentoffset2
                           global ListaIrmsPeak2
                           global potrmsCarga
                           
                           samplings2 = np_array[-1]
                           list_FPVoltage3 = np_array[0:4200]
                           list_FPCurrent3 = np_array[4201:8400]
                           sos = signal.butter(10, 2500, 'low', fs=samplings2, output='sos')
                           list_FPVoltage2 = signal.sosfilt(sos, list_FPVoltage3)
                           list_FPCurrent2 = signal.sosfilt(sos, list_FPCurrent3)
                           list_FPVoltage = list_FPVoltage2[104:4200]
                           list_FPCurrent = list_FPCurrent2 [103:4200]

                           #Valor dc de Voltaje
                           valoresmaximovoltajesinmedia=getMaxValues(list_FPVoltage, 50)
                           valoresminimovoltajesinmedia=getMinValues(list_FPVoltage, 50)
                           maximovoltaje = np.median(valoresmaximovoltajesinmedia)
                           minimovoltaje = np.median(valoresminimovoltajesinmedia)
                           mediadcvoltaje = (maximovoltaje+minimovoltaje)/2
                           # Valores maximo y minimos de voltaje sin componente continua
                           NoVoltageoffset2=list_FPVoltage-mediadcvoltaje
                           vrms2=VoltajeRms(NoVoltageoffset2)*0.94
                           if (len(modamaximovoltaje22)>=5):
                               modavoltaje22=np.median(modamaximovoltaje22)
                               modavoltaje221=modavoltaje22
                               vrms2=VoltRms(modavoltaje221)
                               #print(f'moda voltaje carga: {vrms2}')
                               str_num = {"value":vrms2,"save":1}
                               vrms22 = json.dumps(str_num)
                               #print(f'Vrms Carga: {vrms2}')
                               VoltageFFT(NoVoltageoffset2,samplings2,2)
                               modamaximovoltaje22=[]
                           else:
                               modamaximovoltaje22.append(vrms2)
                            
                           #Valor dc de corriente
                           valoresmaxcorriente=getMaxValues(list_FPCurrent, 50)
                           valoresmincorriente=getMinValues(list_FPCurrent, 50)
                           maximocorriente = np.median(valoresmaxcorriente)
                           minimocorriente = np.median(valoresmincorriente)
       
                           mediadccorrienteCarga = (maximocorriente+minimocorriente)/2
                           #print("mediadccorriente 2: ",mediadccorrienteCarga)
                           # Valores maximo y minimos de corriente
                           NoCurrentoffset2=list_FPCurrent-mediadccorrienteCarga
                           maximocorriente2sinmedia=getMaxValues(NoCurrentoffset2, 50)
                           maximocorriente2 = np.median(maximocorriente2sinmedia)
                           irms2=CorrienteRms(NoCurrentoffset2)

                           if (len(modamaximocorriente22)>=5):
                               modacorriente22=np.median(modamaximocorriente22)
                               irms2=CurrentRms(modacorriente22)
                               str_num = {"value":irms2,"save":1}
                               irms22 = json.dumps(str_num)
                               #print(f'Irms Carga: {irms2}')
                
                               proporción=maximocorriente2/(irms2*np.sqrt(2))
                               ListaIrmsPeak2 = NoCurrentoffset2/proporción
                               maximocorr=getMaxValues(ListaIrmsPeak2, 10)
                               maximocorr = np.median(maximocorr)
                               CurrentFFT(ListaIrmsPeak2,samplings2,2,irms2)
                               potrmsCarga=PotenciaRms(ListaIrmsPeak2,NoVoltageoffset2)
                               Potencias(2,irms2,vrms2)
                               ExcelAllInsertCarga()
                               ExcelDataCarga()
                               try:
                                     Maximo15minCarga()
                               except OSError as err:
                                    print("OS error: {0}".format(err))
                                    continue
                               except ValueError:
                                    print("Could not convert data to an integer.")
                                    continue
                               modamaximocorriente22=[]
                           else:
                               modamaximocorriente22.append(irms2)

                       if (np_array[0] == 33):
                           global modamaximovoltaje33
                           global modamaximocorriente33
                           global vrms3
                           global vrms33
                           global irms3
                           global irms33
                           global modavoltaje33
                           global modacorriente33
                           global samplings3
                           global NoVoltageoffset3
                           global NoCurrentoffset3
                           global ListaIrmsPeak3
                           global potrmsPaneles
                           samplings3 = np_array[-1]
                           list_FPVoltage3 = np_array[0:4200]
                           list_FPCurrent3 = np_array[4201:8400]
                           sos = signal.butter(10, 2500, 'low', fs=samplings3, output='sos')
                           list_FPVoltage2 = signal.sosfilt(sos, list_FPVoltage3)
                           list_FPCurrent2 = signal.sosfilt(sos, list_FPCurrent3)
                           list_FPVoltage = list_FPVoltage2[104:4200]
                           list_FPCurrent = list_FPCurrent2 [103:4200]

                           #Valor dc de Voltaje
                           valoresmaximovoltajesinmedia=getMaxValues(list_FPVoltage, 50)
                           valoresminimovoltajesinmedia=getMinValues(list_FPVoltage, 50)
                           maximovoltaje = np.median(valoresmaximovoltajesinmedia)
                           minimovoltaje = np.median(valoresminimovoltajesinmedia)
                           mediadcvoltaje = (maximovoltaje+minimovoltaje)/2
                           # Valores maximo y minimos de voltaje sin componente continua
                           NoVoltageoffset3=list_FPVoltage-mediadcvoltaje  #Señal Voltaje 
                           
                           vrms3=VoltajeRms(NoVoltageoffset3)*0.94
                           
                           if (len(modamaximovoltaje33)>=5):
                               modavoltaje33=np.median(modamaximovoltaje33)
                               vrms3=VoltRms(modavoltaje33)
                               str_num = {"value":vrms3,"save":1}
                               vrms33 = json.dumps(str_num)
                               #print(f'Vrms Paneles: {vrms3}')
                               VoltageFFT(NoVoltageoffset3,samplings3,3)
                               modamaximovoltaje33=[]
                           else:
                               modamaximovoltaje33.append(vrms3)
                            
                           #Valor dc de corriente
                           valoresmaxcorriente=getMaxValues(list_FPCurrent, 50)
                           valoresmincorriente=getMinValues(list_FPCurrent, 50)
                           maximocorriente = np.median(valoresmaxcorriente)
                           minimocorriente = np.median(valoresmincorriente)
       
                           mediadccorriente = (maximocorriente+minimocorriente)/2
                           
                           # Valores maximo y minimos de corriente
                           NoCurrentoffset3=list_FPCurrent-mediadccorriente
                           maximocorriente2sinmedia=getMaxValues(NoCurrentoffset3, 50)
                           maximocorriente2 = np.median(maximocorriente2sinmedia)
                           irms3=CorrienteRms(NoCurrentoffset3)
                           

                           if (len(modamaximocorriente33)>=5):
                               modacorriente33=np.median(modamaximocorriente33)
                               irms3=CurrentRms(modacorriente33)
                               str_num = {"value":irms3,"save":1}
                               irms33 = json.dumps(str_num)
                               #print(f'Irms Paneles : {irms3}')
                               proporción=maximocorriente2/(irms3*np.sqrt(2))
                               ListaIrmsPeak3 = NoCurrentoffset3/proporción
                               maximocorr=getMaxValues(ListaIrmsPeak3, 10)
                               maximocorr = np.median(maximocorr)
                               CurrentFFT(ListaIrmsPeak3,samplings3,3,irms3)
                               potrmsPaneles=PotenciaRms(ListaIrmsPeak3,NoVoltageoffset3)
                               Potencias(3,irms3,vrms3)
                               ExcelAllInsertPaneles()
                               ExcelDataPaneles()
                               try:
                                    Maximo15minPaneles()
                               except OSError as err:
                                    print("OS error: {0}".format(err))
                                    continue
                               except ValueError:
                                    print("Could not convert data to an integer.")
                               modamaximocorriente33=[]
                           else:
                               modamaximocorriente33.append(irms3)

                       if (np_array[0] == 44):
                           global modamaximovoltajeBateria
                           global modamaximocorrientebateria
                           global modavoltajeBateria
                           global modacorrienteBateria
                           global NoVoltageoffsetBateria
                           global energyBateriaSend
                           global energyBateriaHora
                           global tiempo2Bateria
                           global PotenciaBaterias
                           global VoltajeBaterias
                           global CorrienteBaterias
                           global PotenciaBateriaSend
                           global VoltajeBateriaSend
                           global CorrienteBateriaSend
                           global energyBaterias
                           samplings = np_array[-1]
                           #print(f'samplings: {samplings}')
                           list_FPVoltage3 = np_array[0:4200]
                           list_FPCurrent3 = np_array[4201:8400]
                           sos = signal.butter(10, 500, 'low', fs=samplings, output='sos')
                           list_FPVoltage2 = signal.sosfilt(sos, list_FPVoltage3)
                           list_FPCurrent2 = signal.sosfilt(sos, list_FPCurrent3)
                           list_FPVoltage = list_FPVoltage2[104:4200]
                           list_FPCurrent = list_FPCurrent2 [103:4200]

                           try:
                               mediapotenciometro=(mediadccorrienteCGE+mediadccorrienteCarga)/2
                           except:
                               continue
                           # Valores maximo y minimos de voltaje sin componente continua
                           NoVoltageoffsetBateria=list_FPVoltage-mediapotenciometro
                           MediaVoltageBaterias=np.median(NoVoltageoffsetBateria)
                           #print(f'Voltaje Moda Baterias1: {MediaVoltageBaterias}')
                           if (len(modamaximovoltajeBateria)>=5):
                               modavoltajeBateria=np.median(modamaximovoltajeBateria)
                               print(f'Voltaje Moda Baterias: {modavoltajeBateria}')
                               VoltajeBaterias=modavoltajeBateria/27
                               print(f'Voltaje Baterias: {VoltajeBaterias}')
                               str_num = {"value":VoltajeBaterias,"save":1}
                               VoltajeBateriaSend = json.dumps(str_num)
                               modamaximovoltajeBateria=[]
                           else:
                               modamaximovoltajeBateria.append(MediaVoltageBaterias)
                        
                           #Valor dc de corriente Baterias
                           valoresmaxcorrienteBateria=getMaxValues(list_FPCurrent, 50)
                           valoresmincorrienteBateria=getMinValues(list_FPCurrent, 50)
                           maximocorrientebateria = np.median(valoresmaxcorrienteBateria)
                           minimocorrientebateria = np.median(valoresmincorrienteBateria)
       
                           mediadccorrientebateria = (maximocorrientebateria+minimocorrientebateria)/2
                           
                           # Valores maximo y minimos de corriente
                           NoCurrentoffsetCorrienteBaterias=list_FPCurrent-mediadccorrientebateria
                           mediaCorrienteBaterias = np.median(NoCurrentoffsetCorrienteBaterias)
                           #print(f'Corriente Moda Baterias1: {mediaCorrienteBaterias}')
                           if (len(modamaximocorrientebateria)>=5):
                               modacorrientebateria=np.median(modamaximocorrientebateria)
                               print(f'Corriente Moda Baterias: {modacorrientebateria}')
                               CorrienteBaterias = modacorrientebateria/475
                               print(f'Corriente Baterias: {CorrienteBaterias}')
                               str_num = {"value":CorrienteBaterias,"save":1}
                               CorrienteBateriaSend = json.dumps(str_num)
                               PotenciaBaterias=(modacorrientebateria/475)*(modavoltajeBateria/27)
                               str_num = {"value":PotenciaBaterias,"save":1}
                               PotenciaBateriaSend = json.dumps(str_num)
                               tiempo1Bateria = datetime.datetime.now()
                               delta=(((tiempo1Bateria - tiempo2Bateria).microseconds)/1000+((tiempo1Bateria - tiempo2Bateria).seconds)*1000)/10000000000
                               energyBaterias += np.abs(PotenciaBaterias*delta*2.9)
                               str_num = {"value":energyBaterias,"save":1}
                               energyBateriaSend = json.dumps(str_num)
                               energyBateriaHora += np.abs(PotenciaBaterias*delta*2.9)
                               print(f'Energia Baterias: {energyBaterias}')
                               tiempo2Bateria = datetime.datetime.now()
                               if(tiempo2Bateria.minute==0):
                                    energyBateriaHora=0
                               if(tiempo2Bateria.hour==0 and tiempo2Bateria.minute==0):
                                    energyBaterias=0
                               if(tiempo2Bateria.hour==0 and tiempo2Bateria.minute==1):
                                    energyBaterias=0
                               ExcelAllInsertBaterias()
                               ExcelDataBaterias()
                               try:
                                    Maximo15minBateriasDC()
                               except OSError as err:
                                    print("OS error: {0}".format(err))
                                    continue
                               except ValueError:
                                    print("Error en Baterias")
                               modamaximocorrientebateria=[]
                           else:
                               modamaximocorrientebateria.append(mediaCorrienteBaterias)

                              
                       if (np_array[0] == 55):
                           global modamaximovoltajePaneles
                           global modamaximocorrientePaneles
                           global modavoltajePaneles
                           global modacorrientePaneles
                           global NoVoltageoffsetPaneles
                           global energyPanelesDC
                           global energyPanelesDCSend
                           global energyPanelesHoraDC
                           global tiempo2Paneles
                           global VoltajePanelesDC
                           global VoltajePanelesDCSend
                           global CorrientePanelesDC
                           global PotenciaPanelesDC
                           global CorrientePanelesDCSend
                           global PotenciaPanelesDCSend
                           samplings = np_array[-1]
                           #print(f'samplings Paneles: {samplings}')
                           list_FPCurrent3 = np_array[0:4200]
                           list_FPVoltage3 = np_array[4201:8400]
                           sos = signal.butter(10, 500, 'low', fs=samplings, output='sos')
                           list_FPVoltage2 = signal.sosfilt(sos, list_FPVoltage3)
                           list_FPCurrent2 = signal.sosfilt(sos, list_FPCurrent3)
                           list_FPVoltage = list_FPVoltage2[104:4200]
                           list_FPCurrent = list_FPCurrent2 [103:4200]

                           try:
                               mediapotenciometro=(mediadccorrienteCGE+mediadccorrienteCarga)/2
                               NoCorrienteoffsetPaneles=list_FPCurrent-mediapotenciometro
                               MediaCorrientePaneles=np.median(NoCorrienteoffsetPaneles)
                               #print(f'Voltaje Moda Baterias1: {MediaVoltageBaterias}')
                               if (len(modamaximocorrientePaneles)>=5):
                                   modacorrientePaneles=np.median(modamaximocorrientePaneles)
                                   print(f'Corriente Moda Paneles Directa: {modacorrientePaneles}')
                                   CorrientePanelesDC=modacorrientePaneles/85
                                   print(f'Corriente Paneles Directa: {CorrientePanelesDC}')
                                   str_num = {"value":CorrientePanelesDC,"save":1}
                                   CorrientePanelesDCSend = json.dumps(str_num)
                                   modamaximocorrientePaneles=[]
                               else:
                                   modamaximocorrientePaneles.append(MediaCorrientePaneles)
                           except:
                               continue
                           
                           try:
                               mediapotenciometro=(mediadccorrienteCGE+mediadccorrienteCarga)/2
                               NoVoltageoffsetPaneles=list_FPVoltage-mediapotenciometro
                               MediaVoltagePaneles=np.median(NoVoltageoffsetPaneles)
                               #print(f'Voltaje Moda Baterias1: {MediaVoltageBaterias}')
                               if (len(modamaximovoltajePaneles)>=5):
                                   modavoltajePaneles=np.median(modamaximovoltajePaneles)
                                   print(f'Voltaje Moda Paneles Directa: {modavoltajePaneles}')
                                   VoltajePanelesDC=modavoltajePaneles/4.97
                                   print(f'Voltaje Paneles Directa: {VoltajePanelesDC}')
                                   str_num = {"value":VoltajePanelesDC,"save":1}
                                   VoltajePanelesDCSend = json.dumps(str_num)
                                   PotenciaPanelesDC=(CorrientePanelesDC)*(VoltajePanelesDC)
                                   str_num = {"value":PotenciaPanelesDC,"save":1}
                                   PotenciaPanelesDCSend = json.dumps(str_num)
                                   tiempo1Paneles = datetime.datetime.now()
                                   delta=(((tiempo1Paneles - tiempo2Paneles).microseconds)/1000+((tiempo1Paneles - tiempo2Paneles).seconds)*1000)/10000000000
                                   energyPanelesDC += np.abs(PotenciaPanelesDC*delta*2.9)
                                   str_num = {"value":energyPanelesDC,"save":1}
                                   energyPanelesDCSend = json.dumps(str_num)
                                   energyPanelesHoraDC += np.abs(PotenciaPanelesDC*delta*2.9)
                                   print(f'Energia Paneles: {energyPanelesDC}')
                                   tiempo2Paneles = datetime.datetime.now()
                                   if(tiempo2Paneles.minute==0):
                                        energyPanelesHoraDC=0
                                   if(tiempo2Paneles.hour==0 and tiempo2Paneles.minute==0):
                                        energyPanelesDC=0
                                   if(tiempo2Paneles.hour==0 and tiempo2Paneles.minute==1):
                                        energyPanelesDC=0
                                   ExcelAllInsertPanelesDC()
                                   ExcelDataPanelesDirecta()
                                   try:
                                       Maximo15minPanelesDC()
                                   except OSError as err:
                                       print("OS error: {0}".format(err))
                                       continue
                                   except ValueError:
                                       print("Error en Paneles DC")  
                                   modamaximovoltajePaneles=[]
                               else:
                                   modamaximovoltajePaneles.append(MediaVoltagePaneles)
                           except:
                               continue
                           
                 
                 if (len(np_array)>0 and len(np_array)<=2):
                         global tempESP32
                         global Temp_Raspberry
                         global Temp_Raspberry0
                         global cpu_uso
                         global RAM
                         global RAM1
                         global reinicio
                         #global EstateVentilador
                         Temp_Raspberry0=cpu_temp()
                         cpu_uso=get_cpuload()
                         
                         str_num = {"value":Temp_Raspberry0,"save":0}
                         Temp_Raspberry = json.dumps(str_num)
                         Ventilador()
                         RAM = psutil.virtual_memory()[2]
                         #print(f'RAM memory 2:  {RAM}%') 
                         dataAllVariables()
                         VariablesExcel()
                         if (RAM > 93):
                              os.system("sudo reboot")
                         #temphum()
                         #distance()
                         tempESP320 = round(np_array[0],0)
                         str_num2 = {"value":tempESP320,"save":0}
                         tempESP32 = json.dumps(str_num2)
                         

                 excel=datetime.datetime.now()
                 
                 #display.show()
                 """
                 if(excel.hour==13 and excel.minute==3):
                          if(accesoemail2==0):
                                 accesoemail2=1
                                 print("Entro a SendEmail")
                                 SendEmail()
                 else:
                     accesoemail2=0
                 """
                
                 if(excel.hour==0 and excel.minute==3):
                          if(accesoemail3==0):
                                 accesoemail3=1
                                 print("Entro a SendEmail")
                                 #SendEmail()
                                 time.sleep(5)
                                 #os.remove(dest_filename)
                                 excelcreate()
                 else:
                     accesoemail3=0
                 try:
                       if(excel.minute==1 or excel.minute==16 or excel.minute==31 or excel.minute==46):
                            if(accesoexcel==0): 
                                   ExcelDataCGE15()
                                   ExcelDataCarga15()
                                   ExcelDataPaneles15()
                                   ExcelDataBaterias15()
                                   ExcelDataPanelesDirecta15()
                                   accesoexcel=1
                       else:
                               accesoexcel=0
                 except:
                     continue

                 try:  
                       if(client.connected_flag==True): 
                             publish(client)
                 except:
                     global countbroker
                     print(f'Count Broker: {countbroker}')
                     if(countbroker>=71):
                         reconnectmqtt()
                         countbroker=0
                     else: 
                         countbroker=countbroker+1
                         
                     continue
        #except:
        #    print("Error en Bucle")
        #    continue


b1=time.time()
c1=time.time()
d1=time.time()
e21=time.time()
f1=time.time()
g1=time.time()
h1=time.time()
j1=time.time()
k1=time.time()
z1=time.time()
l1=time.time()
m1=time.time()
n1=time.time()
o21=time.time()
p1=time.time()
q1=time.time()
r1=time.time()
s1=time.time()
t1=time.time()
u1=time.time()
v1=time.time()
v12=time.time()
v13=time.time()
v14=time.time()
v15=time.time()
v16=time.time()
v17=time.time()
w1=time.time()
x1=time.time()
y1=time.time()
#varsLastSend=[b,c,d,e]

def publish(client): 
        global b1, c1 ,d1, e21, f1 ,g1 ,h1 , j1, k1, l1, m1, n1, o21 ,p1, q1, r1, s1, t1, u1, v1 
        global v12, v13, v14, v15, v16, v17,v18,v19,v20,v21,v22,v23, w1, x1, y1, z1
        a1=time.time()
        for i in data["variables"]:

            #    if(data["variables"][i]["variableType"]=="output"):
            #        continue
            if(i["variableFullName"]=="Corriente-CGE"):
                freq = i["variableSendFreq"]
                if(a1 - b1 > float(freq)):
                     b1=time.time()
                     str_variable = i["variable"]
                     topic1 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic1, irms11)
                     status = result[0]            
                     if status == 0:
                         print(f"Send irms1: `{irms11}` to topic `{topic1}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic1}")
        
                   
            if(i["variableFullName"]=="Voltaje-CGE"):
                freq = i["variableSendFreq"]
                if(a1 - c1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     c1=time.time()
                     str_variable2 = i["variable"]
                     topic2 = topicmqtt + str_variable2 + "/sdata"
                     result = client.publish(topic2, vrms11)
                #     status = result[0]
                #     if status == 0:
                #         print(f"Send vrms1: `{vrms11}` to topic `{topic2}` con freq: {freq}")
                #     else:
                #         print(f"Failed to send message to topic {topic2}")
            """
            if(i["variableFullName"]=="Potencia-Reactiva-CGE"):
                freq = i["variableSendFreq"]
                if(a1 - d1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     d1=time.time()
                     str_variable3 = i["variable"]
                     topic3 = topicmqtt + str_variable3 + "/sdata"
                     result = client.publish(topic3, ReactivaCGEFase1)
                     status = result[0]
                     if status == 0:
                         print(f"Send Pot-Reactiva-CGE: `{ReactivaCGEFase1}` to topic `{topic3}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic3}")
            """
            if(i["variableFullName"]=="Pot-Activa-CGE"):
                freq = i["variableSendFreq"]
                if(a1 - e21 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     e21=time.time()
                     str_variable4 = i["variable"]
                     topic4 = topicmqtt + str_variable4 + "/sdata"
                     result = client.publish(topic4, ActivaCGEFase1)
               #      status = result[0]
               #      if status == 0:
               #          print(f"Send Pot-Activa CGE: `{ActivaCGEFase1}` to topic `{topic4}` con freq: {freq}")
               #      else:
               #          print(f"Failed to send message to topic {topic4}")
            
            if(i["variableFullName"]=="Energia-CGE"):
                freq = i["variableSendFreq"]
                if(a1 - f1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     f1=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, energyCGEFase1)
               #      status = result[0]
               #      if status == 0:
               #          print(f"Send energia CGE: `{energyCGEFase1}` to topic `{topic5}` con freq: {freq}")
               #      else:
               #          print(f"Failed to send message to topic {topic5}")
            
            if(i["variableFullName"]=="FP-CGE"):
                freq = i["variableSendFreq"]
                if(a1 - g1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     g1=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, FPCGE)
               #     status = result[0]
               #     if status == 0:
               #         print(f"Send FP-CGE: `{FPCGE}` to topic `{topic5}` con freq: {freq}")
               #     else:
               #         print(f"Failed to send message to topic {topic5}")
            """
            if(i["variableFullName"]=="FD-CGE"):
                freq = i["variableSendFreq"]
                if(a1 - h1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     h1=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, FDCorrienteCGE)
                     status = result[0]
                     if status == 0:
                         print(f"Send FD-CGE: `{FDCorrienteCGE}` to topic `{topic5}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic5}")
            if(i["variableFullName"]=="DAT-CGE"):
                freq = i["variableSendFreq"]
                if(a1 - j1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     j1=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, DATCorrienteCGE)
                     status = result[0]
                     if status == 0:
                         print(f"Send DAT-CGE: `{DATCorrienteCGE}` to topic `{topic5}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic5}")
            """
            if(i["variableFullName"]=="Pot-CGE"):
                freq = i["variableSendFreq"]
                if(a1 - k1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     k1=time.time()
                     str_variable = i["variable"]
                     topic5= topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, AparenteCGEFase1)
                     status = result[0]
                     if status == 0:
                         print(f"Send Pot-Aparente-CGE : `{AparenteCGEFase1}` to topic `{topic5}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic5}")
            
            #SEGUNDA TOMA
            if(i["variableFullName"]=="Corriente-Carga"):
                freq = i["variableSendFreq"]
                if(a1 - l1 > float(freq)):
                     l1=time.time()
                     str_variable = i["variable"]
                     topic = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic, irms22)
              #       status = result[0]
              #       
              #       if status == 0:
              #           print(f"Send Corriente-Carga: `{irms22}` to topic `{topic}` con freq: {freq}")
              #       else:
              #           print(f"Failed to send message to topic {topic}")
        
                   
            if(i["variableFullName"]=="Voltaje-Carga"):
                freq = i["variableSendFreq"]
                if(a1 - m1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     m1=time.time()
                     str_variable2 = i["variable"]
                     topic2 = topicmqtt + str_variable2 + "/sdata"
                     result = client.publish(topic2, vrms22)
              #       status = result[0]
              #       if status == 0:
              #           print(f"Send Voltaje-Carga: `{vrms22}` to topic `{topic2}` con freq: {freq}")
              #       else:
              #           print(f"Failed to send message to topic {topic2}")
            """
            if(i["variableFullName"]=="Potencia-Reactiva-Carga"):
                freq = i["variableSendFreq"]
                if(a1 - n1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     n1=time.time()
                     str_variable3 = i["variable"]
                     topic3 = topicmqtt + str_variable3 + "/sdata"
                     result = client.publish(topic3, ReactivaCargaFase1)
                     status = result[0]
                     if status == 0:
                         print(f"Send Potencia-Reactiva-Carga: `{ReactivaCargaFase1}` to topic `{topic3}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic3}")
            """
            if(i["variableFullName"]=="Pot-Activa-Carga"):
                freq = i["variableSendFreq"]
                if(a1 - o21 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     o21=time.time()
                     str_variable4 = i["variable"]
                     topic4 = topicmqtt + str_variable4 + "/sdata"
                     result = client.publish(topic4, ActivaCargaFase1)
               #      status = result[0]
               #      if status == 0:
               #          print(f"Send Pot-Activa-Carga: `{ActivaCargaFase1}` to topic `{topic4}` con freq: {freq}")
               #      else:
               #          print(f"Failed to send message to topic {topic4}")
            
            if(i["variableFullName"]=="Energia-Carga"):
                freq = i["variableSendFreq"]
                if(a1 - p1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     p1=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, energyCargaFase1)
                 #    status = result[0]
                 #    if status == 0:
                 #        print(f"Send Energia-Carga: `{energyCargaFase1}` to topic `{topic5}` con freq: {freq}")
                 #    else:
                 #        print(f"Failed to send message to topic {topic5}")
            
            if(i["variableFullName"]=="FP-Carga"):
                freq = i["variableSendFreq"]
                if(a1 - q1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     q1=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, FPCarga)
               #      status = result[0]
               #      if status == 0:
               #          print(f"Send FP-Carga: `{FPCarga}` to topic `{topic5}` con freq: {freq}")
               #      else:
               #          print(f"Failed to send message to topic {topic5}")
            """
            if(i["variableFullName"]=="FD-Carga"):
                freq = i["variableSendFreq"]
                if(a1 - r1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     r1=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, FDCorrienteCarga)
                     status = result[0]
                     if status == 0:
                         print(f"Send FD-Carga: `{FDCorrienteCarga}` to topic `{topic5}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic5}")
            if(i["variableFullName"]=="DAT-Carga"):
                freq = i["variableSendFreq"]
                if(a1 - s1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     s1=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, DATCorrienteCarga)
                     status = result[0]
                     if status == 0:
                         print(f"Send DAT-Carga: `{DATCorrienteCarga}` to topic `{topic5}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic5}")
            
            """
            if(i["variableFullName"]=="Pot-Carga"):
                freq = i["variableSendFreq"]
                if(a1 - t1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     t1=time.time()
                     str_variable = i["variable"]
                     topic5= topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, AparenteCargaFase1)
                     status = result[0]
              #       if status == 0:
               #          print(f"Send Pot-Aparente-Carga: `{AparenteCargaFase1}` to topic `{topic5}` con freq: {freq}")
                #     else:
                 #        print(f"Failed to send message to topic {topic5}")
            
            #Tercera Toma
            if(i["variableFullName"]=="Corriente-Paneles"):
                freq = i["variableSendFreq"]
                if(a1 - u1 > float(freq)):
                     u1=time.time()
                     str_variable = i["variable"]
                     topic = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic, irms33)
               #      status = result[0]
               #      
               #      if status == 0:
               #          print(f"Send Corriente-Paneles: `{irms33}` to topic `{topic}` con freq: {freq}")
               #      else:
               #          print(f"Failed to send message to topic {topic}")
        
                   
            if(i["variableFullName"]=="Voltaje-Paneles"):
                freq = i["variableSendFreq"]
                if(a1 - v1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v1=time.time()
                     str_variable2 = i["variable"]
                     topic2 = topicmqtt + str_variable2 + "/sdata"
                     result = client.publish(topic2, vrms33)
                #     status = result[0]
                #     if status == 0:
                #         print(f"Send Voltaje-Paneles: `{vrms33}` to topic `{topic2}` con freq: {freq}")
                #     else:
                #         print(f"Failed to send message to topic {topic2}")
            """
            if(i["variableFullName"]=="Potencia-Reactiva-Paneles"):
                freq = i["variableSendFreq"]
                if(a1 - v12 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v12=time.time()
                     str_variable3 = i["variable"]
                     topic3 = topicmqtt + str_variable3 + "/sdata"
                     result = client.publish(topic3, ReactivaPanelesFase1)
                     status = result[0]
                     if status == 0:
                         print(f"Send Potencia-Reactiva-Paneles: `{ReactivaPanelesFase1}` to topic `{topic3}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic3}")
            """
            if(i["variableFullName"]=="Pot-Activa-Paneles"):
                freq = i["variableSendFreq"]
                if(a1 - v13 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v13=time.time()
                     str_variable4 = i["variable"]
                     topic4 = topicmqtt + str_variable4 + "/sdata"
                     result = client.publish(topic4, ActivaPanelesFase1)
                 #    status = result[0]
                 #    if status == 0:
                 #        print(f"Send Pot-Activa-Paneles: `{ActivaPanelesFase1}` to topic `{topic4}` con freq: {freq}")
                 #    else:
                 #        print(f"Failed to send message to topic {topic4}")
            
            if(i["variableFullName"]=="Energia-Paneles"):
                freq = i["variableSendFreq"]
                if(a1 - v14 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v14=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, energyPanelesFase1)
                  #   status = result[0]
                  #   if status == 0:
                  #       print(f"Send Energia-Paneles: `{energyPanelesFase1}` to topic `{topic5}` con freq: {freq}")
                  #   else:
                  #       print(f"Failed to send message to topic {topic5}")
            
            if(i["variableFullName"]=="FP-Paneles"):
                freq = i["variableSendFreq"]
                if(a1 - v15 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v15=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, FPPaneles)
                  #   status = result[0]
                  #   if status == 0:
                  #       print(f"Send FP-Paneles: `{FPPaneles}` to topic `{topic5}` con freq: {freq}")
                  #   else:
                  #       print(f"Failed to send message to topic {topic5}")
            """
            if(i["variableFullName"]=="FD-Paneles"):
                freq = i["variableSendFreq"]
                if(a1 - v16 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v16=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, FDCorrientePaneles)
                     status = result[0]
                     if status == 0:
                         print(f"Send FD-Paneles: `{FDCorrientePaneles}` to topic `{topic5}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic5}")
            if(i["variableFullName"]=="DAT-Paneles"):
                freq = i["variableSendFreq"]
                if(a1 - v17 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v17=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, DATCorrientePaneles)
                     status = result[0]
                     if status == 0:
                         print(f"Send DAT-Paneles: `{DATCorrientePaneles}` to topic `{topic5}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic5}")
            if(i["variableFullName"]=="Pot-Aparente-Paneles"):
                freq = i["variableSendFreq"]
                if(a1 - w1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     w1=time.time()
                     str_variable = i["variable"]
                     topic5= topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, AparentePanelesFase1)
                     status = result[0]
                     if status == 0:
                         print(f"Send Pot-Aparente-Paneles: `{AparentePanelesFase1}` to topic `{topic5}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic5}")
            
            """
            if(i["variableFullName"]=="Voltaje-Baterias"):
                freq = i["variableSendFreq"]
                if(a1 - v16 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v16=time.time()
                     str_variable4 = i["variable"]
                     topic4 = topicmqtt + str_variable4 + "/sdata"
                     result = client.publish(topic4, VoltajeBateriaSend)  
                     status = result[0]
                     if status == 0:
                         print(f"Send Voltaje-Baterias: `{VoltajeBateriaSend}` to topic `{topic4}` ")
                     else:
                         print(f"Failed to send message to topic {topic4}")    
            if(i["variableFullName"]=="Corriente-Baterias"):
                freq = i["variableSendFreq"]
                if(a1 - v17 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v17=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, CorrienteBateriaSend)     
            if(i["variableFullName"]=="Potencia-Baterias"):
                freq = i["variableSendFreq"]
                if(a1 - v18 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v18=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, PotenciaBateriaSend)
            if(i["variableFullName"]=="Energia-Baterias"):
                freq = i["variableSendFreq"]
                if(a1 - v19 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v19=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, energyBateriaSend)

            if(i["variableFullName"]=="Voltaje-PanelesDC"):
                freq = i["variableSendFreq"]
                if(a1 - v20 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v20=time.time()
                     str_variable4 = i["variable"]
                     topic4 = topicmqtt + str_variable4 + "/sdata"
                     result = client.publish(topic4, VoltajePanelesDCSend)

            if(i["variableFullName"]=="Corriente-PanelesDC"):
                freq = i["variableSendFreq"]
                if(a1 - v21 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v21=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, CorrientePanelesDCSend)    

            if(i["variableFullName"]=="Potencia-PanelesDC"):
                freq = i["variableSendFreq"]
                if(a1 - v22 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v22=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, PotenciaPanelesDCSend)
            if(i["variableFullName"]=="Energia-PanelesDC"):
                freq = i["variableSendFreq"]
                if(a1 - v23 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     v23=time.time()
                     str_variable = i["variable"]
                     topic5 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic5, energyPanelesDCSend)

            if(i["variableFullName"]=="Temperatura-ESP32"):
                freq = i["variableSendFreq"]
                if(a1 - x1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     x1=time.time()
                     str_variable = i["variable"]
                     topic= topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic, tempESP32)
                     status = result[0]
                     #if status == 0:
                         #print(f"Send Temperatura-ESP32: `{tempESP32}` to topic `{topic}` con freq: {freq}")
                     #else:
                       #    print(f"Failed to send message to topic {topic}")

            if(i["variableFullName"]=="Ventilador-Raspberry"):
                freq = i["variableSendFreq"]
                if(a1 - y1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     y1=time.time()
                     str_variable = i["variable"]
                     topic= topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic, EstateVentilador)
                     status = result[0]
                     if status == 0:
                         print(f"Send Ventilador-Raspberry: `{EstateVentilador}` to topic `{topic}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic}")
            
            if(i["variableFullName"]=="Temperatura-Raspberry"):
                freq = i["variableSendFreq"]
                if(a1 - z1 > float(freq)):
                     #print("varlastsend 1: ",varsLastSend[i])
                     z1=time.time()
                     str_variable = i["variable"]
                     topic= topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic, Temp_Raspberry)
                     status = result[0]
                     if status == 0:
                         print(f"Send Temperatura-Raspberry: `{Temp_Raspberry}` to topic `{topic}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic}")
 

dataVariablesAll=[]
dataCGEAll=[]
dataCargaAll=[]
dataPanelesAll=[]
dataBateriasAll=[]
dataPanelesDirectaAll=[]

def dataAllVariables():
        #print(f'Guardando lista')
        #print("Temperatura Raspberry:  ",Temp_Raspberry0)
        #print("Uso de CPU Raspberry:  ",cpu_uso)
        dataVariablesAll.insert(1,Temp_Raspberry0)
        dataVariablesAll.insert(2,cpu_uso)
        dataVariablesAll.insert(3,RAM)
        #dataVariablesAll.insert(4,RAM)

def ExcelAllInsertCGE():
        dataCGEAll.insert(1,round(vrms1,2))
        dataCGEAll.insert(2,round(irms1,2))
        dataCGEAll.insert(3,round(ActivaCGEFase11,2))
        dataCGEAll.insert(4,round(ReactivaCGEFase11,2))
        dataCGEAll.insert(5,round(AparenteCGEFase11,2))
        dataCGEAll.insert(6,round(FPCGE0,2))
        dataCGEAll.insert(7,round(FDCorrienteCGE1,2))
        dataCGEAll.insert(8,round(DATCorrienteCGE1,2))
        dataCGEAll.insert(9,round(cosphiCGE,2))         
        dataCGEAll.insert(10,round(energyCGEFase11,2))
        dataCGEAll.insert(11,round(energyCGEFase11Hour,2))
        
def ExcelAllInsertCarga():        
        dataCargaAll.insert(1,round(vrms2,2))
        dataCargaAll.insert(2,round(irms2,2))
        dataCargaAll.insert(3,round(ActivaCargaFase13,2))
        dataCargaAll.insert(4,round(ReactivaCargaFase13,2))
        dataCargaAll.insert(5,round(AparenteCargaFase13,2))
        dataCargaAll.insert(6,round(FPCarga1,2))
        dataCargaAll.insert(7,round(FDCorrienteCarga1,2))
        dataCargaAll.insert(8,round(DATCorrienteCarga1,2))
        dataCargaAll.insert(9,round(cosphiCarga,2))
        dataCargaAll.insert(10,round(energyCargaFase13,2))
        dataCargaAll.insert(11,round(energyCargaFase13Hour,2))
        
def ExcelAllInsertPaneles():        
        dataPanelesAll.insert(1,round(vrms3,2))
        dataPanelesAll.insert(2,round(irms3,2))
        dataPanelesAll.insert(3,round(ActivaPanelesFase12,2))
        dataPanelesAll.insert(4,round(ReactivaPanelesFase12,2))
        dataPanelesAll.insert(5,round(AparentePanelesFase12,2))
        dataPanelesAll.insert(6,round(FPPaneles1,2))
        dataPanelesAll.insert(7,round(FDCorrientePaneles1,2))
        dataPanelesAll.insert(8,round(DATCorrientePaneles1,2))
        dataPanelesAll.insert(9,round(cosphiPaneles,2))
        dataPanelesAll.insert(10,round(energyPanelesFase12,2))
        dataPanelesAll.insert(11,round(energyPanelesFase12Hour,2))


def ExcelAllInsertBaterias():        
        dataBateriasAll.insert(1,round(VoltajeBaterias,2))
        dataBateriasAll.insert(2,round(CorrienteBaterias,2))
        dataBateriasAll.insert(3,round(PotenciaBaterias,2))
        dataBateriasAll.insert(4,round(energyBaterias,2))
        dataBateriasAll.insert(5,round(energyBateriaHora,2))
        
  
def ExcelAllInsertPanelesDC():        
        dataPanelesDirectaAll.insert(1,round(VoltajePanelesDC,2))
        dataPanelesDirectaAll.insert(2,round(CorrientePanelesDC,2))
        dataPanelesDirectaAll.insert(3,round(PotenciaPanelesDC,2))
        dataPanelesDirectaAll.insert(4,round(energyPanelesDC,2))
        dataPanelesDirectaAll.insert(5,round(energyPanelesHoraDC,2)) 
           
"""
'Voltaje', 'Corriente','Potencia Activa','Potencia Reactiva','Potencia Aparente',
'FP','FD','DAT',
"""

acceso = 0
maximoVoltaje15CGE=0
promedioVoltaje15CGE=0
minimoVoltaje15CGE=0
maximoCorrienteCGE=0
promedioCorrienteCGE=0
minimoCorrienteCGE=0
maximoPotActivaCGE=0
promedioPotActivaCGE=0
minimoPotActivaCGE=0
maximoPotReactivaCGE=0
promedioPotReactivaCGE=0
minimoPotReactivaCGE=0
maximoPotAparenteCGE=0
promedioPotAparenteCGE=0
minimoPotAparenteCGE=0
maximoFPCGEInductivo=-0.99
promedioFPCGEInductivo = -0.99
minimoFPCGEInductivo= -0.99
maximoFPCGEReactivo=0.99
promedioFPCGEReactivo=0.99
minimoFPCGEReactivo=0.99
maximoFDCGE=0
promedioFDCGE=0
minimoFDCGE=0
maximoDATCGE=0
promedioDATCGE=0
minimoDATCGE=0
Volt15CGE=[]
dataCGE=[]
Corriente15CGE=[]
PotActiva15CGE=[]
PotReactiva15CGE=[]
PotAparente15CGE=[]
FP15CGEReactivo=[]
FP15CGEInductivo=[]
FD15CGE=[]
DAT15CGE=[]
def Maximo15minCGE():
    global maximoVoltaje15CGE
    global promedioVoltaje15CGE
    global minimoVoltaje15CGE
    global maximoCorrienteCGE
    global promedioCorrienteCGE
    global minimoCorrienteCGE
    global maximoPotActivaCGE
    global promedioPotActivaCGE
    global minimoPotActivaCGE
    global maximoPotReactivaCGE
    global promedioPotReactivaCGE
    global minimoPotReactivaCGE
    global maximoPotAparenteCGE
    global promedioPotAparenteCGE
    global minimoPotAparenteCGE
    global maximoFDCGE
    global promedioFDCGE
    global minimoFDCGE
    global maximoDATCGE
    global promedioDATCGE
    global minimoDATCGE
    global Corriente15CGE
    global PotActiva15CGE
    global PotReactiva15CGE
    global PotAparente15CGE
    global FP15CGEReactivo
    global FP15CGEInductivo
    global FD15CGE
    global DAT15CGE
    global Volt15CGE
    global acceso
    basea = datetime.datetime.now()
    #print(f'Maximo Voltaje 15 CGE: {maximoVoltaje15CGE}')
    if(basea.minute==0 or basea.minute==15 or basea.minute==30 or basea.minute==45): 
         print("paso if")
         if(acceso == 0):
              print("paso if 2")
              graphVoltage(NoVoltageoffset1,ListaIrmsPeak1,samplings1,1)
              acceso = 1
              maximoVoltaje15CGE=max(Volt15CGE)
              promedioVoltaje15CGE=np.median(Volt15CGE)
              minimoVoltaje15CGE=min(Volt15CGE)
              maximoCorrienteCGE=max(Corriente15CGE)
              promedioCorriente15CGE=np.median(Corriente15CGE)
              minimoCorrienteCGE=min(Corriente15CGE)
              maximoPotActivaCGE=max(PotActiva15CGE)
              promedioPotActivaCGE=np.median(PotActiva15CGE)
              minimoPotActivaCGE=min(PotActiva15CGE)
              maximoPotReactivaCGE=max(PotReactiva15CGE)
              promedioPotReactivaCGE=np.median(PotReactiva15CGE)
              minimoPotReactivaCGE=min(PotReactiva15CGE)
              maximoPotAparenteCGE=max(PotAparente15CGE)
              promedioPotAparenteCGE=np.median(PotAparente15CGE)
              minimoPotAparenteCGE=min(PotAparente15CGE)
              if(len(FP15CGEInductivo)>0):
                     maximoFPCGEInductivo=max(FP15CGEInductivo)
                     promedioFPCGEInductivo=np.median(FP15CGEInductivo)
                     minimoFPCGEInductivo=min(FP15CGEInductivo)
              else:
                     maximoFPCGEInductivo=-0.99
                     promedioFPCGEInductivo=-0.99
                     minimoFPCGEInductivo=-0.99
              if(len(FP15CGEReactivo)>0):
                     minimoFPCGEReactivo=min(FP15CGEReactivo)
                     promedioFPCGEReactivo=np.median(FP15CGEReactivo)
                     maximoFPCGEReactivo=max(FP15CGEReactivo)
              else:
                     maximoFPCGEReactivo=0.99
                     promedioFPCGEReactivo=0.99
                     minimoFPCGEReactivo=0.99
              maximoFDCGE=max(FD15CGE)
              promedioFDCGE=np.median(FD15CGE)
              minimoFDCGE=min(FD15CGE)
              maximoDATCGE=max(DAT15CGE)
              promedioDATCGE=np.median(DAT15CGE)
              minimoDATCGE=min(DAT15CGE)
              dataCGE.insert(1,maximoVoltaje15CGE)
              dataCGE.insert(2,promedioVoltaje15CGE)
              dataCGE.insert(3,minimoVoltaje15CGE)
              dataCGE.insert(4,maximoCorrienteCGE)
              dataCGE.insert(5,promedioCorriente15CGE)
              dataCGE.insert(6,minimoCorrienteCGE)
              dataCGE.insert(7,maximoPotActivaCGE)
              dataCGE.insert(8,promedioPotActivaCGE)
              dataCGE.insert(9,minimoPotActivaCGE)
              dataCGE.insert(10,maximoPotReactivaCGE)
              dataCGE.insert(11,promedioPotReactivaCGE)
              dataCGE.insert(12,minimoPotReactivaCGE)
              dataCGE.insert(13,maximoPotAparenteCGE)
              dataCGE.insert(14,promedioPotAparenteCGE)
              dataCGE.insert(15,minimoPotAparenteCGE)
              dataCGE.insert(16,maximoFPCGEReactivo)
              dataCGE.insert(17,promedioFPCGEReactivo)
              dataCGE.insert(18,minimoFPCGEReactivo)
              dataCGE.insert(19,maximoFPCGEInductivo)
              dataCGE.insert(20,promedioFPCGEInductivo)
              dataCGE.insert(21,minimoFPCGEInductivo)
              dataCGE.insert(22,maximoFDCGE)
              dataCGE.insert(23,promedioFDCGE)
              dataCGE.insert(24,minimoFDCGE)
              dataCGE.insert(25,maximoDATCGE)
              dataCGE.insert(26,promedioDATCGE)
              dataCGE.insert(27,maximoDATCGE)
              dataCGE.insert(28,promedioDATCGE)
              dataCGE.insert(29,minimoDATCGE)
              dataCGE.insert(30,energyCGEFase11)
              dataCGE.insert(31,energyCGEFase11Hour)
              Volt15CGE=[]
              Corriente15CGE=[]
              PotActiva15CGE=[]
              PotReactiva15CGE=[]
              PotAparente15CGE=[]
              FP15CGEReactivo=[]
              FP15CGEInductivo=[]
              FD15CGE=[]
              DAT15CGE=[]
         elif(acceso==1):
              #print("paso elif CGE")
              Volt15CGE.append(vrms1)
              Corriente15CGE.append(irms1)
              PotActiva15CGE.append(ActivaCGEFase11)
              PotReactiva15CGE.append(ReactivaCGEFase11)
              PotAparente15CGE.append(AparenteCGEFase11)
              if(FPCGE0>0.0):
                    FP15CGEReactivo.append(FPCGE0)
              else: 
                    FP15CGEInductivo.append(FPCGE0)
              FD15CGE.append(FDCorrienteCGE1)
              DAT15CGE.append(DATCorrienteCGE1)
 
    else:
        Volt15CGE.append(vrms1)
        Corriente15CGE.append(irms1)
        PotActiva15CGE.append(ActivaCGEFase11)
        PotReactiva15CGE.append(ReactivaCGEFase11)
        PotAparente15CGE.append(AparenteCGEFase11)
        if(FPCGE0>0.0):
              FP15CGEReactivo.append(FPCGE0)
        else: 
              FP15CGEInductivo.append(FPCGE0)
        FD15CGE.append(FDCorrienteCGE1)
        DAT15CGE.append(DATCorrienteCGE1)
        acceso = 0
"""
        if(len(Volt15CGE)>5):
            indice=np.median(Volt15CGE)
            Volt15CGE.pop(indice)
            #print(f'Volt15CGE Despúes: {Volt15CGE}')
            indice=np.median(Corriente15CGE)
            Corriente15CGE.pop(indice)
            indice=np.median(PotActiva15CGE)
            PotActiva15CGE.pop(indice)
            indice=np.median(PotReactiva15CGE)
            PotReactiva15CGE.pop(indice)
            indice=np.median(PotAparente15CGE)
            PotAparente15CGE.pop(indice)
            if(len(FP15CGEReactivo)>=2):
                indice=np.argmax(FP15CGEReactivo)
                FP15CGEReactivo.pop(indice)
            if(len(FP15CGEInductivo)>=2):
                indice=np.argmin(FP15CGEInductivo)
                FP15CGEInductivo.pop(indice)
            indice=np.median(FD15CGE)
            FD15CGE.pop(indice)
            indice=np.median(DAT15CGE)
            DAT15CGE.pop(indice)
"""     


accesoCarga = 0
maximoVoltaje15Carga=0
promedioVoltaje15Carga=0
minimoVoltaje15Carga=0
maximoCorrienteCarga=0
promedioCorrienteCarga=0
minimoCorrienteCarga=0
maximoPotActivaCarga=0
promedioPotActivaCarga=0
minimoPotActivaCarga=0
maximoPotReactivaCarga=0
promedioPotReactivaCarga=0
minimoPotReactivaCarga=0
maximoPotAparenteCarga=0
promedioPotAparenteCarga=0
minimoPotAparenteCarga=0
maximoFPCargaInductivo=-0.99
promedioFPCargaInductivo=-0.99
minimoFPCargaInductivo=-0.99
maximoFPCargaReactivo=0.99
promedioFPCargaReactivo=0.99
minimoFPCargaReactivo=0.99
maximoFDCarga=0
promedioFDCarga=0
minimoFDCarga=0
maximoDATCarga=0
promedioDATCarga=0
minimoDATCarga=0
Volt15Carga=[]
dataCarga=[]
Corriente15Carga=[]
PotActiva15Carga=[]
PotReactiva15Carga=[]
PotAparente15Carga=[]
FP15CargaReactivo=[]
FP15CargaInductivo=[]
FD15Carga=[]
DAT15Carga=[]
def Maximo15minCarga():
    global maximoVoltaje15Carga
    global promedioVoltaje15Carga
    global minimoVoltaje15Carga
    global maximoCorrienteCarga
    global promedioCorrienteCarga
    global minimoCorrienteCarga
    global minimoPotActivaCarga
    global promedioPotActivaCarga
    global maximoPotActivaCarga
    global maximoPotReactivaCarga
    global promedioPotReactivaCarga
    global minimoPotReactivaCarga
    global maximoPotAparenteCarga
    global promedioPotAparenteCarga
    global minimoPotAparenteCarga
    global maximoFDCarga
    global promedioFDCarga
    global minimoFDCarga
    global maximoDATCarga
    global promedioDATCarga
    global minimoDATCarga
    global maximoFPCargaInductivo
    global minimoFPCargaInductivo
    global promedioFPCargaInductivo
    global maximoFPCargaReactivo
    global minimoFPCargaReactivo
    global promedioFPCargaReactivo
    global Corriente15Carga
    global PotActiva15Carga
    global PotReactiva15Carga
    global PotAparente15Carga
    global FP15CargaReactivo
    global FP15CargaInductivo
    global FD15Carga
    global DAT15Carga
    global Volt15Carga
    global accesoCarga
    basea = datetime.datetime.now()
    #print(f'Maximo Voltaje 15 Carga: {maximoVoltaje15Carga}')
    if(basea.minute==0 or basea.minute==15 or basea.minute==30 or basea.minute==45): 
         #print("paso if Carga")
         #if(len(PotAparente15Carga)>2):
               if(accesoCarga == 0):
                    #print("paso if 2 Carga")
                    graphVoltage(NoVoltageoffset2,ListaIrmsPeak2,samplings2,2)
                    accesoCarga = 1
                    maximoVoltaje15Carga=max(Volt15Carga)
                    promedioVoltaje15Carga=np.median(Volt15Carga)
                    minimoVoltaje15Carga=min(Volt15Carga)
                    maximoCorrienteCarga=max(Corriente15Carga)
                    promedioCorrienteCarga=np.median(Corriente15Carga)
                    minimoCorrienteCarga=min(Corriente15Carga)
                    maximoPotActivaCarga=max(PotActiva15Carga)
                    promedioPotActivaCarga=np.median(PotActiva15Carga)
                    minimoPotActivaCarga=min(PotActiva15Carga)
                    maximoPotReactivaCarga=max(PotReactiva15Carga)
                    promedioPotReactivaCarga=np.median(PotReactiva15Carga)
                    minimoPotReactivaCarga=min(PotReactiva15Carga)
                    maximoPotAparenteCarga=max(PotAparente15Carga)
                    promedioPotAparenteCarga=np.median(PotAparente15Carga)
                    minimoPotAparenteCarga=min(PotAparente15Carga)
                    if(len(FP15CargaInductivo)>0):
                           maximoFPCargaInductivo=max(FP15CargaInductivo)
                           promedioFPCargaInductivo=np.median(FP15CargaInductivo)
                           minimoFPCargaInductivo=min(FP15CargaInductivo)
                    else:
                           maximoFPCargaInductivo=-0.99
                           promedioFPCargaInductivo=-0.99
                           minimoFPCargaInductivo=-0.99
                    if(len(FP15CargaReactivo)>0):
                           maximoFPCargaReactivo=min(FP15CargaReactivo)
                           promedioFPCargaReactivo=min(FP15CargaReactivo)
                           minimoFPCargaReactivo=min(FP15CargaReactivo)
                    else:
                           maximoFPCargaReactivo=0.99
                           promedioFPCargaReactivo=0.99
                           minimoFPCargaReactivo=0.99
                    maximoFDCarga=max(FD15Carga)
                    promedioFDCarga=np.median(FD15Carga)
                    minimoFDCarga=min(FD15Carga)
                    maximoDATCarga=max(DAT15Carga)
                    promedioDATCarga=np.median(DAT15Carga)
                    minimoDATCarga=min(DAT15Carga)
                    dataCarga.insert(1,maximoVoltaje15Carga)
                    dataCarga.insert(2,promedioVoltaje15Carga)
                    dataCarga.insert(3,minimoVoltaje15Carga)
                    dataCarga.insert(4,maximoCorrienteCarga)
                    dataCarga.insert(5,promedioCorrienteCarga)
                    dataCarga.insert(6,minimoCorrienteCarga)
                    dataCarga.insert(7,maximoPotActivaCarga)
                    dataCarga.insert(8,promedioPotActivaCarga)
                    dataCarga.insert(9,minimoPotActivaCarga)
                    dataCarga.insert(10,maximoPotReactivaCarga)
                    dataCarga.insert(11,promedioPotReactivaCarga)
                    dataCarga.insert(12,minimoPotReactivaCarga)
                    dataCarga.insert(13,maximoPotAparenteCarga)
                    dataCarga.insert(14,promedioPotAparenteCarga)
                    dataCarga.insert(15,minimoPotAparenteCarga)
                    dataCarga.insert(16,maximoFPCargaReactivo)
                    dataCarga.insert(17,promedioFPCargaReactivo)
                    dataCarga.insert(18,minimoFPCargaReactivo)
                    dataCarga.insert(19,maximoFPCargaInductivo)
                    dataCarga.insert(20,promedioFPCargaInductivo)
                    dataCarga.insert(21,minimoFPCargaInductivo)
                    dataCarga.insert(22,maximoFDCarga)
                    dataCarga.insert(23,promedioFDCarga)
                    dataCarga.insert(24,minimoFDCarga)
                    dataCarga.insert(25,maximoDATCarga)
                    dataCarga.insert(26,promedioDATCarga)
                    dataCarga.insert(27,minimoDATCarga)
                    dataCarga.insert(28,energyCargaFase13)
                    dataCarga.insert(29,energyCargaFase13Hour)
                    Volt15Carga=[]
                    Corriente15Carga=[]
                    PotActiva15Carga=[]
                    PotReactiva15Carga=[]
                    PotAparente15Carga=[]
                    FP15CargaReactivo=[]
                    FP15CargaInductivo=[]
                    FD15Carga=[]
                    DAT15Carga=[]
               elif(accesoCarga==1):
                    #print("paso elif Carga")
                    Volt15Carga.append(vrms1)
                    Corriente15Carga.append(irms2)
                    PotActiva15Carga.append(ActivaCargaFase13)
                    PotReactiva15Carga.append(ReactivaCargaFase13)
                    PotAparente15Carga.append(AparenteCargaFase13)
                    if(FPCarga1>0.0):
                          FP15CargaReactivo.append(FPCarga1)
                    else: 
                          FP15CargaInductivo.append(FPCarga1)
                    FD15Carga.append(FDCorrienteCarga1)
                    DAT15Carga.append(DATCorrienteCarga1)
              
    else:
        Volt15Carga.append(vrms2)
        Corriente15Carga.append(irms2)
        PotActiva15Carga.append(ActivaCargaFase13)
        PotReactiva15Carga.append(ReactivaCargaFase13)
        PotAparente15Carga.append(AparenteCargaFase13)
        if(FPCarga1>0.0):
                    FP15CargaReactivo.append(FPCarga1)
        else: 
              FP15CargaInductivo.append(FPCarga1)
        FD15Carga.append(FDCorrienteCarga1)
        DAT15Carga.append(DATCorrienteCarga1)
        accesoCarga = 0
        """
        if(len(Volt15Carga)>4):
            indice=np.argmin(Volt15Carga)
            Volt15Carga.pop(indice)
            ##print(f'Volt15Carga Despúes: {Volt15Carga}')
            indice=np.argmin(Corriente15Carga)
            Corriente15Carga.pop(indice)
            indice=np.argmin(PotActiva15Carga)
            PotActiva15Carga.pop(indice)
            indice=np.argmin(PotReactiva15Carga)
            PotReactiva15Carga.pop(indice)
            indice=np.argmin(PotAparente15Carga)
            PotAparente15Carga.pop(indice)
            if(len(FP15CargaReactivo)>=2):
                indice=np.argmax(FP15CargaReactivo)
                FP15CargaReactivo.pop(indice)
            if(len(FP15CargaInductivo)>=2):
                indice=np.argmin(FP15CargaInductivo)
                FP15CargaInductivo.pop(indice)
            indice=np.argmin(FD15Carga)
            FD15Carga.pop(indice)
            indice=np.argmin(DAT15Carga)
            DAT15Carga.pop(indice)
        """


accesoPaneles = 0
maximoVoltaje15Paneles=0
promedioVoltaje15Paneles=0
minimoVoltaje15Paneles=0
maximoCorrientePaneles=0
promedioCorrientePaneles=0
minimoCorrientePaneles=0
maximoPotActivaPaneles=0
promedioPotActivaPaneles=0
minimoPotActivaPaneles=0
maximoPotReactivaPaneles=0
promedioPotReactivaPaneles=0
minimoPotReactivaPaneles=0
maximoPotAparentePaneles=0
promedioPotAparentePaneles=0
minimoPotAparentePaneles=0
maximoFPPanelesInductivo=-0.99
promedioFPPanelesInductivo=-0.99
minimoFPPanelesInductivo=-0.99
maximoFPPanelesReactivo=0.99
promedioFPPanelesReactivo=0.99
minimoFPPanelesReactivo=0.99
maximoFDPaneles=0
promedioFDPaneles=0
minimoFDPaneles=0
maximoDATPaneles=0
promedioDATPaneles=0
minimoDATPaneles=0
Volt15Paneles=[]
dataPaneles=[]
Corriente15Paneles=[]
PotActiva15Paneles=[]
PotReactiva15Paneles=[]
PotAparente15Paneles=[]
FP15PanelesReactivo=[]
FP15PanelesInductivo=[]
FD15Paneles=[]
DAT15Paneles=[]
def Maximo15minPaneles():
    global maximoVoltaje15Paneles
    global promedioVoltaje15Paneles
    global minimoVoltaje15Paneles
    global maximoCorrientePaneles
    global promedioCorrientePaneles
    global minimoCorrientePaneles
    global maximoPotActivaPaneles
    global promedioPotActivaPaneles
    global minimoPotActivaPaneles
    global maximoPotReactivaPaneles
    global promedioPotReactivaPaneles
    global minimoPotReactivaPaneles
    global maximoPotAparentePaneles
    global promedioPotAparentePaneles
    global minimoPotAparentePaneles
    global maximoFDPaneles
    global promedioFDPaneles
    global minimoFDPaneles
    global maximoDATPaneles
    global promedioDATPaneles
    global minimoDATPaneles
    global Corriente15Paneles
    global PotActiva15Paneles
    global PotReactiva15Paneles
    global PotAparente15Paneles
    global FP15PanelesReactivo
    global FP15PanelesInductivo
    global FD15Paneles
    global DAT15Paneles
    global Volt15Paneles
    global accesoPaneles
    basea = datetime.datetime.now()
    #print(f'Maximo Voltaje 15 Paneles: {maximoVoltaje15Paneles}')
    if(basea.minute==0 or basea.minute==15 or basea.minute==30 or basea.minute==45):  
         #print("paso if Paneles")
         if(accesoPaneles == 0):
              #print("paso if 2 Paneles")
              try:
                     graphVoltage(NoVoltageoffset3,ListaIrmsPeak3,samplings3,3)
                     accesoPaneles = 1
                     maximoVoltaje15Paneles=max(Volt15Paneles)
                     promedioVoltaje15Paneles=np.median(Volt15Paneles)
                     minimoVoltaje15Paneles=min(Volt15Paneles)
                     maximoCorrientePaneles=max(Corriente15Paneles)
                     promedioCorrientePaneles=np.median(Corriente15Paneles)
                     minimoCorrientePaneles=min(Corriente15Paneles)
                     maximoPotActivaPaneles=max(PotActiva15Paneles)
                     promedioPotActivaPaneles=np.median(PotActiva15Paneles)
                     minimoPotActivaPaneles=min(PotActiva15Paneles)
                     maximoPotReactivaPaneles=max(PotReactiva15Paneles)
                     promedioPotReactivaPaneles=min(PotReactiva15Paneles)
                     minimoPotReactivaPaneles=np.median(PotReactiva15Paneles)
                     maximoPotAparentePaneles=max(PotAparente15Paneles)
                     promedioPotAparentePaneles=np.median(PotAparente15Paneles)
                     minimoPotAparentePaneles=min(PotAparente15Paneles)
                     if(len(FP15PanelesInductivo)>0):
                            maximoFPPanelesInductivo=max(FP15PanelesInductivo)
                            promedioFPPanelesInductivo=np.median(FP15PanelesInductivo)
                            minimoFPPanelesInductivo=min(FP15PanelesInductivo)
                     else:
                            maximoFPPanelesInductivo=-0.99
                            promedioFPPanelesInductivo=-0.99
                            minimoFPPanelesInductivo=-0.99
                     if(len(FP15PanelesReactivo)>0):
                            maximoFPPanelesReactivo=min(FP15PanelesReactivo)
                            promedioFPPanelesReactivo=np.median(FP15PanelesReactivo)
                            minimoFPPanelesReactivo=min(FP15PanelesReactivo)
                     else:
                            maximoFPPanelesReactivo=0.99
                            promedioFPPanelesReactivo=0.99
                            minimoFPPanelesReactivo=0.99
                     maximoFDPaneles=max(FD15Paneles)
                     promedioFDPaneles=np.median(FD15Paneles)
                     minimoFDPaneles=min(FD15Paneles)
                     maximoDATPaneles=max(DAT15Paneles)
                     promedioDATPaneles=np.median(DAT15Paneles)
                     minimoDATPaneles=min(DAT15Paneles)
                     dataPaneles.insert(1,maximoVoltaje15Paneles)
                     dataPaneles.insert(1,promedioVoltaje15Paneles)
                     dataPaneles.insert(1,minimoVoltaje15Paneles)
                     dataPaneles.insert(2,maximoCorrientePaneles)
                     dataPaneles.insert(2,promedioCorrientePaneles)
                     dataPaneles.insert(2,minimoCorrientePaneles)
                     dataPaneles.insert(3,maximoPotActivaPaneles)
                     dataPaneles.insert(3,promedioPotActivaPaneles)
                     dataPaneles.insert(3,minimoPotActivaPaneles)
                     dataPaneles.insert(4,maximoPotReactivaPaneles)
                     dataPaneles.insert(4,promedioPotReactivaPaneles)
                     dataPaneles.insert(4,minimoPotReactivaPaneles)
                     dataPaneles.insert(5,maximoPotAparentePaneles)
                     dataPaneles.insert(5,promedioPotAparentePaneles)
                     dataPaneles.insert(5,minimoPotAparentePaneles)
                     dataPaneles.insert(6,maximoFPPanelesReactivo)
                     dataPaneles.insert(6,promedioFPPanelesReactivo)
                     dataPaneles.insert(6,minimoFPPanelesReactivo)
                     dataPaneles.insert(7,maximoFPPanelesInductivo)
                     dataPaneles.insert(7,promedioFPPanelesInductivo)
                     dataPaneles.insert(7,minimoFPPanelesInductivo)
                     dataPaneles.insert(8,maximoFDPaneles)
                     dataPaneles.insert(8,promedioFDPaneles)
                     dataPaneles.insert(8,minimoFDPaneles)
                     dataPaneles.insert(9,maximoDATPaneles)
                     dataPaneles.insert(9,promedioDATPaneles)
                     dataPaneles.insert(9,minimoDATPaneles)
                     dataPaneles.insert(10,energyPanelesFase12)
                     dataPaneles.insert(10,energyPanelesFase12Hour)
              except:
                  print("no hay maximos")
              Volt15Paneles=[]
              Corriente15Paneles=[]
              PotActiva15Paneles=[]
              PotReactiva15Paneles=[]
              PotAparente15Paneles=[]
              FP15PanelesReactivo=[]
              FP15PanelesInductivo=[]
              FD15Paneles=[]
              DAT15Paneles=[]
         elif(accesoPaneles==1):
              #print("paso elif Paneles")
              Volt15Paneles.append(vrms3)
              Corriente15Paneles.append(irms3)
              PotActiva15Paneles.append(ActivaPanelesFase12)
              PotReactiva15Paneles.append(ReactivaPanelesFase12)
              PotAparente15Paneles.append(AparentePanelesFase12)
              if(FPPaneles1>0.0):
                    FP15PanelesReactivo.append(FPPaneles1)
              else: 
                    FP15PanelesInductivo.append(FPPaneles1)
              FD15Paneles.append(FDCorrientePaneles1)
              DAT15Paneles.append(DATCorrientePaneles1)
              
              
 
    else:
        Volt15Paneles.append(vrms3)
        print(f'Irms Paneles : {irms3}') 
        Corriente15Paneles.append(irms3)
        PotActiva15Paneles.append(ActivaPanelesFase12)
        PotReactiva15Paneles.append(ReactivaPanelesFase12)
        PotAparente15Paneles.append(AparentePanelesFase12)
        if(FPPaneles1>0.0):
                    FP15PanelesReactivo.append(FPPaneles1)
        else: 
              FP15PanelesInductivo.append(FPPaneles1)
        FD15Paneles.append(FDCorrientePaneles1)
        DAT15Paneles.append(DATCorrientePaneles1)      
        accesoPaneles = 0
        """
        if(len(Volt15Paneles)>4):
            indice=np.argmin(Volt15Paneles)
            Volt15Paneles.pop(indice)
            #print(f'Volt15 paneles Despúes: {Volt15Paneles}')
            indice=np.argmin(Corriente15Paneles)
            Corriente15Paneles.pop(indice)
            #print(f'Corriente15 paneles Despúes: {Corriente15Paneles}')
            indice=np.argmin(PotActiva15Paneles)
            PotActiva15Paneles.pop(indice)
            indice=np.argmin(PotReactiva15Paneles)
            PotReactiva15Paneles.pop(indice)
            indice=np.argmin(PotAparente15Paneles)
            PotAparente15Paneles.pop(indice)
            if(len(FP15PanelesReactivo)>=2):
                indice=np.argmax(FP15PanelesReactivo)
                FP15PanelesReactivo.pop(indice)
            if(len(FP15PanelesInductivo)>=2):
                indice=np.argmin(FP15PanelesInductivo)
                FP15PanelesInductivo.pop(indice)
            indice=np.argmin(FD15Paneles)
            FD15Paneles.pop(indice)
            indice=np.argmin(DAT15Paneles)
            DAT15Paneles.pop(indice)
        """

VoltMax15BateriasDC=[]
CorrienteMax15BateriasDC=[]
PotMax15BateriasDC=[]
dataBateriasDC=[]
accesoBateriasDC = 0
maximoVoltaje15BateriasDC=0
promedioVoltaje15BateriasDC=0
minimoVoltaje15BateriasDC=0
maximoCorrienteBateriasDC=0
promedioCorrienteBateriasDC=0
minimoCorrienteBateriasDC=0
maximoPotBateriasDC=0
promedioPotBateriasDC=0
minimoPotBateriasDC=0
def Maximo15minBateriasDC():
    global maximoVoltaje15BateriasDC
    global promedioVoltaje15BateriasDC
    global minimoVoltaje15BateriasDC
    global maximoBateriasPanelesDC
    global maximoPotBateriasDC
    global CorrienteMax15BateriasDC
    global PotMax15BateriasDC
    global VoltMax15BateriasDC
    global accesoBateriasDC
    global PotMax15BateriasDC
    basea = datetime.datetime.now()
    #print(f'Maximo Voltaje 15 Paneles: {maximoVoltaje15Paneles}')
    if(basea.minute==0 or basea.minute==15 or basea.minute==30 or basea.minute==45):  
         #print("paso if Paneles")
         if(accesoBateriasDC == 0):
              #print("paso if 2 Paneles")
              try:
                     accesoPanelesDC = 1
                     maximoVoltaje15BateriasDC=max(VoltMax15BateriasDC)
                     promedioVoltaje15BateriasDC=np.median(VoltMax15BateriasDC)
                     minimoVoltaje15BateriasDC=min(VoltMax15BateriasDC)
                     maximoCorrienteBateriasDC=max(CorrienteMax15BateriasDC)
                     promedioCorrienteBateriasDC=np.median(CorrienteMax15BateriasDC)
                     minimoCorrienteBateriasDC=min(CorrienteMax15BateriasDC)
                     maximoPotBateriasDC=max(PotMax15BateriasDC)
                     promedioPotBateriasDC=np.median(PotMax15BateriasDC)
                     minimoPotBateriasDC=min(PotMax15BateriasDC)
                     dataBateriasDC.insert(1,maximoVoltaje15BateriasDC)
                     dataBateriasDC.insert(2,promedioVoltaje15BateriasDC)
                     dataBateriasDC.insert(3,minimoVoltaje15BateriasDC)
                     dataBateriasDC.insert(4,maximoCorrienteBateriasDC)
                     dataBateriasDC.insert(5,promedioCorrienteBateriasDC)
                     dataBateriasDC.insert(6,minimoCorrienteBateriasDC)
                     dataBateriasDC.insert(7,maximoPotBateriasDC)
                     dataBateriasDC.insert(8,promedioPotBateriasDC)
                     dataBateriasDC.insert(10,minimoPotBateriasDC)
                     dataBateriasDC.insert(11,energyBateria)
                     dataBateriasDC.insert(12,energyBateriaHora)
              except:
                  print("no hay maximos")
              VoltMax15BateriasDC=[]
              CorrienteMax15BateriasDC=[]
              PotMax15BateriasDC=[]
         elif(accesoPanelesDC==1):
              #print("paso elif Paneles")
              VoltMax15BateriasDC.append(VoltajeBaterias)
              CorrienteMax15BateriasDC.append(CorrienteBaterias)
              PotMax15BateriasDC.append(PotenciaBaterias)
 
    else:
        VoltMax15BateriasDC.append(VoltajeBaterias)
        CorrienteMax15BateriasDC.append(CorrienteBaterias)
        PotMax15BateriasDC.append(PotenciaBaterias)     
        accesoPanelesDC = 0


 
        """
        if(len(VoltMax15PanelesDC)>4):
            indice=np.argmin(VoltMax15BateriasDC)
            VoltMax15BateriasDC.pop(indice)
            indice=np.argmin(CorrienteMax15PanelesDC)
            CorrienteMax15PanelesDC.pop(indice)
            indice=np.argmin(PotMax15BateriasDC)
            PotMax15BateriasDC.pop(indice)
        """

VoltMax15PanelesDC=[]
CorrienteMax15PanelesDC=[]
PotMax15PanelesDC=[]
dataPanelesDirecta15=[]
accesoPanelesDC = 0
maximoVoltaje15PanelesDC=0
promedioVoltaje15PanelesDC=0
minimoVoltaje15PanelesDC=0
maximoCorrientePanelesDC=0
promedioCorrientePanelesDC=0
minimoCorrientePanelesDC=0
maximoPotPanelesDC=0
promedioPotPanelesDC=0
minimoPotPanelesDC=0
def Maximo15minPanelesDC():
    global maximoVoltaje15PanelesDC
    global promedioVoltaje15PanelesDC
    global minimoVoltaje15PanelesDC
    global maximoCorrientePanelesDC
    global promedioCorrientePanelesDC
    global minimoCorrientePanelesDC
    global maximoPotPanelesDC
    global promedioPotPanelesDC
    global minimoPotPanelesDC
    global CorrienteMax15PanelesDC
    global VoltMax15PanelesDC
    global accesoPanelesDC
    global PotMax15PanelesDC
    basea = datetime.datetime.now()
    #print(f'Maximo Voltaje 15 Paneles: {maximoVoltaje15Paneles}')
    if(basea.minute==0 or basea.minute==15 or basea.minute==30 or basea.minute==45):  
         #print("paso if Paneles")
         if(accesoPanelesDC == 0):
              #print("paso if 2 Paneles")
              try:
                     accesoPanelesDC = 1
                     maximoVoltaje15PanelesDC=max(VoltMax15PanelesDC)
                     promedioVoltaje15PanelesDC=np.median(VoltMax15PanelesDC)
                     minimoVoltaje15PanelesDC=min(VoltMax15PanelesDC)
                     maximoCorrientePanelesDC=max(CorrienteMax15PanelesDC)
                     promedioCorrientePanelesDC=np.median(CorrienteMax15PanelesDC)
                     minimoCorrientePanelesDC=min(CorrienteMax15PanelesDC)
                     maximoPotActivaPanelesDC=max(PotMax15PanelesDC)
                     promedioPotActivaPanelesDC=np.median(PotMax15PanelesDC)
                     minimoPotActivaPanelesDC=min(PotMax15PanelesDC)
                     dataPanelesDirecta15.insert(1,maximoVoltaje15PanelesDC)
                     dataPanelesDirecta15.insert(2,promedioVoltaje15PanelesDC)
                     dataPanelesDirecta15.insert(3,minimoVoltaje15PanelesDC)
                     dataPanelesDirecta15.insert(4,maximoCorrientePanelesDC)
                     dataPanelesDirecta15.insert(5,promedioCorrientePanelesDC)
                     dataPanelesDirecta15.insert(6,minimoCorrientePanelesDC)
                     dataPanelesDirecta15.insert(7,maximoPotActivaPanelesDC)
                     dataPanelesDirecta15.insert(8,promedioPotActivaPanelesDC)
                     dataPanelesDirecta15.insert(9,minimoPotActivaPanelesDC)
                     dataPanelesDirecta15.insert(10,energyPanelesDC)
                     dataPanelesDirecta15.insert(10,energyPanelesHoraDC)
              except:
                  print("no hay maximos")
              VoltMax15PanelesDC=[]
              CorrienteMax15PanelesDC=[]
              PotMax15PanelesDC=[]
         elif(accesoPanelesDC==1):
              #print("paso elif Paneles")
              VoltMax15PanelesDC.append(VoltajePanelesDC)
              CorrienteMax15PanelesDC.append(CorrientePanelesDC)
              PotMax15PanelesDC.append(PotenciaPanelesDC)
 
    else:
        VoltMax15PanelesDC.append(VoltajePanelesDC)
        CorrienteMax15PanelesDC.append(CorrientePanelesDC)
        PotMax15PanelesDC.append(PotenciaPanelesDC)     
        accesoPanelesDC = 0
        print("VoltMax15PanelesDC: " , VoltMax15PanelesDC)
        """
        if(len(VoltMax15PanelesDC)>4):
            indice=np.argmin(VoltMax15PanelesDC)
            VoltMax15PanelesDC.pop(indice)
            indice=np.argmin(CorrienteMax15PanelesDC)
            CorrienteMax15PanelesDC.pop(indice)
            indice=np.argmin(PotMax15PanelesDC)
            PotMax15PanelesDC.pop(indice)
        """
        
    

def excelcreate():
    global sheet1
    global sheet2
    global dest_filename
    global sheet3
    global sheet4
    global sheet5
    global sheet6
    global sheet7
    global sheet8
    global sheet9
    global sheet10
    global sheet11
    exceltime=date.today()
    book = Workbook()
    dest_filename = f'{exceltime}.xlsx'
    #sheet1 = book.active
    sheet1  = book.create_sheet("Variables Raspberry")
    sheet2 = book.create_sheet("CGE Maximos 15 Min")
    sheet3 = book.create_sheet("Carga Maximos 15 Min")
    sheet4 = book.create_sheet("Paneles Maximos 15 Min")
    sheet5 = book.create_sheet("CGE")
    sheet6 = book.create_sheet("Carga")
    sheet7 = book.create_sheet("Paneles")
    sheet8 = book.create_sheet("Baterias")
    sheet9 = book.create_sheet("Baterias15")
    sheet10 = book.create_sheet("PanelesDC")
    sheet11 = book.create_sheet("Paneles15DC")
    headings0 = ['Fecha y Hora'] + list(['T° Raspberry','Uso CPU %','RAM2'])
    headings=['Fecha y Hora'] + list(['Voltaje Maximo','Voltaje Promedio','Voltaje Minimo', 'Corriente Maximo','Corriente Promedio','Corriente Minimo','Potencia Activa Maxima','Potencia Activa Promedio','Potencia Activa Minima','Potencia Reactiva Maxima','Potencia Reactiva Promedio','Potencia Reactiva Minima','Potencia Aparente Maxima','Potencia Aparente Promedio','Potencia Aparente Minima','FPReact Maxima','FPReact Promedio','FPReact Minima','FPInduct Maxima','FPInduct Promedio','FPInduct Minima','FD Maximo''FD Promedio','FD Minima','DAT Maximo','DAT Promedio','DAT Minimo','Energia'])
    headings2=['Fecha y Hora'] + list(['Voltaje', 'Corriente','Potencia Activa','Potencia Reactiva','Potencia Aparente',
    'FP','FD','DAT','cos(phi)','Energia','Energia por Hora'])
    headings3=['Fecha y Hora'] + list(['Voltaje', 'Corriente','Potencia','Energia','Energia por Hora'])
    headings4=['Fecha y Hora'] + list(['Voltaje Maximo', 'Voltaje promedio', 'Voltaje minimo', 'Corriente Maxima','Corriente Promedio', 'Corriente Minima','Potencia Maxima','Potencia Promedio', 'Potencia Maxima','Energia total','Energia acumulada en 15'])
    ceros=list([0,0,0,0,0,0,0,0,0,0,0])
    sheet1.append(headings0)
    sheet2.append(headings)
    sheet3.append(headings)
    sheet4.append(headings)
    sheet5.append(headings2)
    sheet6.append(headings2)
    sheet7.append(headings3)
    sheet8.append(headings3)
    sheet9.append(headings4)
    sheet10.append(headings3)
    sheet11.append(headings4)
    sheet1.append(list([0,0,0]))
    sheet2.append(ceros)
    sheet3.append(ceros)
    sheet4.append(ceros)
    sheet5.append(ceros)
    sheet6.append(ceros)
    sheet7.append(ceros)
    sheet8.append(ceros)
    sheet9.append(ceros)
    sheet10.append(ceros)
    sheet11.append(ceros)

    book.save(filename = dest_filename)



def AbrirExcel():
    global dest_filename
    global energyCGEFase11
    global energyCargaFase13
    global energyPanelesFase12
    global energyBateria
    global energyPaneles
    dia=date.today()
    if(os.path.exists(f'{dia}.xlsx')):
            dest_filename = f'{dia}.xlsx'
            print("Existe")
            workbook=openpyxl.load_workbook(filename = dest_filename)
            sheet5 = workbook["CGE"]
            sheet6 = workbook["Carga"]
            sheet7 = workbook["Paneles"]
            sheet8 = workbook["Baterias"]
            sheet10 = workbook["PanelesDC"]
            largoexcelCGE=len(sheet5["FP"])
            largoexcelCarga=len(sheet6["FP"])
            largoexcelPaneles=len(sheet7["FP"])
            largoExcelBateria=len(sheet8["Voltaje"])
            largoExcelPanelesDC=len(sheet10["Voltaje"])
            #print(f'Numero de filas de paneles: {largoexcelPaneles} ')
            energyCGEFase11 = float(sheet5[f'k{largoexcelCGE}'].value)
            energyCargaFase13 = float(sheet6[f'k{largoexcelCarga}'].value)
            energyPanelesFase12 = float(sheet7[f'k{largoexcelPaneles}'].value)
            energyBateria = float(sheet8[f'k{largoExcelBateria}'].value)
            energyPaneles = float(sheet10[f'k{largoExcelPanelesDC}'].value)
            print(f'Valor Energia Paneles Acumulado: {energyPaneles} ')
    else:
            excelcreate()
            print("No Existe")

AbrirExcel()


def VariablesExcel():
       global dataVariablesAll                      
       workbook=openpyxl.load_workbook(filename = dest_filename)
       sheet1 = workbook["Variables Raspberry"]
       dataVariablesAll.insert(0,datetime.datetime.now())
       sheet1.append(list(dataVariablesAll))
       workbook.save(filename = dest_filename)
       dataVariablesAll=[]

def ExcelDataCGE():
       global dataCGEAll                      
       workbook=openpyxl.load_workbook(filename = dest_filename)
       sheet5 = workbook["CGE"]
       dataCGEAll.insert(0,datetime.datetime.now())
       sheet5.append(list(dataCGEAll))
       #print(f'Data CGE: {dataCGEAll}')
       #print("Datos Insertados Correctamente!")
       workbook.save(filename = dest_filename)
       dataCGEAll=[]

def ExcelDataCGE15():
       global dataCGE                        
       workbook=openpyxl.load_workbook(filename = dest_filename)
       sheet2 = workbook["CGE Maximos 15 Min"]
       dataCGE.insert(0,datetime.datetime.now())
       sheet2.append(list(dataCGE))
       #print(f'Data CGE: {dataCGE}')
       #print("Datos Insertados Correctamente!")
       workbook.save(filename = dest_filename)
       dataCGE=[]
      
def ExcelDataCarga():
       global dataCargaAll
       workbook=openpyxl.load_workbook(filename = dest_filename)
       sheet6 = workbook["Carga"]
       dataCargaAll.insert(0,datetime.datetime.now())
       sheet6.append(list(dataCargaAll))
       #print(f'Data Carga: {dataCarga}')
       #print("Datos Insertados Correctamente!")
       workbook.save(filename = dest_filename)
       dataCargaAll=[]

def ExcelDataCarga15():
       global dataCarga
       workbook=openpyxl.load_workbook(filename = dest_filename)
       sheet3 = workbook["Carga Maximos 15 Min"]
       dataCarga.insert(0,datetime.datetime.now())
       sheet3.append(list(dataCarga))
       #print(f'Data Carga: {dataCarga}')
       #print("Datos Insertados Correctamente!")
       workbook.save(filename = dest_filename)
       dataCarga=[]
     
def ExcelDataPaneles():
       global dataPanelesAll       
       workbook=openpyxl.load_workbook(filename = dest_filename)
       sheet7 = workbook["Paneles"]
       dataPanelesAll.insert(0,datetime.datetime.now())
       sheet7.append(list(dataPanelesAll))
       #print(f'Numero de filas de paneles: {len(sheet7["FP"])} ')
       #print(f'Data paneles: {dataPaneles}')
       #print("Datos Insertados Correctamente!")
       workbook.save(filename = dest_filename)
       dataPanelesAll=[]

def ExcelDataPaneles15():
       global dataPaneles       
       workbook=openpyxl.load_workbook(filename = dest_filename)
       sheet4 = workbook["Paneles Maximos 15 Min"]
       dataPaneles.insert(0,datetime.datetime.now())
       sheet4.append(list(dataPaneles))
       #print(f'Data paneles: {dataPaneles}')
       #print("Datos Insertados Correctamente!")
       workbook.save(filename = dest_filename)
       dataPaneles=[]

def ExcelDataBaterias():
       global dataBateriasAll       
       workbook=openpyxl.load_workbook(filename = dest_filename)
       sheet8 = workbook["Baterias"]
       dataBateriasAll.insert(0,datetime.datetime.now())
       sheet8.append(list(dataBateriasAll))
       workbook.save(filename = dest_filename)
       dataBateriasAll=[]

def ExcelDataBaterias15():
       global dataBateriasDC       
       workbook=openpyxl.load_workbook(filename = dest_filename)
       sheet9 = workbook["Baterias15"]
       dataBateriasDC.insert(0,datetime.datetime.now())
       sheet9.append(list(dataBateriasDC))
       workbook.save(filename = dest_filename)
       dataBateriasDC=[]

def ExcelDataPanelesDirecta():
       global dataPanelesDirectaAll       
       workbook=openpyxl.load_workbook(filename = dest_filename)
       sheet10 = workbook["PanelesDC"]
       dataPanelesDirectaAll.insert(0,datetime.datetime.now())
       sheet10.append(list(dataPanelesDirectaAll))
       workbook.save(filename = dest_filename)
       dataPanelesDirectaAll=[]

def ExcelDataPanelesDirecta15():
       global dataPanelesDirecta15      
       workbook=openpyxl.load_workbook(filename = dest_filename)
       sheet11 = workbook["Paneles15DC"]
       dataPanelesDirecta15.insert(0,datetime.datetime.now())
       sheet11.append(list(dataPanelesDirecta15))
       workbook.save(filename = dest_filename)
       dataPanelesDirecta15=[]



def SendEmail():
    #global dest_filename
    Lugar="Santa Cristina"
    username = "empresasserspa@gmail.com"
    password = "empresasserspa"
    destinatario = "ricardovera.93@hotmail.com"
    #destinatario2 = "ricardovera.93@hotmail.com"
    #destinatario2 = "demetrio.vera@serm.cl"
    
    mensaje = MIMEMultipart("Alternative")
    mensaje["Subject"] = "Reportes "+str(Lugar)+" "+str(datetime.date.today())
    mensaje["From"] = username
    mensaje["To"] = destinatario
    
    html = f"""
    <html>
    <body>
         <p> Hola <i>{destinatario}</i> <br>
         Reportes desde {Lugar} </b>
    </body>
    </html>
    """
    
    parte_html = MIMEText(html, "html")
    mensaje.attach(parte_html)
    
    archivo = dest_filename
    
    with open(archivo, "rb") as adjunto:
         contenido_adjunto = MIMEBase("application", "octet-stream")
         contenido_adjunto.set_payload(adjunto.read())
    
    encoders.encode_base64(contenido_adjunto)
    
    contenido_adjunto.add_header(
         "Content-Disposition",
         f"attachment; filename= {archivo}",
    )

    mensaje.attach(contenido_adjunto)
    text = mensaje.as_string()
    
    
    context = ssl.create_default_context()

    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
          server.login(username,password)
          print("Sesión Iniciada Correctamente !")
          #server.sendmail(username, destinatario, mensaje)
          server.sendmail(username, destinatario, text)
          #server.sendmail(username, destinatario2, text)
          print("Mensaje Enviado Correctamente !")


"""
while client.connected_flag: 
    #print("In loop")
    #received()
    publish(client)
    #time.sleep(5)
"""   


fecha=str(datetime.datetime.now())

f = open('mi_fichero2.txt', 'w')
try:
    f.write(fecha)
finally:
    f.close()

if __name__ == '__main__':
    received()
    #t = threading.Thread(target=received)
    #t.daemon = True
    #t.start()

