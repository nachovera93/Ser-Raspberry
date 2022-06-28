import requests
from datetime import date, timedelta
from datetime import datetime
import json
from openpyxl.chart import (
    AreaChart,
    BarChart,
    Reference,
    Series,
    PieChart,
    ProjectedPieChart,
    Reference
)
from ast import For
import xlsxwriter
import random
import time
import paho.mqtt.client as mqtt
import os
import threading
import datetime
from scipy import interpolate
from scipy.fft import fft, fftfreq
from scipy.interpolate import lagrange
from scipy import signal
from scipy.signal import savgol_filter
import numpy as np
import subprocess
from time import sleep
import sys
import RPi.GPIO as GPIO
import time
import os
import serial
import openpyxl
import smtplib, ssl
import getpass
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.mime.base import MIMEBase
import datetime
import matplotlib.pyplot as plt
import psutil
"""
    0: connection succeeded
    1: connection failed - incorrect protocol version
    2: connection failed - invalid client identifier
    3: connection failed - the broker is not available
    4: connection failed - wrong username or password
    5: connection failed - unauthorized
    6-255: undefined
"""



def BorrarArchivos():

      current_time = time.time()
      for f in os.listdir():
            creation_time = os.path.getctime(f)
            if (((current_time - creation_time) // (24 * 3600)) >= 7):
                os.unlink(f)

#BorrarArchivos()


try:                           
     esp32 = serial.Serial('/dev/ttyUSB0', 230400, timeout=0.5)
     esp32.flushInput()
except:
     esp32 = serial.Serial('/dev/ttyUSB1', 230400, timeout=0.5)
     esp32.flushInput()
     
horasetup=datetime.datetime.now()
print("Hora de comienzo:", horasetup)

broker = '18.228.175.193'   #'192.168.1.85' #mqtt server
port = 1883
dId = '123454321'
passw = 'RyfwT0cRAX'
webhook_endpoint = 'http://18.228.175.193:3001/api/getdevicecredentials'


def get_mqtt_credentials():
    global usernamemqtt
    global passwordmqtt
    global mqttopic
    global str_client_id
    global topicmqtt
    global data
    print("Getting MQTT Credentials from WebHook")
    time.sleep(2)
    toSend = {"dId": dId, "password": passw}
    respuesta = requests.post(webhook_endpoint, data=toSend)

    if(respuesta.status_code < 0):
          print("Error Sending Post Request ", respuesta.status_code)
          respuesta.close()
          return False
    if(respuesta.status_code != 200):
          print("Error in response ", respuesta.status_code)
          respuesta.close()
          return False
    if(respuesta.status_code == 200):
          print("Mqtt Credentials Obtained Successfully :)   ")
          #print("json: " ,resp.content)
          #print('- ' * 20)
          my_bytes_value = respuesta.content      #Contenido entero del json
          my_new_string = my_bytes_value.decode("utf-8").replace("'", '"')
          data = json.loads(my_new_string)
          s = json.dumps(data, indent=4, sort_keys=True)
          print(s)
          usernamemqtt = data["username"]
          #print("username:",usernamemqtt)
          passwordmqtt = data["password"]
          topicmqtt = data["topic"]   #topico al que nos vamos a suscribir
          mqttopic = f"{topicmqtt}+/actdata"
          str_client_id = f'device_{dId}_{random.randint(0, 9999)}'
          #print(mqttopic)
          respuesta.close()
          print("Ends mqtt credentials")
    return True


rcConnect = 1  
def on_disconnect(client, userdata, rc):
    if (rc != 0 and rc != 5):
        print("Unexpected disconnection, will auto-reconnect")
    elif(rc==5):
        print("Getting new credentials!")
        #get_mqtt_credentials()
        client.username_pw_set(usernamemqtt, passwordmqtt)
 
                     
def on_connected(client, userdata, flags, rc):
    if rc==0:
        client.connected_flag=True #set flag
        client.subscribe(mqttopic)
        print("connected OK")
        print("rc =",client.connected_flag)
    else:
        global rcConnect
        rcConnect=rcConnect+1
        print(f"rcConnection = {rcConnect}")
        print("Bad connection Returned code=",rc)
        client.bad_connection_flag=False

get_mqtt_credentials()     
client = mqtt.Client(str_client_id)   #Creación cliente
client.connect(broker, port)     #Conexión al broker
client.on_disconnect = on_disconnect
client.username_pw_set(usernamemqtt, passwordmqtt)
client.on_connect = on_connected
client.loop_start()



def get_cpuload():
    cpuload = psutil.cpu_percent(interval=1, percpu=False)
    return str(cpuload)

def cpu_temp():
	thermal_zone = subprocess.Popen(
	    ['cat', '/sys/class/thermal/thermal_zone0/temp'], stdout=subprocess.PIPE)
	out, err = thermal_zone.communicate()
	cpu_temp = int(out.decode())/1000
	return cpu_temp

CPU_temp = 0.0
EstateVentilador="OFF"

def Ventilador():
    global EstateVentilador
    CPU_temp = round(cpu_temp(),0)
    if CPU_temp > 53:
        GPIO.output(23, True)
        str_num = {"value":"ON","save":0}
        EstateVentilador = json.dumps(str_num)
    elif CPU_temp <= 38:
        GPIO.output(23, False)
        str_num = {"value":"OFF","save":0}
        EstateVentilador = json.dumps(str_num)

def getMaxValues(myList, quantity):
        return(sorted(list(set(myList)), reverse=True)[:quantity]) 
        #print(f'max : {max(myList)}')

def getMinValues(myList, quantity):
        return(sorted(list(set(myList)))[:quantity]) 
        #print(f'max : {max(myList)}')

Vrms=0.0
def VoltRms(MaxVoltage2):
    if(MaxVoltage2<13):
        Vrms=0
        #print(f'Vrms : {Vrms}')
        return Vrms
    Vrms=MaxVoltage2*0.71
    #Vrms=(-1.16 + 0.179*(MaxVoltage2) +  -0.00718*(MaxVoltage2**2) + 0.000155*(MaxVoltage2**3) + -0.00000203*(MaxVoltage2**4) + 0.0000000168*(MaxVoltage2**5) + -0.0000000000912*(MaxVoltage2**6) + 0.00000000000032*(MaxVoltage2**7) + -0.000000000000000703*(MaxVoltage2**8) + 0.000000000000000000875*(MaxVoltage2**9) + -0.000000000000000000000472*(MaxVoltage2**10))*MaxVoltage2
    #print(f'Vrms : {Vrms}')
    return Vrms

Irms=0.0
def CurrentRms(MaxCurrent2):
    #Irms=(-0.0248 + 0.00402*MaxCurrent2 - 0.000176*(MaxCurrent2**2) + 0.00000392*(MaxCurrent2**3) - 0.000000046*(MaxCurrent2**4) + 0.000000000284*(MaxCurrent2**5) - 0.00000000000069*(MaxCurrent2**6))*MaxCurrent2
    #print(f'Irms : {Irms}')
    if(MaxCurrent2>430):
         Irms = MaxCurrent2*0.0133
    else:
         Irms=(0.0046 + 0.000282*MaxCurrent2 - 0.00000328*(MaxCurrent2**2) + 0.0000000167*(MaxCurrent2**3) - 0.0000000000382*(MaxCurrent2**4) + 0.0000000000000322*(MaxCurrent2**5))*MaxCurrent2
    
    return Irms



Vrms0=0.0
def VoltajeRms(listVoltage):
    global Vrms0
    N = len(listVoltage)
    Squares = []

    for i in range(0,N,1):    #elevamos al cuadrado cada termino y lo amacenamos
         listsquare = listVoltage[i]*listVoltage[i]
         Squares.append(listsquare)
    
    SumSquares=0
    for i in range(0,N,1):    #Sumatoria de todos los terminos al cuadrado
         SumSquares = SumSquares + Squares[i]

    MeanSquares = (1/N)*SumSquares #Dividimos por N la sumatoria

    Vrms0=np.sqrt(MeanSquares)
   
    #print(f'Voltaje RMS : {Vrms0}')
    return Vrms0


Irms0=0.0
def CorrienteRms(listCurrent):
    global Irms0
    
    N = len(listCurrent)
    Squares = []

    for i in range(0,N,1):    #elevamos al cuadrado cada termino y lo amacenamos
         listsquare = listCurrent[i]*listCurrent[i]
         Squares.append(listsquare)
    
    SumSquares=0
    for i in range(0,N,1):    #Sumatoria de todos los terminos al cuadrado
         SumSquares = SumSquares + Squares[i]

    MeanSquares = (1/N)*SumSquares #Dividimos por N la sumatoria

    Irms0=np.sqrt(MeanSquares)
    
    #print(f'Corriente RMS : {Irms0}')
    return Irms0

#potrms=0.0
def PotenciaRms(listCurrent,listVoltage):
    global Squares
    
    N = len(listCurrent)
    Squares = []

    for i in range(0,N,1):    #elevamos al cuadrado cada termino y lo amacenamos
         listsquare = listCurrent[i]*listVoltage[i]
         Squares.append(listsquare)
    
    SumSquares=0
    for i in range(0,N,1):    #Sumatoria de todos los terminos al cuadrado
         SumSquares = SumSquares + Squares[i]

    MeanSquares = (1/N)*SumSquares #Dividimos por N la sumatoria

    #potrms=np.sqrt(MeanSquares)
    
    
    return MeanSquares

    

FDVoltage_1 = 0.0
DATVoltage_1= 0.0
FDVoltage_2 = 0.0
DATVoltage_2= 0.0
FDVoltage_3 = 0.0
DATVoltage_3= 0.0
FDVoltage_4 = 0.0
DATVoltage_4= 0.0
FDVoltage_5 = 0.0
DATVoltage_5= 0.0
FDVoltage_6 = 0.0
DATVoltage_6= 0.0
FDVoltage_7 = 0.0
DATVoltage_7= 0.0
FDVoltage_8 = 0.0
DATVoltage_8= 0.0
FDVoltage_9 = 0.0
DATVoltage_9= 0.0
sincvoltaje=0
PhaseVoltage=0.0
def VoltageFFT(list_fftVoltages, samplings,i):
    global PhaseVoltage
    global j
    global FDVoltage_1 
    global DATVoltage_1
    global FDVoltage_2
    global DATVoltage_2
    global FDVoltage_3
    global DATVoltage_3
    global FDVoltage_4 
    global DATVoltage_4
    global FDVoltage_5 
    global DATVoltage_5
    global FDVoltage_6 
    global DATVoltage_6
    global FDVoltage_7 
    global DATVoltage_7
    global FDVoltage_8 
    global DATVoltage_8
    global FDVoltage_9 
    global DATVoltage_9
    global sincvoltaje1
    p = i
    N = len(list_fftVoltages)
    T = 1 / samplings
    list_fftVoltages -= np.mean(list_fftVoltages)
    datosfft = list_fftVoltages * np.hamming(4096)
    yf = np.fft.rfft(datosfft)
    xf = fftfreq(N, T)[:N//2] 
    if (samplings > 5100):
           f = interpolate.interp1d(xf, yf[:N//2] )
           xnewv = np.arange(0, 2575, 1)  # 2550
           ynew = f(xnewv)
           ejeyabsolutv =  2.0/4096 * np.abs(ynew)
           FD = []
           complejo = []
           real=[]
           imag=[]
           z=0
           for i in range(45, 2575, 50):
                 Time1b = max(ynew[i:i+10])
                 arra = max(ejeyabsolutv[i:i+10])
                 complejo.append(Time1b)
                 real1 = Time1b.real
                 real.append(real1)
                 imag1 = Time1b.imag
                 imag.append(imag1)
                 z=z+1
                 FD.append(arra)
           FD2=[]       
           for i in range(0,len(FD)):
               if(FD[i]>(FD[0]/10)):
                   FD2.append(FD[i])
                   
           SumMagnitudEficaz = (np.sum([FD[0:len(FD)]]))
           Magnitud1 = FD[0]            
           PhaseVoltage = np.arctan(real[0]/(imag[0]))
           FDVoltage = Magnitud1/SumMagnitudEficaz
           DATVoltage= np.sqrt(((SumMagnitudEficaz**2)-(Magnitud1**2))/(Magnitud1**2))
           sincvoltaje1 = 1
           if (p == 1):
               #print("FD Voltage: ",FDVoltage)
               FDVoltage_1 = FDVoltage
               DATVoltage_1= DATVoltage
           elif (p == 2):
               FDVoltage_2 = FDVoltage
               DATVoltage_2= DATVoltage         
           elif (p == 3):
               FDVoltage_3 = FDVoltage
               DATVoltage_3= DATVoltage        
           elif (p == 4):
               FDVoltage_4 = FDVoltage
               DATVoltage_4= DATVoltage         
           elif (p == 5):
               FDVoltage_5 = FDVoltage
               DATVoltage_5= DATVoltage        
           elif (p == 6):
               FDVoltage_6 = FDVoltage
               DATVoltage_6= DATVoltage         
           elif (p == 7):
               FDVoltage_7 = FDVoltage
               DATVoltage_7= DATVoltage             
           elif (p == 8):
               FDVoltage_8 = FDVoltage
               DATVoltage_8= DATVoltage            
           elif (p == 9):
               FDVoltage_9 = FDVoltage
               DATVoltage_9= DATVoltage
                       
           


CosPhi_1= 0.0
FP_1= 0.0
DATCurrent_1= 0.0
FDCurrent_1= 0.0
CoFP_2= 0.0
DATCurrent_2= 0.0
FDCurrent_2= 0.0
CosPhi_3= 0.0
FP_3= 0.0
DATCurrent_3= 0.0
FDCurrent_3= 0.0
CosPhi_4= 0.0
FP_4= 0.0
DATCurrent_4= 0.0
FDCurrent_4= 0.0
CosPhi_5= 0.0
FP_5= 0.0
DATCurrent_5= 0.0
FDCurrent_5= 0.0
CosPhi_6= 0.0
FP_6= 0.0
DATCurrent_6= 0.0
FDCurrent_6= 0.0
CosPhi_7= 0.0
FP_7= 0.0
DATCurrent_7= 0.0
FDCurrent_7= 0.0
CosPhi_8= 0.0
FP_8= 0.0
DATCurrent_8= 0.0
FDCurrent_8= 0.0
CosPhi_9= 0.0
FP_9= 0.0
DATCurrent_9= 0.0
FDCurrent_9= 0.0
CosPhi=0.0
PhaseCurrent=0.0

def CurrentFFT(list_fftVoltages, samplings, i,Irms):
    global CosPhi
    global CosPhi_1
    global FP_1
    global DATCurrent_1
    global FDCurrent_1
    global CosPhi_2
    global FP_2
    global DATCurrent_2
    global FDCurrent_2
    global CosPhi_3
    global FP_3
    global DATCurrent_3
    global FDCurrent_3
    global CosPhi_4
    global FP_4
    global DATCurrent_4
    global FDCurrent_4
    global CosPhi_5
    global FP_5
    global DATCurrent_5
    global FDCurrent_5
    global CosPhi_6
    global FP_6
    global DATCurrent_6
    global FDCurrent_6
    global CosPhi_7
    global FP_7
    global DATCurrent_7
    global FDCurrent_7
    global CosPhi_8
    global FP_8
    global DATCurrent_8
    global FDCurrent_8
    global CosPhi_9
    global FP_9
    global DATCurrent_9
    global FDCurrent_9
    global PhaseCurrent
    global sincvoltaje1
    q = i
    N = len(list_fftVoltages)
    T = 1 / samplings
    list_fftVoltages -= np.mean(list_fftVoltages)
    datosfft = list_fftVoltages * np.hamming(4096)
    yf = np.fft.rfft(datosfft)
    
    xf = fftfreq(N, T)[:N//2]
    if (samplings > 5100):
         f = interpolate.interp1d(xf,yf[:N//2])
         xnew = np.arange(0, 2575, 1)
         ynew = f(xnew)
         ejeyabsolut =  2.0/N * np.abs(ynew)
         p = int(i)
         z=0
         FD= []
         complejo = []
         real=[]
         imag=[]
         for i in range(45, 2575, 50):
               Time1b = max(ynew[i:i+10])
               arra = max(ejeyabsolut[i:i+10])
               complejo.append(Time1b)
               real1 = Time1b.real
               real.append(real1)
               imag1 = Time1b.imag
               imag.append(imag1)
               FD.append(arra)
         repite = list(ejeyabsolut)
         #SumMagnitudEficaz = (np.sum([FD[0:len(FD)]]))*0.01
         SumMagnitudEficaz2 = (np.sum([FD[0:len(FD)]]))
         Magnitud1 = FD[0]#*0.01
         ArmonicosRestantes=SumMagnitudEficaz2-Magnitud1
         proporcion = Irms/(np.sqrt(Magnitud1**2+ArmonicosRestantes**2))
         Irmsarmonico1prop=Magnitud1*proporcion
         Irmstotalproporcionado=np.sqrt((Irmsarmonico1prop**2)+(ArmonicosRestantes*proporcion)**2)
         FDCurrent = Irmsarmonico1prop/Irms
         DATCurrent = np.sqrt((SumMagnitudEficaz2**2-Magnitud1**2)/(Magnitud1**2))
         PhaseCurrent = np.arctan(real[0]/(imag[0]))
         if (sincvoltaje1 == 1):
             if(PhaseVoltage-(PhaseCurrent)>=0):
                 desfaseCGE = "Corriente Adelantada a Voltaje"
             else:
                 desfaseCGE = "Voltaje Adelantado a Corriente"
             FP=np.cos(PhaseVoltage-PhaseCurrent)*FDCurrent
             CosPhi=np.cos(PhaseVoltage-PhaseCurrent)
             str_num_FP = {"value":FP,"save":0}
             JsonFP = json.dumps(str_num_FP)
             sincvoltaje1=0  
             if (p == 1):
                 CosPhi_1=CosPhi
                 FP_1=FP
                 DATCurrent_1=DATCurrent
                 FDCurrent_1=FDCurrent
             elif (p == 2):
                 CosPhi_2=CosPhi
                 FP_2=FP
                 DATCurrent_2=DATCurrent
                 FDCurrent_2=FDCurrent
             elif (p == 3):
                 CosPhi_3=CosPhi
                 FP_3=FP
                 DATCurrent_3=DATCurrent
                 FDCurrent_3=FDCurrent           
             elif (p == 4):
                 CosPhi_4=CosPhi
                 FP_4=FP
                 DATCurrent_4=DATCurrent
                 FDCurrent_4=FDCurrent           
             elif (p == 5):
                 CosPhi_5=CosPhi
                 FP_5=FP
                 DATCurrent_5=DATCurrent
                 FDCurrent_5=FDCurrent         
             elif (p == 6):
                 CosPhi_6=CosPhi
                 FP_6=FP
                 DATCurrent_6=DATCurrent
                 FDCurrent_6=FDCurrent          
             elif (p == 7):
                 CosPhi_7=CosPhi
                 FP_7=FP
                 DATCurrent_7=DATCurrent
                 FDCurrent_7=FDCurrent           
             elif (p == 8):
                 CosPhi_8=CosPhi
                 FP_8=FP
                 DATCurrent_8=DATCurrent
                 FDCurrent_8=FDCurrent    
             elif (p == 9):
                 CosPhi_9=CosPhi
                 FP_9=FP
                 DATCurrent_9=DATCurrent
                 FDCurrent_9=FDCurrent       


Time1a = datetime.datetime.now()
Time2a = datetime.datetime.now()
Time3a = datetime.datetime.now()
Time4a = datetime.datetime.now()
Time5a = datetime.datetime.now()
Time6a = datetime.datetime.now()
Time7a = datetime.datetime.now()
Time8a = datetime.datetime.now()
Time9a = datetime.datetime.now()
Energy_1 = 0.0
Energy_2 = 0.0
Energy_3 = 0.0
Energy_4 = 0.0
Energy_5 = 0.0
Energy_6 = 0.0
Energy_7 = 0.0
Energy_8 = 0.0
Energy_9 = 0.0
OneHourEnergy_1 = 0.0
OneHourEnergy_2 = 0.0
OneHourEnergy_3 = 0.0
OneHourEnergy_4 = 0.0
OneHourEnergy_5 = 0.0
OneHourEnergy_6 = 0.0
OneHourEnergy_7 = 0.0
OneHourEnergy_8 = 0.0
OneHourEnergy_9 = 0.0
AparentPower = 0.0
ActivePower = 0.0
ReactivePower = 0.0
acceshourenergy=0
Energy_1_TotalMes = []
Energy_2_TotalMes = []
Energy_3_TotalMes = []
Energy_4_TotalMes = []
Energy_5_TotalMes = []
Energy_6_TotalMes = []
Energy_7_TotalMes = []
Energy_8_TotalMes = []
Energy_9_TotalMes = []
def Potencias(i,Irms,Vrms,potrmsCGE):
    global vt1
    global vt2
    global vt3
    global a
    global Energy
    global Energy_1
    global Energy_2
    global Energy_3
    global Energy_4
    global Energy_5
    global Energy_6
    global Energy_7
    global Energy_8
    global Energy_9
    global OneHourEnergy_1
    global OneHourEnergy_2
    global OneHourEnergy_3
    global OneHourEnergy_4
    global OneHourEnergy_5
    global OneHourEnergy_6
    global OneHourEnergy_7
    global OneHourEnergy_8
    global OneHourEnergy_9
    global ActivePower
    global AparentPower
    global ReactivePower
    global Time1a
    global Time2a
    global Time3a
    global Time4a
    global Time5a
    global Time6a
    global Time7a
    global Time8a
    global Time9a
    global optionsave
    global vt15
    global acceshourenergy
    #print([round(OneHourEnergy_1,5),round(OneHourEnergy_2,5),round(OneHourEnergy_3,5),round(OneHourEnergy_4,5),round(OneHourEnergy_5,5),round(OneHourEnergy_6,5),round(OneHourEnergy_7,5),round(OneHourEnergy_8,5),round(OneHourEnergy_9,5)])
    TimeEnergy = datetime.datetime.now()
    if(TimeEnergy.minute==4):
        acceshourenergy=0
    if(TimeEnergy.minute==20):
        if(acceshourenergy==0):
            print("Entrando a graficar")
            workbook=openpyxl.load_workbook(filename = dest_filename)
            sheet20 = workbook[f"MaxHora Fase 1 Diario"]
            sheet21 = workbook[f"MaxHora Fase 2 Diario"]
            sheet22 = workbook[f"MaxHora Fase 3 Diario"] 
            datetim=datetime.datetime.now()-datetime.timedelta(minutes=3)
            dataHourFase1=[f'{datetim.hour}:{datetim.minute}{datetim.minute}',round(OneHourEnergy_1,5),round(OneHourEnergy_4,5),round(OneHourEnergy_7,5)]
            dataHourFase2=[f'{datetim.hour}:{datetim.minute}{datetim.minute}',round(OneHourEnergy_2,5),round(OneHourEnergy_5,5),round(OneHourEnergy_8,5)]
            dataHourFase3=[f'{datetim.hour}:{datetim.minute}{datetim.minute}',round(OneHourEnergy_3,5),round(OneHourEnergy_6,5),round(OneHourEnergy_9,5)]
            sheet20.append(list(dataHourFase1))
            sheet21.append(list(dataHourFase2))
            sheet22.append(list(dataHourFase3))
            if(TimeEnergy.hour==21 and TimeEnergy.minute==20):
                print("Entrando a GRAPH EXCEL")
                sheet23 = workbook[f"MaxHora Fase 1 Mensual"]
                sheet24 = workbook[f"MaxHora Fase 2 Mensual"]
                sheet25 = workbook[f"MaxHora Fase 3 Mensual"] 
                ##Fase 1
                chart = BarChart()
                chart.title = "Grafico Energias Fase 1"
                chart.style = 13
                chart.x_axis.title = 'Horas'
                chart.y_axis.title = 'KWh'
                chart.type = "col"
                chart.height = 10
                chart.width = 30
                Pos=len(sheet20['A'])
                print(f' Pos A :{Pos}')
                cats = Reference(sheet20, min_col=1, min_row=2, max_row=Pos+1)
                data = Reference(sheet20, min_col=2, min_row=1, max_col=4, max_row=Pos+1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                print("Graficando Fase 1")
                sheet20.add_chart(chart, f"F1")
                ##Fase 2
                chart2 = BarChart()
                chart2.title = "Grafico Energias Fase 2"
                chart2.style = 13
                chart2.x_axis.title = 'Horas'
                chart2.y_axis.title = 'KWh'
                chart2.type = "col"
                chart2.height = 10
                chart2.width = 30
                Pos2=len(sheet21['A'])
                cats2 = Reference(sheet21, min_col=1, min_row=2, max_row=Pos2+1)
                data2 = Reference(sheet21, min_col=2, min_row=1, max_col=4, max_row=Pos2+1)
                chart2.add_data(data2, titles_from_data=True)
                chart2.set_categories(cats2)
                print("Graficando Fase 2")
                sheet21.add_chart(chart2, f"F1")
                
                
                #Fase 3
                chart3 = BarChart()
                chart3.title = "Grafico Energias Fase 1"
                chart3.style = 12
                chart3.x_axis.title = 'Horas'
                chart3.y_axis.title = 'KWh'
                chart3.type = "col"
                chart3.height = 10
                chart3.width = 30
                Pos3=len(sheet22['A'])
                cats3 = Reference(sheet22, min_col=1, min_row=2, max_row=Pos3+1)
                data3 = Reference(sheet22, min_col=2, min_row=1, max_col=4, max_row=Pos3+1)
                chart3.add_data(data3, titles_from_data=True)
                chart3.set_categories(cats3)
                print("Graficando Fase 3")
                sheet22.add_chart(chart3, f"F1") 
                
                dia0=date.today()
                dia=str(dia0-timedelta(1))
                x=0
                for f in os.listdir('/home/pi/Desktop/Ser-Raspberry/'):
                    if(dia[5:7]==f[5:7]):
                        x = x+1
                        #print(f)  
                        workbook=openpyxl.load_workbook(filename = f'{f}')
                        try:
                            
                            sheet1 = workbook[f"REDCompañia-Fase-1"]
                            sheet2 = workbook[f"REDCompañia-Fase-2"]
                            sheet3 = workbook[f"REDCompañia-Fase-3"]
                            sheet4 = workbook[f"CentralFotovoltaica-Fase-1"]
                            sheet5 = workbook[f"CentralFotovoltaica-Fase-2"]
                            sheet6 = workbook[f"CentralFotovoltaica-Fase-3"]
                            sheet7 = workbook[f"ConsumoCliente-Fase-1"]
                            sheet8 = workbook[f"ConsumoCliente-Fase-2"]
                            sheet9 = workbook[f"ConsumoCliente-Fase-3"]
                            LargeSheet11=len(sheet1["FP"])
                            LargeSheet21=len(sheet2["FP"])
                            LargeSheet31=len(sheet3["FP"])
                            LargeSheet41=len(sheet4["FP"])
                            LargeSheet51=len(sheet5["FP"])
                            LargeSheet61=len(sheet6["FP"])
                            LargeSheet71=len(sheet7["FP"])
                            LargeSheet81=len(sheet8["FP"])
                            LargeSheet91=len(sheet9["FP"])
                            Energy_1 = float(sheet1[f'm{LargeSheet11}'].value)
                            Energy_2 = float(sheet2[f'm{LargeSheet21}'].value)
                            Energy_3 = float(sheet3[f'm{LargeSheet31}'].value)
                            Energy_4 = float(sheet4[f'm{LargeSheet41}'].value)
                            Energy_5 = float(sheet5[f'm{LargeSheet51}'].value)
                            Energy_6 = float(sheet6[f'm{LargeSheet61}'].value)
                            Energy_7 = float(sheet7[f'm{LargeSheet71}'].value)
                            Energy_8 = float(sheet8[f'm{LargeSheet81}'].value)
                            Energy_9 = float(sheet9[f'm{LargeSheet91}'].value)   
                            Energy_1_TotalMes.append(Energy_1)
                            Energy_2_TotalMes.append(Energy_2)
                            Energy_3_TotalMes.append(Energy_3)
                            Energy_4_TotalMes.append(Energy_4)
                            Energy_5_TotalMes.append(Energy_5)
                            Energy_6_TotalMes.append(Energy_6)
                            Energy_7_TotalMes.append(Energy_7)
                            Energy_8_TotalMes.append(Energy_8)
                            Energy_9_TotalMes.append(Energy_9)
                            print(f)
                        except:
                            continue
                Suma_Mes_1=np.sum(Energy_1_TotalMes)
                Suma_Mes_2=np.sum(Energy_2_TotalMes)
                Suma_Mes_3=np.sum(Energy_3_TotalMes)
                Suma_Mes_4=np.sum(Energy_4_TotalMes)
                Suma_Mes_5=np.sum(Energy_5_TotalMes)
                Suma_Mes_6=np.sum(Energy_6_TotalMes)
                Suma_Mes_7=np.sum(Energy_7_TotalMes)
                Suma_Mes_8=np.sum(Energy_8_TotalMes)
                Suma_Mes_9=np.sum(Energy_9_TotalMes)
                 
                sheet23['E1'] = 'Acumulado Energia Mes REDCompañia-Fase-1'  
                sheet23['F1'] = 'Acumulado Energia REDCompañia-Fase-2'
                sheet23['G1'] = 'Acumulado Energia REDCompañia-Fase-3'
                sheet24['E1'] = 'Acumulado Energia CentralFotovoltaica-Fase-1'
                sheet24['F1'] = 'Acumulado Energia CentralFotovoltaica-Fase-2'
                sheet24['G1'] = 'Acumulado Energia CentralFotovoltaica-Fase-3'
                sheet25['E1'] = 'Acumulado Energia ConsumoCliente-Fase-1'
                sheet25['F1'] = 'Acumulado Energia ConsumoCliente-Fase-2'
                sheet25['G1'] = 'Acumulado Energia ConsumoCliente-Fase-3'
                
                sheet23['E2'] = Suma_Mes_1
                sheet23['F2'] = Suma_Mes_2
                sheet23['G2'] = Suma_Mes_3
                sheet24['E2'] = Suma_Mes_4
                sheet24['F2'] = Suma_Mes_5
                sheet24['G2'] = Suma_Mes_6
                sheet25['E2'] = Suma_Mes_7
                sheet25['F2'] = Suma_Mes_8
                sheet25['G2'] = Suma_Mes_9
                Energy_1_TotalMes = []
                Energy_2_TotalMes = []
                Energy_3_TotalMes = []
                Energy_4_TotalMes = []
                Energy_5_TotalMes = []
                Energy_6_TotalMes = []
                Energy_7_TotalMes = []
                Energy_8_TotalMes = []
                Energy_9_TotalMes = []
            workbook.save(filename = dest_filename)
            SendDataToBroker(q=1,k=k1,f=f1,EnergiaHora=f'{OneHourEnergy_1}')
            print("Enviando Hora Max Energia 1")
            OneHourEnergy_1=0
            SendDataToBroker(q=2,k=k1,f=f2,EnergiaHora=f'{OneHourEnergy_2}')
            print("Enviando Hora Max Energia 2")
            OneHourEnergy_2=0   
            SendDataToBroker(q=3,k=k1,f=f3,EnergiaHora=f'{OneHourEnergy_3}')
            print("Enviando Hora Max Energia 3")
            OneHourEnergy_3=0   
            SendDataToBroker(q=4,k=k2,f=f1,EnergiaHora=f'{OneHourEnergy_4}')
            print("Enviando Hora Max Energia 4")
            OneHourEnergy_4=0    
            SendDataToBroker(q=5,k=k2,f=f2,EnergiaHora=f'{OneHourEnergy_5}')
            print("Enviando Hora Max Energia 5")
            OneHourEnergy_5=0   
            SendDataToBroker(q=6,k=k2,f=f3,EnergiaHora=f'{OneHourEnergy_6}')
            print("Enviando Hora Max Energia 6")
            OneHourEnergy_6=0     
            SendDataToBroker(q=7,k=k3,f=f1,EnergiaHora=f'{OneHourEnergy_7}')
            print("Enviando Hora Max Energia 7")
            OneHourEnergy_7=0     
            SendDataToBroker(q=8,k=k3,f=f2,EnergiaHora=f'{OneHourEnergy_8}')
            print("Enviando Hora Max Energia 8")
            OneHourEnergy_8=0       
            SendDataToBroker(q=9,k=k3,f=f3,EnergiaHora=f'{OneHourEnergy_9}')
            print("Enviando Hora Max Energia 9")
            OneHourEnergy_9=0
            vt15=time.time()
            dataHourFase1=[]
            dataHourFase2=[]
            dataHourFase3=[]
            acceshourenergy=1
            
    if(TimeEnergy.hour==0 and TimeEnergy.minute==4):
            Energy_1=0
            Energy_2=0
            Energy_3=0
            Energy_4=0
            Energy_5=0
            Energy_6=0
            Energy_7=0
            Energy_8=0
            Energy_9=0
    AparentPower = Vrms*Irms
    if (potrmsCGE>=0):
          ActivePower = Vrms*Irms*CosPhi
          ActivePower = np.abs(ActivePower)
    else:
          ActivePower = Vrms*Irms*CosPhi
          ActivePower = np.abs(ActivePower)
          ActivePower = ActivePower*(-1)
    ReactivePower = Vrms*Irms*np.sin(PhaseVoltage-PhaseCurrent)
    optionsave=1
    if (i == 1):
        Time1b = datetime.datetime.now()
        delta=(((Time1b - Time1a).microseconds)/1000+((Time1b - Time1a).seconds)*1000)/10000000000
        Energy_1 += np.abs(AparentPower*delta*2.9) #ActivePower
        OneHourEnergy_1 += np.abs(AparentPower*delta*2.9) #ActivePower
        Time1a = datetime.datetime.now()
        AparentPower_1 = AparentPower
        ActivePower_1 = ActivePower
        ReactivePower_1 = ReactivePower
        SaveDataCsv(Vrms,Irms,ActivePower_1,ReactivePower_1,AparentPower_1,FP_1,CosPhi_1,FDVoltage_1,FDCurrent_1,DATVoltage_1,DATCurrent_1,Energy_1,OneHourEnergy_1,i,k1,f1)
        SendDataToBroker(q=i,k=k1,f=f1,Voltaje=f"{Vrms}",Corriente=f"{Irms}",Potencia=f"{AparentPower_1}",Energia=f"{Energy_1}")
        vt1=time.time() 
        Maximo15min_1(Vrms,Irms,ActivePower_1,ReactivePower_1,AparentPower_1,FP_1,FDVoltage_1,FDCurrent_1,DATVoltage_1,DATCurrent_1,OneHourEnergy_1,Energy_1,i,k1,f1)
        #{key}-{q}-{f}-{k}
    elif (i == 2):
        Time2b = datetime.datetime.now()
        delta=(((Time2b - Time2a).microseconds)/1000+((Time2b - Time2a).seconds)*1000)/10000000000
        Energy_2 += np.abs(AparentPower*delta*2.9)
        OneHourEnergy_2 += np.abs(AparentPower*delta*2.9)
        Time2a = datetime.datetime.now()
        AparentPower_2 = AparentPower
        ActivePower_2 = ActivePower
        ReactivePower_2 = ReactivePower 
        SaveDataCsv(Vrms,Irms,ActivePower_2,ReactivePower_2,AparentPower_2,FP_2,CosPhi_2,FDVoltage_2,FDCurrent_2,DATVoltage_2,DATCurrent_2,Energy_2,OneHourEnergy_2,i,k1,f2)
        SendDataToBroker(q=i,k=k1,f=f2,Voltaje=f"{Vrms}",Corriente=f"{Irms}",Potencia=f"{AparentPower_2}",Energia=f"{Energy_2}")
        vt2=time.time() 
        Maximo15min_2(Vrms,Irms,ActivePower_2,ReactivePower_2,AparentPower_2,FP_2,FDVoltage_2,FDCurrent_2,DATVoltage_2,DATCurrent_2,OneHourEnergy_2,Energy_2,i,k1,f2)       
    elif (i == 3):
        Time3b = datetime.datetime.now()
        delta=(((Time3b - Time3a).microseconds)/1000+((Time3b - Time3a).seconds)*1000)/10000000000
        Energy_3 += np.abs(AparentPower*delta*2.9)
        OneHourEnergy_3 += np.abs(AparentPower*delta*2.9)
        Time3a = datetime.datetime.now()
        AparentPower_3 = AparentPower
        ActivePower_3 = ActivePower
        ReactivePower_3 = ReactivePower
        SaveDataCsv(Vrms,Irms,ActivePower_3,ReactivePower_3,AparentPower_3,FP_3,CosPhi_3,FDVoltage_3,FDCurrent_3,DATVoltage_3,DATCurrent_3,Energy_3,OneHourEnergy_3,i,k1,f3)
        SendDataToBroker(q=i,k=k1,f=f3,Voltaje=f"{Vrms}",Corriente=f"{Irms}",Potencia=f"{AparentPower_3}",Energia=f"{Energy_3}")
        vt3=time.time() 
        Maximo15min_3(Vrms,Irms,ActivePower_3,ReactivePower_3,AparentPower_3,FP_3,FDVoltage_3,FDCurrent_3,DATVoltage_3,DATCurrent_3,OneHourEnergy_3,Energy_3,i,k1,f3)             
    elif (i == 4):
        Time4b = datetime.datetime.now()
        delta=(((Time4b - Time4a).microseconds)/1000+((Time4b - Time4a).seconds)*1000)/10000000000
        Energy_4 += np.abs(AparentPower*delta*2.9)
        OneHourEnergy_4 += np.abs(AparentPower*delta*2.9)
        Time4a = datetime.datetime.now()
        AparentPower_4 = AparentPower
        ActivePower_4 = ActivePower
        ReactivePower_4 = ReactivePower
        SaveDataCsv(Vrms,Irms,ActivePower_4,ReactivePower_4,AparentPower_4,FP_4,CosPhi_4,FDVoltage_4,FDCurrent_4,DATVoltage_4,DATCurrent_4,Energy_4,OneHourEnergy_4,i,k2,f1)
        SendDataToBroker(q=i,k=k2,f=f1,Voltaje=f"{Vrms}",Corriente=f"{Irms}",Potencia=f"{AparentPower_4}",Energia=f"{Energy_4}")
        #SendDataToBroker(q=i,k=k2,f=f1,Voltaje=f"{Vrms}",Corriente=f"{Irms}",Potencia=f"{AparentPower_4}",Energia=f"{Energy_4}")
        Maximo15min_4(Vrms,Irms,ActivePower_4,ReactivePower_4,AparentPower_4,FP_4,FDVoltage_4,FDCurrent_4,DATVoltage_4,DATCurrent_4,OneHourEnergy_4,Energy_4,i,k2,f1)              
    elif (i == 5):
        Time5b = datetime.datetime.now()
        delta=(((Time5b - Time5a).microseconds)/1000+((Time5b - Time5a).seconds)*1000)/10000000000
        Energy_5 += np.abs(AparentPower*delta*2.9)
        OneHourEnergy_5 += np.abs(AparentPower*delta*2.9)
        Time5a = datetime.datetime.now()
        AparentPower_5 = AparentPower
        ActivePower_5 = ActivePower
        ReactivePower_5 = ReactivePower
        SaveDataCsv(Vrms,Irms,ActivePower_5,ReactivePower_5,AparentPower_5,FP_5,CosPhi_5,FDVoltage_5,FDCurrent_5,DATVoltage_5,DATCurrent_5,Energy_5,OneHourEnergy_5,i,k2,f2)
        SendDataToBroker(q=i,k=k2,f=f2,Voltaje=f"{Vrms}",Corriente=f"{Irms}",Potencia=f"{AparentPower_5}",Energia=f"{Energy_5}")
        Maximo15min_5(Vrms,Irms,ActivePower_5,ReactivePower_5,AparentPower_5,FP_5,FDVoltage_5,FDCurrent_5,DATVoltage_5,DATCurrent_5,OneHourEnergy_5,Energy_5,i,k2,f2)               
    elif (i == 6):
        Time6b = datetime.datetime.now()
        delta=(((Time6b - Time6a).microseconds)/1000+((Time6b - Time6a).seconds)*1000)/10000000000
        Energy_6 += np.abs(AparentPower*delta*2.9)
        OneHourEnergy_6 += np.abs(AparentPower*delta*2.9)
        Time6a = datetime.datetime.now()
        AparentPower_6 = AparentPower
        ActivePower_6 = ActivePower
        ReactivePower_6 = ReactivePower
        SaveDataCsv(Vrms,Irms,ActivePower_6,ReactivePower_6,AparentPower_6,FP_6,CosPhi_6,FDVoltage_6,FDCurrent_6,DATVoltage_6,DATCurrent_6,Energy_6,OneHourEnergy_6,i,k2,f3)
        SendDataToBroker(q=i,k=k2,f=f3,Voltaje=f"{Vrms}",Corriente=f"{Irms}",Potencia=f"{AparentPower_6}",Energia=f"{Energy_6}")
        Maximo15min_6(Vrms,Irms,ActivePower_6,ReactivePower_6,AparentPower_6,FP_6,FDVoltage_6,FDCurrent_6,DATVoltage_6,DATCurrent_6,OneHourEnergy_6,Energy_6,i,k2,f3)           
    elif (i == 7):
        Time7b = datetime.datetime.now()
        delta=(((Time7b - Time7a).microseconds)/1000+((Time7b - Time7a).seconds)*1000)/10000000000
        Energy_7 += np.abs(AparentPower*delta*2.9)
        OneHourEnergy_7 += np.abs(AparentPower*delta*2.9)
        Time7a = datetime.datetime.now()
        AparentPower_7 = AparentPower
        ActivePower_7 = ActivePower
        ReactivePower_7 = ReactivePower
        SaveDataCsv(Vrms,Irms,ActivePower_7,ReactivePower_7,AparentPower_7,FP_7,CosPhi_7,FDVoltage_7,FDCurrent_7,DATVoltage_7,DATCurrent_7,Energy_7,OneHourEnergy_7,i,k3,f1)
        SendDataToBroker(q=i,k=k3,f=f1,Voltaje=f"{Vrms}",Corriente=f"{Irms}",Potencia=f"{AparentPower_7}",Energia=f"{Energy_7}")
        Maximo15min_7(Vrms,Irms,ActivePower_7,ReactivePower_7,AparentPower_7,FP_7,FDVoltage_7,FDCurrent_7,DATVoltage_7,DATCurrent_7,OneHourEnergy_7,Energy_7,i,k3,f1)           
    elif (i == 8):
        Time8b = datetime.datetime.now()
        delta=(((Time8b - Time8a).microseconds)/1000+((Time8b - Time8a).seconds)*1000)/10000000000
        Energy_8 += np.abs(AparentPower*delta*2.9)
        OneHourEnergy_8 += np.abs(AparentPower*delta*2.9)
        Time8a = datetime.datetime.now()
        AparentPower_8 = AparentPower
        ActivePower_8 = ActivePower
        ReactivePower_8 = ReactivePower
        SaveDataCsv(Vrms,Irms,ActivePower_8,ReactivePower_8,AparentPower_8,FP_8,CosPhi_8,FDVoltage_8,FDCurrent_8,DATVoltage_8,DATCurrent_8,Energy_8,OneHourEnergy_8,i,k3,f2)
        SendDataToBroker(q=i,k=k3,f=f2,Voltaje=f"{Vrms}",Corriente=f"{Irms}",Potencia=f"{AparentPower_8}",Energia=f"{Energy_8}")
        Maximo15min_8(Vrms,Irms,ActivePower_8,ReactivePower_8,AparentPower_8,FP_8,FDVoltage_8,FDCurrent_8,DATVoltage_8,DATCurrent_8,OneHourEnergy_8,Energy_8,i,k3,f2)             
    elif (i == 9):
        Time9b = datetime.datetime.now()
        delta=(((Time9b - Time9a).microseconds)/1000+((Time9b - Time9a).seconds)*1000)/10000000000
        Energy_9 += np.abs(AparentPower*delta*2.9)
        OneHourEnergy_9 += np.abs(AparentPower*delta*2.9)
        Time9a = datetime.datetime.now()
        AparentPower_9 = AparentPower
        ActivePower_9 = ActivePower
        ReactivePower_9 = ReactivePower
        SaveDataCsv(Vrms,Irms,ActivePower_9,ReactivePower_9,AparentPower_9,FP_9,CosPhi_9,FDVoltage_9,FDCurrent_9,DATVoltage_9,DATCurrent_9,Energy_9,OneHourEnergy_9,i,k3,f3)
        SendDataToBroker(q=i,k=k3,f=f3,Voltaje=f"{Vrms}",Corriente=f"{Irms}",Potencia=f"{AparentPower_9}",Energia=f"{Energy_9}")
        Maximo15min_9(Vrms,Irms,ActivePower_9,ReactivePower_9,AparentPower_9,FP_9,FDVoltage_9,FDCurrent_9,DATVoltage_9,DATCurrent_9,OneHourEnergy_9,Energy_9,i,k3,f3) 
       
    
    
font = {'family': 'serif',
        'color':  'darkred',
        'weight': 'normal',
        'size': 8,
        }

def graphVoltage(list_fftVoltage,list_FinalCurrent,samplings,i):
        global ax
        global imagenVoltaje
        i = str(i)
        global render
        fig=plt.figure(figsize=(8,6))
        tiempo = 1/(samplings*(0.001/4200))
        tiempoms = np.arange(0,tiempo,tiempo/4096)


        ax = fig.add_subplot(9,1,1)
        ax.plot(tiempoms,list_FinalCurrent,color="green", label="Corriente")
        if(i=="1"):
             plt.title(f'Corriente | I: {round(Irms,2)}  |  P-Activa: {round(ActivePower,2)} | P-Aparente: {round(AparentPower1,2)}  |  P-Reactiva:{round(ReactivePower1,2)}  ',fontdict=font)
        if(i=="2"):
             plt.title(f'Corriente | I: {round(Irms2,2)}  |  P-Activa: {round(Activa2Fase13,2)} | P-Aparente: {round(Aparente2Fase13,2)}  |  P-Reactiva:{round(Reactiva2Fase13,2)}  ',fontdict=font)
        if(i=="3"):
             plt.title(f'Corriente | I: {round(Irms3,2)}  |  P-Activa: {round(ActivaPanelesFase12,2)} | P-Aparente: {round(AparentePanelesFase12,2)}  |  P-Reactiva:{round(ReactivaPanelesFase12,2)}  ',fontdict=font)
             
        ax.legend(loc='upper left')
        ax.set_xlabel('Tiempo (mS)',fontdict=font)
        ax = fig.add_subplot(9,1,3)
        ax.plot(tiempoms,list_fftVoltage,color="blue", label="Voltaje")
        if(i=="1"):
             plt.title(f'Voltaje | V: {round(Vrms,2)} |  FP: {round(FP,2)}',fontdict=font)
        if(i=="2"):
             plt.title(f'Voltaje | V: {round(Vrms2,2)} |  FP: {round(FP21,2)}',fontdict=font)
        if(i=="3"):
             plt.title(f'Voltaje | V: {round(Vrms3,2)} |  FP: {round(FPPaneles1,2)}',fontdict=font)
        
        ax = fig.add_subplot(9,1,5)
        ax.plot(tiempoms,Squares,color="red", label="Pot-Activa")

        ax = fig.add_subplot(9,1,7)
        if(i=="1"):
            plt.title(f'FFT Corriente | DAT: {round(DATCurrent,2)}, FD: {round(FDCurrent,2)} |   cos phi: {round(CosPhi,2)} | phase voltaje CGE : {round(PhaseVoltage,2)} | phase Corriente CGE : {round(PhaseCurrent,2)}',fontdict=font)
        if(i=="2"):
            plt.title(f'FFT Corriente | DAT: {round(DATCorriente21,2)}, FD: {round(FDCorriente21,2)} |   cos phi: {round(cosphi2,2)} | phase voltaje 2 : {round(phasevoltaje2,2)} | phase Corriente 2 : {round(phasecorriente2,2)} ',fontdict=font)
        if(i=="3"):
            plt.title(f'FFT Corriente | DAT: {round(DATCorrientePaneles1,2)}, FD: {round(FDCorrientePaneles1,2)}  |   cos phi: {round(cosphiPaneles,2)} | phase voltaje Paneles: {round(phasevoltajePaneles,2)} | phase Corriente Paneles : {round(phasecorrientePaneles,2)}',fontdict=font)

        ax.plot(xnew,ejeyabsolut)
        rangex = np.zeros(28)
        n=0
        for h in range(50, 2600, 100):
           rangex[n]=h
           n = n+1
        ax.xaxis.set_ticks(rangex)  
        ax.grid(True)
        ax.set_xlabel('Frecuencia (Hz)',fontdict=font)
        
        ax = fig.add_subplot(9,1,9)
        if(i=="1"):
            plt.title(f'{desfaseCGE}',fontdict=font)
        if(i=="2"):
            plt.title(f'{desfase2}',fontdict=font)
        if(i=="3"):
            plt.title(f'{desfasePaneles} ',fontdict=font)

        ax.plot(xnewv,ejeyabsolutv)
        rangex = np.zeros(28)
        n=0
        for h in range(50, 2600, 100):
           rangex[n]=h
           n = n+1
        ax.xaxis.set_ticks(rangex)  
        ax.grid(True)
        ax.set_xlabel('Frecuencia (Hz)',fontdict=font)
        #ax.set_ylabel('|dB|',fontdict=font) 

        oldepoch = time.time()
        st = datetime.datetime.fromtimestamp(oldepoch).strftime('%Y-%m-%d-%H:%M:%S') 
        #plt.xlabel("Tiempo(ms)",fontdict=font)
        #plt.title(f'FFT Voltaje |    DAT: {DATVoltage}     |     FD: {FDVoltage}    |   cos phi: {round(CosPhi,2)}',fontdict=font)
        ax.set_xlabel(f'Frecuencia (Hz)  |    fecha: {st} ',fontdict=font)
        #ax.set_ylabel('Pk-Pk',fontdict=font) 
        imagenVoltaje = f'images{i}/{st}.png'
        plt.savefig(imagenVoltaje)
        
    


vt=time.time()
vt1=time.time()
vt2=time.time()
vt3=time.time()
vt4=time.time()
vt5=time.time()
vt6=time.time()
vt7=time.time()
vt8=time.time()
vt9=time.time()
vt15=time.time()
vt115=time.time()
vt215=time.time()
vt315=time.time()
vt415=time.time()
vt515=time.time()
vt615=time.time()
vt715=time.time()
vt815=time.time()
vt915=time.time()


optionsave=1
k1="REDCompañia"
k2="CentralFotovoltaica"
k3="ConsumoCliente"
f1="Fase-1"
f2="Fase-2"
f3="Fase-3"
#{key}-{f}-{k} 90 Variables
#Voltaje-Fase-1-REDCompañia // Corriente-Fase-1-REDCompañia // Potencia-Fase-1-REDCompañia // Potencia-Fase-1-REDCompañia
#VoltajeMax-Fase-1-REDCompañia // VoltajePromedio-Fase-1-REDCompañia // VoltajeMin-Fase-1-REDCompañia // CorrienteMax-Fase-1-REDCompañia // PotenciaMax-Fase-1-REDCompañia // EnergiaMax-Fase-1-REDCompañia
#Voltaje-Fase-2-REDCompañia // Corriente-Fase-2-REDCompañia // Potencia-Fase-2-REDCompañia // Potencia-Fase-2-REDCompañia
#VoltajeMax-Fase-2-REDCompañia // VoltajePromedio-Fase-2-REDCompañia // VoltajeMin-Fase-2-REDCompañia // CorrienteMax-Fase-2-REDCompañia // PotenciaMax-Fase-2-REDCompañia // EnergiaMax-Fase-2-REDCompañia
#Voltaje-Fase-3-REDCompañia // Corriente-Fase-3-REDCompañia // Potencia-Fase-3-REDCompañia // Potencia-Fase-3-REDCompañia
#VoltajeMax-Fase-3-REDCompañia // VoltajePromedio-Fase-3-REDCompañia // VoltajeMin-Fase-3-REDCompañia // CorrienteMax-Fase-3-REDCompañia // PotenciaMax-Fase-3-REDCompañia // EnergiaMax-Fase-3-REDCompañia

#Voltaje-Fase-1-CentralFotovoltaica // Corriente-Fase-1-CentralFotovoltaica // Potencia-Fase-1-CentralFotovoltaica // Potencia-Fase-1-CentralFotovoltaica
#VoltajeMax-Fase-1-CentralFotovoltaica // VoltajePromedio-Fase-1-CentralFotovoltaica // VoltajeMin-Fase-1-CentralFotovoltaica // CorrienteMax-Fase-1-CentralFotovoltaica // PotenciaMax-Fase-1-CentralFotovoltaica // EnergiaMax-Fase-1-CentralFotovoltaica
#Voltaje-Fase-2-CentralFotovoltaica // Corriente-Fase-2-CentralFotovoltaica // Potencia-Fase-2-CentralFotovoltaica // Potencia-Fase-2-CentralFotovoltaica
#VoltajeMax-Fase-2-CentralFotovoltaica // VoltajePromedio-Fase-2-CentralFotovoltaica // VoltajeMin-Fase-2-CentralFotovoltaica // CorrienteMax-Fase-2-CentralFotovoltaica // PotenciaMax-Fase-2-CentralFotovoltaica // EnergiaMax-Fase-2-CentralFotovoltaica
#Voltaje-Fase-3-CentralFotovoltaica // Corriente-Fase-3-CentralFotovoltaica // Potencia-Fase-3-CentralFotovoltaica // Potencia-Fase-3-CentralFotovoltaica
#VoltajeMax-Fase-3-CentralFotovoltaica // VoltajePromedio-Fase-3-CentralFotovoltaica // VoltajeMin-Fase-3-CentralFotovoltaica // CorrienteMax-Fase-3-CentralFotovoltaica // PotenciaMax-Fase-3-CentralFotovoltaica // EnergiaMax-Fase-3-CentralFotovoltaica

#Voltaje-Fase-1-ConsumoCliente // Corriente-Fase-1-ConsumoCliente // Potencia-Fase-1-ConsumoCliente // Potencia-Fase-1-ConsumoCliente
#VoltajeMax-Fase-1-ConsumoCliente // VoltajePromedio-Fase-1-ConsumoCliente // VoltajeMin-Fase-1-ConsumoCliente // CorrienteMax-Fase-1-ConsumoCliente // PotenciaMax-Fase-1-ConsumoCliente // EnergiaMax-Fase-1-ConsumoCliente
#Voltaje-Fase-2-ConsumoCliente // Corriente-Fase-2-ConsumoCliente // Potencia-Fase-2-ConsumoCliente // Potencia-Fase-2-ConsumoCliente
#VoltajeMax-Fase-2-ConsumoCliente // VoltajePromedio-Fase-2-ConsumoCliente // VoltajeMin-Fase-2-ConsumoCliente // CorrienteMax-Fase-2-ConsumoCliente // PotenciaMax-Fase-2-ConsumoCliente // EnergiaMax-Fase-2-ConsumoCliente
#Voltaje-Fase-3-ConsumoCliente // Corriente-Fase-3-ConsumoCliente // Potencia-Fase-3-ConsumoCliente // Potencia-Fase-3-ConsumoCliente
#VoltajeMax-Fase-3-ConsumoCliente // VoltajePromedio-Fase-3-ConsumoCliente // VoltajeMin-Fase-3-ConsumoCliente // CorrienteMax-Fase-3-ConsumoCliente // PotenciaMax-Fase-3-ConsumoCliente // EnergiaMax-Fase-3-ConsumoCliente

def SendDataToBroker(q,k,f,**kwargs):
        
        def publish(client): 
            g=0
            global vt3,vt4,vt5,vt6,vt7,vt8,vt9,vt,vt15
            print(len(kwargs.values()))
            if(len(kwargs.values())==4):
                if(q==1):
                    vt = vt1 #0 // 10
                    #print(f'vt {vt}')
                elif(q==2):
                    vt = vt2
                elif(q==3):
                    vt = vt3
                elif(q==4):
                    vt = vt4
                elif(q==5):
                    vt = vt5
                elif(q==6):
                    vt = vt6
                elif(q==7):
                    vt = vt7
                elif(q==8):
                    vt = vt8
                elif(q==9):
                    vt = vt9
            elif(len(kwargs.values())<2):
                vt = vt15 #0 // 10
                #print(f'vt15 {vt}')
            elif(len(kwargs.values())>9):
                if(q==1):
                    vt = vt115 #0 // 10
                    #print(f'vt115 {vt}')
                elif(q==2):
                    vt = vt215
                elif(q==3):
                    vt = vt315
                elif(q==4):
                    vt = vt415
                elif(q==5):
                    vt = vt515
                elif(q==6):
                    vt = vt615
                elif(q==7):
                    vt = vt715
                elif(q==8):
                    vt = vt815
                elif(q==9):
                    vt = vt915
            timeToSend=time.time() #10 // 20
            #print(f'timetoSend: {round(timeToSend)}')
            #print(f'Largo Kwargs {len(kwargs.values())}')
            for key, value in kwargs.items():
                g=g+1
                #print(f'g = {g}')
                #print(f"Preparando Envio - {key}-{q} {value} {f} - {k}")
                str_num = {"value":value,"save":optionsave}
                valueJson = json.dumps(str_num)
                #print(f'{key}-{f}-{k}')
                for i in data["variables"]:
                    if(i["variableFullName"]==f'{key}-{f}-{k}'):
                        print(f"Preparando Envio en publish de variable {key}-{q}-{f}-{k}")
                        freq = i["variableSendFreq"]  
                        #print(f'Tiempo {round(timeToSend - vt)}')  #10-0=10 // 20-10=10 
                        if(timeToSend - vt > float(freq)): 
                             #print(f"Entrando a envio {key}-{q}")
                             str_variable = i["variable"]
                             topic = topicmqtt + str_variable + "/sdata"
                             result = client.publish(topic, valueJson)
                             status = result[0]            
                             if status == 0:
                                 print(f"Send {key}-{f}-{k} {value}")#`{valueJson}` to topic `{topic}` freq: {freq} to {key}-{q} ")  
                             else:
                                 print(f"Failed to send message to topic {topic}")       

        try:  
            if(client.connected_flag==True): 
                publish(client)
        except:
            pass

#    if(data["variables"][i]["variableType"]=="output"):
#        continue
        
    
    
    
    
    

#def HistoricalExcel():
#    today = date.today()
#    exceltime=date.today()
#    book = Workbook()
#    dest_filename = f'{exceltime}.xlsx'
#    print("El mes actual es {}".format(today.month))
#    for f in os.listdir('/home/pi/Desktop/Ser-Raspberry/'):
#            if():

    
            
MaxVoltage15_1=0.0
MeanVoltage15_1=0.0
MinVoltage15_1=0.0
MaxCurrent15_1=0.0
MeanCurrent15_1=0.0
MinCurrent15_1=0.0
MaxActivePower_1=0.0
MeanActivePower_1=0.0
MinActivePower_1=0.0
MaxReactivePower_1=0.0
MeanReactivePower_1=0.0
MinReactivePower_1=0.0
MaxAparentPower_1=0.0
MeanAparentPower_1=0.0
MinAparentPower_1=0.0
MaxFPInductive_1=-0.99
MeanFPInductive_1=-0.99
MinFPInductive_1=-0.99
MaxFPReactive_1=0.99
MeanFPReactive_1=0.99
MinFPReactive_1=0.99
MaxFD_1=0.0
MeanFD_1=0.0
MinFD_1=0.0
MaxDAT_1=0.0
MeanDAT_1=0.0
MinDAT_1=0.0
Access_1 = 0
Volt15_1=[]
data15_1=[]
Current15_1=[]
ActivePower15_1=[]
ReactivePower15_1=[]
AparentPower15_1=[]
FP15_Reactive_1=[]
FP15_Inductive_1=[]
FDVoltage15_1=[]
FDCurrent15_1=[]
DAT15Voltage_1=[]
DAT15Current_1=[]

def Maximo15min_1(Vrms,Irms,ActivePower,ReactivePower,AparentPower,FP,FDVoltage,FDCurrent,DATVoltage,DATCurrent,OneHourEnergy,Energy,i,k,f):
    global data15_1
    global Volt15_1
    global data15_1
    global Current15_1
    global ActivePower15_1
    global ReactivePower15_1
    global AparentPower15_1
    global FP15_Reactive_1
    global FP15_Inductive_1
    global FDVoltage15_1
    global FDCurrent15_1
    global DAT15Voltage_1
    global DAT15Current_1
    global Access_1
    #global OneHourEnergy_1
    global optionsave
    global vt115
    basea = datetime.datetime.now()
    if(basea.minute==0 or basea.minute==1 or basea.minute==2 or basea.minute==15 or basea.minute==16 or basea.minute==17 or basea.minute==30 or basea.minute==31 or basea.minute==32 or basea.minute==45 or basea.minute==46 or basea.minute==47): 
               if(Access_1 == 0):
                    #graphVoltage(NoVoltageoffset2,ListaIrmsPeak2,samplings2,2)
                    Access_1 = 1
                    MaxVoltage15_1=max(Volt15_1)
                    MeanVoltage15_1=np.median(Volt15_1)
                    MinVoltage15_1=min(Volt15_1)
                    MaxCurrent15_1=max(Current15_1)
                    MeanCurrent15_1=np.median(Current15_1)
                    MinCurrent15_1=min(Current15_1)
                    MaxActivePower_1=max(ActivePower15_1)
                    MeanActivePower_1=np.median(ActivePower15_1)
                    MinActivePower_1=min(ActivePower15_1)
                    MaxReactivePower_1=max(ReactivePower15_1)
                    MeanReactivePower_1=np.median(ReactivePower15_1)
                    MinReactivePower_1=min(ReactivePower15_1)
                    MaxAparentPower_1=max(AparentPower15_1)
                    MeanAparentPower_1=np.median(AparentPower15_1)
                    MinAparentPower_1=min(AparentPower15_1)
                    if(len(FP15_Inductive_1)>0):
                           MaxFPInductive_1=max(FP15_Inductive_1)
                           MeanFPInductive_1=np.median(FP15_Inductive_1)
                           MinFPInductive_1=min(FP15_Inductive_1)
                    else:
                           MaxFPInductive_1=-0.99
                           MeanFPInductive_1=-0.99
                           MinFPInductive_1=-0.99
                    if(len(FP15_Reactive_1)>0):
                           MaxFPReactive_1=max(FP15_Reactive_1)
                           MeanFPReactive_1=np.median(FP15_Reactive_1)
                           MinFPReactive_1=min(FP15_Reactive_1)
                    else:
                           MaxFPReactive_1=0.99
                           MeanFPReactive_1=0.99
                           MinFPReactive_1=0.99
                    MaxFDVoltage_1=max(FDVoltage15_1)
                    MeanFDVoltage_1=np.median(FDVoltage15_1)
                    MinFDVoltage_1=min(FDVoltage15_1)
                    MaxFDCurrent_1=max(FDCurrent15_1)
                    MeanFDCurrent_1=np.median(FDCurrent15_1)
                    MinFDCurrent_1=min(FDCurrent15_1)
                    MaxDATVoltage_1=max(DAT15Voltage_1)
                    MeanDATVoltage_1=np.median(DAT15Voltage_1)
                    MinDATVoltage_1=min(DAT15Voltage_1)
                    MaxDATCurrent_1=max(DAT15Current_1)
                    MeanDATCurrent_1=np.median(DAT15Current_1)
                    MinDATCurrent_1=min(DAT15Current_1)
                    data15_1.insert(1,MaxVoltage15_1)
                    data15_1.insert(2,MeanVoltage15_1)
                    data15_1.insert(3,MinVoltage15_1)
                    data15_1.insert(4,MaxCurrent15_1)
                    data15_1.insert(5,MeanCurrent15_1)
                    data15_1.insert(6,MinCurrent15_1)
                    data15_1.insert(7,MaxActivePower_1)
                    data15_1.insert(8,MeanActivePower_1)
                    data15_1.insert(9,MinActivePower_1)
                    data15_1.insert(10,MaxReactivePower_1)
                    data15_1.insert(11,MeanReactivePower_1)
                    data15_1.insert(12,MinReactivePower_1)
                    data15_1.insert(13,MaxAparentPower_1)
                    data15_1.insert(14,MeanAparentPower_1)
                    data15_1.insert(15,MinAparentPower_1)
                    data15_1.insert(16,MaxFPInductive_1)
                    data15_1.insert(17,MeanFPInductive_1)
                    data15_1.insert(18,MinFPInductive_1)
                    data15_1.insert(19,MaxFPReactive_1)
                    data15_1.insert(20,MeanFPReactive_1)
                    data15_1.insert(21,MinFPReactive_1)
                    data15_1.insert(22,MaxFDVoltage_1)
                    data15_1.insert(23,MeanFDVoltage_1)
                    data15_1.insert(24,MinFDVoltage_1)
                    data15_1.insert(25,MaxFDCurrent_1)
                    data15_1.insert(26,MeanFDCurrent_1)
                    data15_1.insert(27,MinFDCurrent_1)
                    data15_1.insert(28,MaxDATVoltage_1)
                    data15_1.insert(29,MeanDATVoltage_1)
                    data15_1.insert(30,MinDATVoltage_1)
                    data15_1.insert(31,MaxDATCurrent_1)
                    data15_1.insert(32,MeanDATCurrent_1)
                    data15_1.insert(33,MinDATCurrent_1)
                    data15_1.insert(34,OneHourEnergy)
                    data15_1.insert(35,Energy)
                    data15_1.insert(0,datetime.datetime.now())
                    workbook=openpyxl.load_workbook(filename = dest_filename)
                    sheet2 = workbook[f"Mean-{k}-{f}"]
                    sheet2.append(list(data15_1))
                    print(f'Data 1: Guardando Promedios')
                    optionsave=1
                    SendDataToBroker(q=i,k=k,f=f,VoltajeMax=f'{MaxVoltage15_1}',VoltajePromedio=f'{MeanVoltage15_1}',VoltajeMin=f'{MinVoltage15_1}',MaxCorriente=f'{MaxCurrent15_1}',PromedioCorriente=f'{MeanCurrent15_1}',MinimoCorriente=f'{MinCurrent15_1}',PotenciaMax=f'{MaxAparentPower_1}',PromedioPotenciaAparente=f'{MeanAparentPower_1}',MinPotenciaAparente=f'{MinAparentPower_1}',Energia=f'{Energy}')
                    #SendDataToBroker(MaxVoltage15_1,MeanVoltage15_1,MinVoltage15_1,MaxCurrent15_1,MeanCurrent15_1,MinCurrent15_1,MaxAparentPower_1,MeanAparentPower_1,MinAparentPower_1,OneHourEnergy,Energy,k,f,_)
                    vt115=time.time()
                    workbook.save(filename = dest_filename)
                    data15_1=[]
                    Volt15_1=[]
                    Current15_1=[]
                    ActivePower15_1=[]
                    ReactivePower15_1=[]
                    AparentPower15_1=[]
                    FP15_Reactive_1=[]
                    FP15_Inductive_1=[]
                    FDVoltage15_1=[]
                    FDCurrent15_1=[]
                    DAT15Voltage_1=[]
                    DAT15Current_1=[]
                    #OneHourEnergy_1=0
               elif(Access_1==1):
                    #print("paso elif 2")
                    Volt15_1.append(Vrms)
                    Current15_1.append(Irms)
                    ActivePower15_1.append(ActivePower)
                    ReactivePower15_1.append(ReactivePower)
                    AparentPower15_1.append(AparentPower)
                    if(FP>0.0):
                          FP15_Reactive_1.append(FP)
                    else: 
                          FP15_Inductive_1.append(FP)
                    FDVoltage15_1.append(FDVoltage)
                    FDCurrent15_1.append(FDCurrent)
                    DAT15Voltage_1.append(DATVoltage)
                    DAT15Current_1.append(DATCurrent)
              
    else:
        Volt15_1.append(Vrms)
        #print(Volt15_1)
        Current15_1.append(Irms)
        ActivePower15_1.append(ActivePower)
        ReactivePower15_1.append(ReactivePower)
        AparentPower15_1.append(AparentPower)
        if(FP>0.0):
              FP15_Reactive_1.append(FP)
        else: 
              FP15_Inductive_1.append(FP)
        FDVoltage15_1.append(FDVoltage)
        FDCurrent15_1.append(FDCurrent)
        DAT15Voltage_1.append(DATVoltage)
        DAT15Current_1.append(DATCurrent)
        Access_1 = 0
        
        if(len(Volt15_1)>4):
            Volt15_1.sort()
            indice=np.argmin(Volt15_1)
            Volt15_1.pop(indice+1)
            indice=np.argmax(Volt15_1)
            Volt15_1.pop(indice-1)
        if(len(Current15_1)>4):
            Current15_1.sort()
            indice=np.argmin(Current15_1)
            Current15_1.pop(indice+1)
            indice=np.argmax(Current15_1)
            Current15_1.pop(indice-1)
        if(len(ActivePower15_1)>4):
            ActivePower15_1.sort()
            indice=np.argmin(ActivePower15_1)
            ActivePower15_1.pop(indice+1)
            indice=np.argmax(ActivePower15_1)
            ActivePower15_1.pop(indice-1)
        if(len(ReactivePower15_1)>4):
            ReactivePower15_1.sort()
            indice=np.argmin(ReactivePower15_1)
            ReactivePower15_1.pop(indice+1)
            indice=np.argmax(ReactivePower15_1)
            ReactivePower15_1.pop(indice-1)
        if(len(AparentPower15_1)>4):
            AparentPower15_1.sort()
            indice=np.argmin(AparentPower15_1)
            AparentPower15_1.pop(indice+1)
            indice=np.argmax(AparentPower15_1)
            AparentPower15_1.pop(indice-1)
        if(len(FP15_Reactive_1)>4):
            FP15_Reactive_1.sort()
            indice=np.argmin(FP15_Reactive_1)
            FP15_Reactive_1.pop(indice+1)
            indice=np.argmax(FP15_Reactive_1)
            FP15_Reactive_1.pop(indice-1)
        if(len(FP15_Inductive_1)>4):
            FP15_Inductive_1.sort()
            indice=np.argmin(FP15_Inductive_1)
            FP15_Inductive_1.pop(indice+1)
            indice=np.argmax(FP15_Inductive_1)
            FP15_Inductive_1.pop(indice-1)
        if(len(FDVoltage15_1)>4):
            FDVoltage15_1.sort()
            indice=np.argmin(FDVoltage15_1)
            FDVoltage15_1.pop(indice+1)
            indice=np.argmax(FDVoltage15_1)
            FDVoltage15_1.pop(indice-1)
        if(len(FDCurrent15_1)>4):
            FDCurrent15_1.sort()
            indice=np.argmin(FDCurrent15_1)
            FDCurrent15_1.pop(indice+1)
            indice=np.argmax(FDCurrent15_1)
            FDCurrent15_1.pop(indice-1)
        if(len(DAT15Voltage_1)>4):
            DAT15Voltage_1.sort()
            indice=np.argmin(DAT15Voltage_1)
            DAT15Voltage_1.pop(indice+1)
            indice=np.argmax(DAT15Voltage_1)
            DAT15Voltage_1.pop(indice-1)
        if(len(DAT15Current_1)>4):
            DAT15Current_1.sort()
            indice=np.argmin(DAT15Current_1)
            DAT15Current_1.pop(indice+1)
            indice=np.argmax(DAT15Current_1)
            DAT15Current_1.pop(indice-1)
        
        
Access_2 = 0
MaxVoltage15_2=0.0
MeanVoltage15_2=0.0
MinVoltage15_2=0.0
MaxCurrent15_2=0.0
MeanCurrent15_2=0.0
MinCurrent15_2=0.0
MaxActivePower_2=0.0
MeanActivePower_2=0.0
MinActivePower_2=0.0
MaxReactivePower_2=0.0
MeanReactivePower_2=0.0
MinReactivePower_2=0.0
MaxAparentPower_2=0.0
MeanAparentPower_2=0.0
MinAparentPower_2=0.0
MaxFPInductive_2=-0.99
MeanFPInductive_2=-0.99
MinFPInductive_2=-0.99
MaxFPReactive_2=0.99
MeanFPReactive_2=0.99
MinFPReactive_2=0.99
MaxFD_2=0.0
MeanFD_2=0.0
MinFD_2=0.0
MaxDAT_2=0.0
MeanDAT_2=0.0
MinDAT_2=0.0
Volt15_2=[]
data15_2=[]
Current15_2=[]
ActivePower15_2=[]
ReactivePower15_2=[]
AparentPower15_2=[]
FP15_Reactive_2=[]
FP15_Inductive_2=[]
FDVoltage15_2=[]
FDCurrent15_2=[]
DAT15Voltage_2=[]
DAT15Current_2=[]
def Maximo15min_2(Vrms,Irms,ActivePower,ReactivePower,AparentPower,FP,FDVoltage,FDCurrent,DATVoltage,DATCurrent,OneHourEnergy,Energy,i,k,f):
    global data15_2
    global Volt15_2
    global data15_2
    global Current15_2
    global ActivePower15_2
    global ReactivePower15_2
    global AparentPower15_2
    global FP15_Reactive_2
    global FP15_Inductive_2
    global FDVoltage15_2
    global FDCurrent15_2
    global DAT15Voltage_2
    global DAT15Current_2
    global Access_2
    #global OneHourEnergy_2
    global optionsave
    global vt215
    basea = datetime.datetime.now()
    if(basea.minute==0 or basea.minute==1 or basea.minute==2 or basea.minute==15 or basea.minute==16 or basea.minute==17 or basea.minute==30 or basea.minute==31 or basea.minute==32 or basea.minute==45 or basea.minute==46 or basea.minute==47):
               if(Access_2 == 0):
                    #graphVoltage(NoVoltageoffset2,ListaIrmsPeak2,samplings2,2)
                    Access_2 = 1
                    MaxVoltage15_2=max(Volt15_2)
                    MeanVoltage15_2=np.median(Volt15_2)
                    MinVoltage15_2=min(Volt15_2)
                    MaxCurrent15_2=max(Current15_2)
                    MeanCurrent15_2=np.median(Current15_2)
                    MinCurrent15_2=min(Current15_2)
                    MaxActivePower_2=max(ActivePower15_2)
                    MeanActivePower_2=np.median(ActivePower15_2)
                    MinActivePower_2=min(ActivePower15_2)
                    MaxReactivePower_2=max(ReactivePower15_2)
                    MeanReactivePower_2=np.median(ReactivePower15_2)
                    MinReactivePower_2=min(ReactivePower15_2)
                    MaxAparentPower_2=max(AparentPower15_2)
                    MeanAparentPower_2=np.median(AparentPower15_2)
                    MinAparentPower_2=min(AparentPower15_2)
                    if(len(FP15_Inductive_2)>0):
                           MaxFPInductive_2=max(FP15_Inductive_2)
                           MeanFPInductive_2=np.median(FP15_Inductive_2)
                           MinFPInductive_2=min(FP15_Inductive_2)
                    else:
                           MaxFPInductive_2=-0.99
                           MeanFPInductive_2=-0.99
                           MinFPInductive_2=-0.99
                    if(len(FP15_Reactive_2)>0):
                           MaxFPReactive_2=max(FP15_Reactive_2)
                           MeanFPReactive_2=np.median(FP15_Reactive_2)
                           MinFPReactive_2=min(FP15_Reactive_2)
                    else:
                           MaxFPReactive_2=0.99
                           MeanFPReactive_2=0.99
                           MinFPReactive_2=0.99
                    MaxFDVoltage_2=max(FDVoltage15_2)
                    MeanFDVoltage_2=np.median(FDVoltage15_2)
                    MinFDVoltage_2=min(FDVoltage15_2)
                    MaxFDCurrent_2=max(FDCurrent15_2)
                    MeanFDCurrent_2=np.median(FDCurrent15_2)
                    MinFDCurrent_2=min(FDCurrent15_2)
                    MaxDATVoltage_2=max(DAT15Voltage_2)
                    MeanDATVoltage_2=np.median(DAT15Voltage_2)
                    MinDATVoltage_2=min(DAT15Voltage_2)
                    MaxDATCurrent_2=max(DAT15Current_2)
                    MeanDATCurrent_2=np.median(DAT15Current_2)
                    MinDATCurrent_2=min(DAT15Current_2)
                    data15_2.insert(1,MaxVoltage15_2)
                    data15_2.insert(2,MeanVoltage15_2)
                    data15_2.insert(3,MinVoltage15_2)
                    data15_2.insert(4,MaxCurrent15_2)
                    data15_2.insert(5,MeanCurrent15_2)
                    data15_2.insert(6,MinCurrent15_2)
                    data15_2.insert(7,MaxActivePower_2)
                    data15_2.insert(8,MeanActivePower_2)
                    data15_2.insert(9,MinActivePower_2)
                    data15_2.insert(10,MaxReactivePower_2)
                    data15_2.insert(11,MeanReactivePower_2)
                    data15_2.insert(12,MinReactivePower_2)
                    data15_2.insert(13,MaxAparentPower_2)
                    data15_2.insert(14,MeanAparentPower_2)
                    data15_2.insert(15,MinAparentPower_2)
                    data15_2.insert(16,MaxFPInductive_2)
                    data15_2.insert(17,MeanFPInductive_2)
                    data15_2.insert(18,MinFPInductive_2)
                    data15_2.insert(19,MaxFPReactive_2)
                    data15_2.insert(20,MeanFPReactive_2)
                    data15_2.insert(21,MinFPReactive_2)
                    data15_2.insert(22,MaxFDVoltage_2)
                    data15_2.insert(23,MeanFDVoltage_2)
                    data15_2.insert(24,MinFDVoltage_2)
                    data15_2.insert(25,MaxFDCurrent_2)
                    data15_2.insert(26,MeanFDCurrent_2)
                    data15_2.insert(27,MinFDCurrent_2)
                    data15_2.insert(28,MaxDATVoltage_2)
                    data15_2.insert(29,MeanDATVoltage_2)
                    data15_2.insert(30,MinDATVoltage_2)
                    data15_2.insert(31,MaxDATCurrent_2)
                    data15_2.insert(32,MeanDATCurrent_2)
                    data15_2.insert(33,MinDATCurrent_2)
                    data15_1.insert(34,OneHourEnergy)
                    data15_2.insert(35,Energy)
                    data15_2.insert(0,datetime.datetime.now())
                    workbook=openpyxl.load_workbook(filename = dest_filename)
                    sheet3 = workbook[f"Mean-{k}-{f}"]
                    sheet3.append(list(data15_2))
                    print(f'Data 2: Guardando Promedios')
                    optionsave=1
                    SendDataToBroker(q=i,k=k,f=f,VoltajeMax=f'{MaxVoltage15_2}',VoltajePromedio=f'{MeanVoltage15_2}',VoltajeMin=f'{MinVoltage15_2}',MaxCorriente=f'{MaxCurrent15_2}',PromedioCorriente=f'{MeanCurrent15_2}',MinimoCorriente=f'{MinCurrent15_2}',PotenciaMax=f'{MaxAparentPower_2}',PromedioPotenciaAparente=f'{MeanAparentPower_2}',MinPotenciaAparente=f'{MinAparentPower_2}',Energia=f'{Energy}')
                    vt215=time.time()
                    #print("Datos Insertados Correctamente!")
                    workbook.save(filename = dest_filename)
                    data15_2=[]
                    Volt15_2=[]
                    Current15_2=[]
                    ActivePower15_2=[]
                    ReactivePower15_2=[]
                    AparentPower15_2=[]
                    FP15_Reactive_2=[]
                    FP15_Inductive_2=[]
                    FDVoltage15_2=[]
                    FDCurrent15_2=[]
                    DAT15Voltage_2=[]
                    DAT15Current_2=[]
                    #OneHourEnergy_2=0
               elif(Access_2==1):
                    #print("paso elif 2")
                    Volt15_2.append(Vrms)
                    Current15_2.append(Irms)
                    ActivePower15_2.append(ActivePower)
                    ReactivePower15_2.append(ReactivePower)
                    AparentPower15_2.append(AparentPower)
                    if(FP>0.0):
                          FP15_Reactive_2.append(FP)
                    else: 
                          FP15_Inductive_2.append(FP)
                    FDVoltage15_2.append(FDVoltage)
                    FDCurrent15_2.append(FDCurrent)
                    DAT15Voltage_2.append(DATVoltage)
                    DAT15Current_2.append(DATCurrent)
              
    else:
        Volt15_2.append(Vrms)
        Current15_2.append(Irms)
        ActivePower15_2.append(ActivePower)
        ReactivePower15_2.append(ReactivePower)
        AparentPower15_2.append(AparentPower)
        if(FP>0.0):
              FP15_Reactive_2.append(FP)
        else: 
              FP15_Inductive_2.append(FP)
        FDVoltage15_2.append(FDVoltage)
        FDCurrent15_2.append(FDCurrent)
        DAT15Voltage_2.append(DATVoltage)
        DAT15Current_2.append(DATCurrent)
        Access_2 = 0
    
        if(len(Volt15_2)>4):
            Volt15_2.sort()
            indice=np.argmin(Volt15_2)
            Volt15_2.pop(indice+1)
            indice=np.argmax(Volt15_2)
            Volt15_2.pop(indice-1)
        if(len(Current15_2)>4):
            Current15_2.sort()
            indice=np.argmin(Current15_2)
            Current15_2.pop(indice+1)
            indice=np.argmax(Current15_2)
            Current15_2.pop(indice-1)
        if(len(ActivePower15_2)>4):
            ActivePower15_2.sort()
            indice=np.argmin(ActivePower15_2)
            ActivePower15_2.pop(indice+1)
            indice=np.argmax(ActivePower15_2)
            ActivePower15_2.pop(indice-1)
        if(len(ReactivePower15_2)>4):
            ReactivePower15_2.sort()
            indice=np.argmin(ReactivePower15_2)
            ReactivePower15_2.pop(indice+1)
            indice=np.argmax(ReactivePower15_2)
            ReactivePower15_2.pop(indice-1)
        if(len(AparentPower15_2)>4):
            AparentPower15_2.sort()
            indice=np.argmin(AparentPower15_2)
            AparentPower15_2.pop(indice+1)
            indice=np.argmax(AparentPower15_2)
            AparentPower15_2.pop(indice-1)
        if(len(FP15_Reactive_2)>4):
            FP15_Reactive_2.sort()
            indice=np.argmin(FP15_Reactive_2)
            FP15_Reactive_2.pop(indice+1)
            indice=np.argmax(FP15_Reactive_2)
            FP15_Reactive_2.pop(indice-1)
        if(len(FP15_Inductive_2)>4):
            FP15_Inductive_2.sort()
            indice=np.argmin(FP15_Inductive_2)
            FP15_Inductive_2.pop(indice+1)
            indice=np.argmax(FP15_Inductive_2)
            FP15_Inductive_2.pop(indice-1)
        if(len(FDVoltage15_2)>4):
            FDVoltage15_2.sort()
            indice=np.argmin(FDVoltage15_2)
            FDVoltage15_2.pop(indice+1)
            indice=np.argmax(FDVoltage15_2)
            FDVoltage15_2.pop(indice-1)
        if(len(FDCurrent15_2)>4):
            FDCurrent15_2.sort()
            indice=np.argmin(FDCurrent15_2)
            FDCurrent15_2.pop(indice+1)
            indice=np.argmax(FDCurrent15_2)
            FDCurrent15_2.pop(indice-1)
        if(len(DAT15Voltage_2)>4):
            DAT15Voltage_2.sort()
            indice=np.argmin(DAT15Voltage_2)
            DAT15Voltage_2.pop(indice+1)
            indice=np.argmax(DAT15Voltage_2)
            DAT15Voltage_2.pop(indice-1)
        if(len(DAT15Current_2)>4):
            DAT15Current_2.sort()
            indice=np.argmin(DAT15Current_2)
            DAT15Current_2.pop(indice+1)
            indice=np.argmax(DAT15Current_2)
            DAT15Current_2.pop(indice-1)
    
Access_3 = 0
MaxVoltage15_3=0.0
MeanVoltage15_3=0.0
MinVoltage15_3=0.0
MaxCurrent15_3=0.0
MeanCurrent15_3=0.0
MinCurrent15_3=0.0
MaxActivePower_3=0.0
MeanActivePower_3=0.0
MinActivePower_3=0.0
MaxReactivePower_3=0.0
MeanReactivePower_3=0.0
MinReactivePower_3=0.0
MaxAparentPower_3=0.0
MeanAparentPower_3=0.0
MinAparentPower_3=0.0
MaxFPInductive_3=-0.99
MeanFPInductive_3=-0.99
MinFPInductive_3=-0.99
MaxFPReactive_3=0.99
MeanFPReactive_3=0.99
MinFPReactive_3=0.99
MaxFD_3=0.0
MeanFD_3=0.0
MinFD_3=0.0
MaxDAT_3=0.0
MeanDAT_3=0.0
MinDAT_3=0.0
Volt15_3=[]
data15_3=[]
Current15_3=[]
ActivePower15_3=[]
ReactivePower15_3=[]
AparentPower15_3=[]
FP15_Reactive_3=[]
FP15_Inductive_3=[]
FDVoltage15_3=[]
FDCurrent15_3=[]
DAT15Voltage_3=[]
DAT15Current_3=[]
def Maximo15min_3(Vrms,Irms,ActivePower,ReactivePower,AparentPower,FP,FDVoltage,FDCurrent,DATVoltage,DATCurrent,OneHourEnergy,Energy,i,k,f):
    global data15_3
    global Volt15_3
    global data15_3
    global Current15_3
    global ActivePower15_3
    global ReactivePower15_3
    global AparentPower15_3
    global FP15_Reactive_3
    global FP15_Inductive_3
    global FDVoltage15_3
    global FDCurrent15_3
    global DAT15Voltage_3
    global DAT15Current_3
    global Access_3
    #global OneHourEnergy_3
    global optionsave
    basea = datetime.datetime.now()
    if(basea.minute==0 or basea.minute==1 or basea.minute==2 or basea.minute==15 or basea.minute==16 or basea.minute==17 or basea.minute==30 or basea.minute==31 or basea.minute==32 or basea.minute==45 or basea.minute==46 or basea.minute==47):
               if(Access_3 == 0):
                    #graphVoltage(NoVoltageoffset2,ListaIrmsPeak2,samplings2,2)
                    Access_3 = 1
                    MaxVoltage15_3=max(Volt15_3)
                    MeanVoltage15_3=np.median(Volt15_3)
                    MinVoltage15_3=min(Volt15_3)
                    MaxCurrent15_3=max(Current15_3)
                    MeanCurrent15_3=np.median(Current15_3)
                    MinCurrent15_3=min(Current15_3)
                    MaxActivePower_3=max(ActivePower15_3)
                    MeanActivePower_3=np.median(ActivePower15_3)
                    MinActivePower_3=min(ActivePower15_3)
                    MaxReactivePower_3=max(ReactivePower15_3)
                    MeanReactivePower_3=np.median(ReactivePower15_3)
                    MinReactivePower_3=min(ReactivePower15_3)
                    MaxAparentPower_3=max(AparentPower15_3)
                    MeanAparentPower_3=np.median(AparentPower15_3)
                    MinAparentPower_3=min(AparentPower15_3)
                    if(len(FP15_Inductive_3)>0):
                           MaxFPInductive_3=max(FP15_Inductive_3)
                           MeanFPInductive_3=np.median(FP15_Inductive_3)
                           MinFPInductive_3=min(FP15_Inductive_3)
                    else:
                           MaxFPInductive_3=-0.99
                           MeanFPInductive_3=-0.99
                           MinFPInductive_3=-0.99
                    if(len(FP15_Reactive_3)>0):
                           MaxFPReactive_3=max(FP15_Reactive_3)
                           MeanFPReactive_3=np.median(FP15_Reactive_3)
                           MinFPReactive_3=min(FP15_Reactive_3)
                    else:
                           MaxFPReactive_3=0.99
                           MeanFPReactive_3=0.99
                           MinFPReactive_3=0.99
                    MaxFDVoltage_3=max(FDVoltage15_3)
                    MeanFDVoltage_3=np.median(FDVoltage15_3)
                    MinFDVoltage_3=min(FDVoltage15_3)
                    MaxFDCurrent_3=max(FDCurrent15_3)
                    MeanFDCurrent_3=np.median(FDCurrent15_3)
                    MinFDCurrent_3=min(FDCurrent15_3)
                    MaxDATVoltage_3=max(DAT15Voltage_3)
                    MeanDATVoltage_3=np.median(DAT15Voltage_3)
                    MinDATVoltage_3=min(DAT15Voltage_3)
                    MaxDATCurrent_3=max(DAT15Current_3)
                    MeanDATCurrent_3=np.median(DAT15Current_3)
                    MinDATCurrent_3=min(DAT15Current_3)
                    data15_3.insert(1,MaxVoltage15_3)
                    data15_3.insert(2,MeanVoltage15_3)
                    data15_3.insert(3,MinVoltage15_3)
                    data15_3.insert(4,MaxCurrent15_3)
                    data15_3.insert(5,MeanCurrent15_3)
                    data15_3.insert(6,MinCurrent15_3)
                    data15_3.insert(7,MaxActivePower_3)
                    data15_3.insert(8,MeanActivePower_3)
                    data15_3.insert(9,MinActivePower_3)
                    data15_3.insert(10,MaxReactivePower_3)
                    data15_3.insert(11,MeanReactivePower_3)
                    data15_3.insert(12,MinReactivePower_3)
                    data15_3.insert(13,MaxAparentPower_3)
                    data15_3.insert(14,MeanAparentPower_3)
                    data15_3.insert(15,MinAparentPower_3)
                    data15_3.insert(16,MaxFPInductive_3)
                    data15_3.insert(17,MeanFPInductive_3)
                    data15_3.insert(18,MinFPInductive_3)
                    data15_3.insert(19,MaxFPReactive_3)
                    data15_3.insert(20,MeanFPReactive_3)
                    data15_3.insert(21,MinFPReactive_3)
                    data15_3.insert(22,MaxFDVoltage_3)
                    data15_3.insert(23,MeanFDVoltage_3)
                    data15_3.insert(24,MinFDVoltage_3)
                    data15_3.insert(25,MaxFDCurrent_3)
                    data15_3.insert(26,MeanFDCurrent_3)
                    data15_3.insert(27,MinFDCurrent_3)
                    data15_3.insert(28,MaxDATVoltage_3)
                    data15_3.insert(29,MeanDATVoltage_3)
                    data15_3.insert(30,MinDATVoltage_3)
                    data15_3.insert(31,MaxDATCurrent_3)
                    data15_3.insert(32,MeanDATCurrent_3)
                    data15_3.insert(33,MinDATCurrent_3)
                    data15_3.insert(34,OneHourEnergy)
                    data15_3.insert(35,Energy)
                    data15_3.insert(0,datetime.datetime.now())
                    workbook=openpyxl.load_workbook(filename = dest_filename)
                    sheet4 = workbook[f"Mean-{k}-{f}"]
                    sheet4.append(list(data15_3))
                    print(f'Data 3: Guardando Promedios')
                    optionsave=1
                    SendDataToBroker(q=i,k=k,f=f,VoltajeMax=f'{MaxVoltage15_3}',VoltajePromedio=f'{MeanVoltage15_3}',VoltajeMin=f'{MinVoltage15_3}',MaxCorriente=f'{MaxCurrent15_3}',PromedioCorriente=f'{MeanCurrent15_3}',MinimoCorriente=f'{MinCurrent15_3}',PotenciaMax=f'{MaxAparentPower_3}',PromedioPotenciaAparente=f'{MeanAparentPower_3}',MinPotenciaAparente=f'{MinAparentPower_3}',Energia=f'{Energy}')
                    #print("Datos Insertados Correctamente!")
                    workbook.save(filename = dest_filename)
                    data15_3=[]
                    Volt15_3=[]
                    Current15_3=[]
                    ActivePower15_3=[]
                    ReactivePower15_3=[]
                    AparentPower15_3=[]
                    FP15_Reactive_3=[]
                    FP15_Inductive_3=[]
                    FDVoltage15_3=[]
                    FDCurrent15_3=[]
                    DAT15Voltage_3=[]
                    DAT15Current_3=[]
                    #OneHourEnergy_3=0
               elif(Access_3==1):
                    #print("paso elif 2")
                    Volt15_3.append(Vrms)
                    Current15_3.append(Irms)
                    ActivePower15_3.append(ActivePower)
                    ReactivePower15_3.append(ReactivePower)
                    AparentPower15_3.append(AparentPower)
                    if(FP>0.0):
                          FP15_Reactive_3.append(FP)
                    else: 
                          FP15_Inductive_3.append(FP)
                    FDVoltage15_3.append(FDVoltage)
                    FDCurrent15_3.append(FDCurrent)
                    DAT15Voltage_3.append(DATVoltage)
                    DAT15Current_3.append(DATCurrent)
              
    else:
        Volt15_3.append(Vrms)
        Current15_3.append(Irms)
        ActivePower15_3.append(ActivePower)
        ReactivePower15_3.append(ReactivePower)
        AparentPower15_3.append(AparentPower)
        if(FP>0.0):
              FP15_Reactive_3.append(FP)
        else: 
              FP15_Inductive_3.append(FP)
        FDVoltage15_3.append(FDVoltage)
        FDCurrent15_3.append(FDCurrent)
        DAT15Voltage_3.append(DATVoltage)
        DAT15Current_3.append(DATCurrent)
        Access_3 = 0
        
        if(len(Volt15_3)>4):
            Volt15_3.sort()
            indice=np.argmin(Volt15_3)
            Volt15_3.pop(indice+1)    
            indice=np.argmax(Volt15_3)
            Volt15_3.pop(indice-1)
        if(len(Current15_3)>4):
            Current15_3.sort()
            indice=np.argmin(Current15_3)
            Current15_3.pop(indice+1)
            indice=np.argmax(Current15_3)
            Current15_3.pop(indice-1)
        if(len(ActivePower15_3)>4):
            ActivePower15_3.sort()
            indice=np.argmin(ActivePower15_3)
            ActivePower15_3.pop(indice+1)
            indice=np.argmax(ActivePower15_3)
            ActivePower15_3.pop(indice-1)
        if(len(ReactivePower15_3)>4):
            ReactivePower15_3.sort()
            indice=np.argmin(ReactivePower15_3)
            ReactivePower15_3.pop(indice+1)
            indice=np.argmax(ReactivePower15_3)
            ReactivePower15_3.pop(indice-1)
        if(len(AparentPower15_3)>4):
            AparentPower15_3.sort()
            indice=np.argmin(AparentPower15_3)
            AparentPower15_3.pop(indice+1)
            indice=np.argmax(AparentPower15_3)
            AparentPower15_3.pop(indice-1)
        if(len(FP15_Reactive_3)>4):
            FP15_Reactive_3.sort()
            indice=np.argmin(FP15_Reactive_3)
            FP15_Reactive_3.pop(indice+1)
            indice=np.argmax(FP15_Reactive_3)
            FP15_Reactive_3.pop(indice-1)
        if(len(FP15_Inductive_3)>4):
            FP15_Inductive_3.sort()
            indice=np.argmin(FP15_Inductive_3)
            FP15_Inductive_3.pop(indice+1)
            indice=np.argmax(FP15_Inductive_3)
            FP15_Inductive_3.pop(indice-1)
        if(len(FDVoltage15_3)>4):
            FDVoltage15_3.sort()
            indice=np.argmin(FDVoltage15_3)
            FDVoltage15_3.pop(indice+1)
            indice=np.argmax(FDVoltage15_3)
            FDVoltage15_3.pop(indice-1)
        if(len(FDCurrent15_3)>4):
            FDCurrent15_3.sort()
            indice=np.argmin(FDCurrent15_3)
            FDCurrent15_3.pop(indice+1)
            indice=np.argmax(FDCurrent15_3)
            FDCurrent15_3.pop(indice-1)
        if(len(DAT15Voltage_3)>4):
            DAT15Voltage_3.sort()
            indice=np.argmin(DAT15Voltage_3)
            DAT15Voltage_3.pop(indice+1)
            indice=np.argmax(DAT15Voltage_3)
            DAT15Voltage_3.pop(indice-1)
        if(len(DAT15Current_3)>4):
            DAT15Current_3.sort()
            indice=np.argmin(DAT15Current_3)
            DAT15Current_3.pop(indice+1)
            indice=np.argmax(DAT15Current_3)
            DAT15Current_3.pop(indice-1)

Access_4 = 0
MaxVoltage15_4=0.0
MeanVoltage15_4=0.0
MinVoltage15_4=0.0
MaxCurrent15_4=0.0
MeanCurrent15_4=0.0
MinCurrent15_4=0.0
MaxActivePower_4=0.0
MeanActivePower_4=0.0
MinActivePower_4=0.0
MaxReactivePower_4=0.0
MeanReactivePower_4=0.0
MinReactivePower_4=0.0
MaxAparentPower_4=0.0
MeanAparentPower_4=0.0
MinAparentPower_4=0.0
MaxFPInductive_4=-0.99
MeanFPInductive_4=-0.99
MinFPInductive_4=-0.99
MaxFPReactive_4=0.99
MeanFPReactive_4=0.99
MinFPReactive_4=0.99
MaxFD_4=0.0
MeanFD_4=0.0
MinFD_4=0.0
MaxDAT_4=0.0
MeanDAT_4=0.0
MinDAT_4=0.0
Volt15_4=[]
data15_4=[]
Current15_4=[]
ActivePower15_4=[]
ReactivePower15_4=[]
AparentPower15_4=[]
FP15_Reactive_4=[]
FP15_Inductive_4=[]
FDVoltage15_4=[]
FDCurrent15_4=[]
DAT15Voltage_4=[]
DAT15Current_4=[]
def Maximo15min_4(Vrms,Irms,ActivePower,ReactivePower,AparentPower,FP,FDVoltage,FDCurrent,DATVoltage,DATCurrent,OneHourEnergy,Energy,i,k,f):
    global data15_4
    global Volt15_4
    global data15_4
    global Current15_4
    global ActivePower15_4
    global ReactivePower15_4
    global AparentPower15_4
    global FP15_Reactive_4
    global FP15_Inductive_4
    global FDVoltage15_4
    global FDCurrent15_4
    global DAT15Voltage_4
    global DAT15Current_4
    global Access_4
    global OneHourEnergy_4
    global optionsave
    basea = datetime.datetime.now()
    if(basea.minute==0 or basea.minute==1 or basea.minute==2 or basea.minute==15 or basea.minute==16 or basea.minute==17 or basea.minute==30 or basea.minute==31 or basea.minute==32 or basea.minute==45 or basea.minute==46 or basea.minute==47):
               if(Access_4 == 0):
                    #graphVoltage(NoVoltageoffset2,ListaIrmsPeak2,samplings2,2)
                    Access_4 = 1
                    MaxVoltage15_4=max(Volt15_4)
                    MeanVoltage15_4=np.median(Volt15_4)
                    MinVoltage15_4=min(Volt15_4)
                    MaxCurrent15_4=max(Current15_4)
                    MeanCurrent15_4=np.median(Current15_4)
                    MinCurrent15_4=min(Current15_4)
                    MaxActivePower_4=max(ActivePower15_4)
                    MeanActivePower_4=np.median(ActivePower15_4)
                    MinActivePower_4=min(ActivePower15_4)
                    MaxReactivePower_4=max(ReactivePower15_4)
                    MeanReactivePower_4=np.median(ReactivePower15_4)
                    MinReactivePower_4=min(ReactivePower15_4)
                    MaxAparentPower_4=max(AparentPower15_4)
                    MeanAparentPower_4=np.median(AparentPower15_4)
                    MinAparentPower_4=min(AparentPower15_4)
                    if(len(FP15_Inductive_4)>0):
                           MaxFPInductive_4=max(FP15_Inductive_4)
                           MeanFPInductive_4=np.median(FP15_Inductive_4)
                           MinFPInductive_4=min(FP15_Inductive_4)
                    else:
                           MaxFPInductive_4=-0.99
                           MeanFPInductive_4=-0.99
                           MinFPInductive_4=-0.99
                    if(len(FP15_Reactive_4)>0):
                           MaxFPReactive_4=max(FP15_Reactive_4)
                           MeanFPReactive_4=np.median(FP15_Reactive_4)
                           MinFPReactive_4=min(FP15_Reactive_4)
                    else:
                           MaxFPReactive_4=0.99
                           MeanFPReactive_4=0.99
                           MinFPReactive_4=0.99
                    MaxFDVoltage_4=max(FDVoltage15_4)
                    MeanFDVoltage_4=np.median(FDVoltage15_4)
                    MinFDVoltage_4=min(FDVoltage15_4)
                    MaxFDCurrent_4=max(FDCurrent15_4)
                    MeanFDCurrent_4=np.median(FDCurrent15_4)
                    MinFDCurrent_4=min(FDCurrent15_4)
                    MaxDATVoltage_4=max(DAT15Voltage_4)
                    MeanDATVoltage_4=np.median(DAT15Voltage_4)
                    MinDATVoltage_4=min(DAT15Voltage_4)
                    MaxDATCurrent_4=max(DAT15Current_4)
                    MeanDATCurrent_4=np.median(DAT15Current_4)
                    MinDATCurrent_4=min(DAT15Current_4)
                    data15_4.insert(1,MaxVoltage15_4)
                    data15_4.insert(2,MeanVoltage15_4)
                    data15_4.insert(3,MinVoltage15_4)
                    data15_4.insert(4,MaxCurrent15_4)
                    data15_4.insert(5,MeanCurrent15_4)
                    data15_4.insert(6,MinCurrent15_4)
                    data15_4.insert(7,MaxActivePower_4)
                    data15_4.insert(8,MeanActivePower_4)
                    data15_4.insert(9,MinActivePower_4)
                    data15_4.insert(10,MaxReactivePower_4)
                    data15_4.insert(11,MeanReactivePower_4)
                    data15_4.insert(12,MinReactivePower_4)
                    data15_4.insert(13,MaxAparentPower_4)
                    data15_4.insert(14,MeanAparentPower_4)
                    data15_4.insert(15,MinAparentPower_4)
                    data15_4.insert(16,MaxFPInductive_4)
                    data15_4.insert(17,MeanFPInductive_4)
                    data15_4.insert(18,MinFPInductive_4)
                    data15_4.insert(19,MaxFPReactive_4)
                    data15_4.insert(20,MeanFPReactive_4)
                    data15_4.insert(21,MinFPReactive_4)
                    data15_4.insert(22,MaxFDVoltage_4)
                    data15_4.insert(23,MeanFDVoltage_4)
                    data15_4.insert(24,MinFDVoltage_4)
                    data15_4.insert(25,MaxFDCurrent_4)
                    data15_4.insert(26,MeanFDCurrent_4)
                    data15_4.insert(27,MinFDCurrent_4)
                    data15_4.insert(28,MaxDATVoltage_4)
                    data15_4.insert(29,MeanDATVoltage_4)
                    data15_4.insert(30,MinDATVoltage_4)
                    data15_4.insert(31,MaxDATCurrent_4)
                    data15_4.insert(32,MeanDATCurrent_4)
                    data15_4.insert(33,MinDATCurrent_4)
                    data15_4.insert(34,OneHourEnergy)
                    data15_4.insert(35,Energy)
                    data15_4.insert(0,datetime.datetime.now())
                    workbook=openpyxl.load_workbook(filename = dest_filename)
                    sheet5 = workbook[f"Mean-{k}-{f}"]
                    sheet5.append(list(data15_4))
                    print(f'Data 4: Guardando Promedios')
                    optionsave=1
                    SendDataToBroker(q=i,k=k,f=f,VoltajeMax=f'{MaxVoltage15_4}',VoltajePromedio=f'{MeanVoltage15_4}',VoltajeMin=f'{MinVoltage15_4}',MaxCorriente=f'{MaxCurrent15_4}',PromedioCorriente=f'{MeanCurrent15_4}',MinimoCorriente=f'{MinCurrent15_4}',PotenciaMax=f'{MaxAparentPower_4}',PromedioPotenciaAparente=f'{MeanAparentPower_4}',MinPotenciaAparente=f'{MinAparentPower_4}',OneHourEnergy=f'{OneHourEnergy}',Energia=f'{Energy}')
                    #print("Datos Insertados Correctamente!")
                    workbook.save(filename = dest_filename)
                    data15_4=[]
                    Volt15_4=[]
                    Current15_4=[]
                    ActivePower15_4=[]
                    ReactivePower15_4=[]
                    AparentPower15_4=[]
                    FP15_Reactive_4=[]
                    FP15_Inductive_4=[]
                    FDVoltage15_4=[]
                    FDCurrent15_4=[]
                    DAT15Voltage_4=[]
                    DAT15Current_4=[]
                    #OneHourEnergy_4=0
               elif(Access_4==1):
                    #print("paso elif 2")
                    Volt15_4.append(Vrms)
                    Current15_4.append(Irms)
                    ActivePower15_4.append(ActivePower)
                    ReactivePower15_4.append(ReactivePower)
                    AparentPower15_4.append(AparentPower)
                    if(FP>0.0):
                          FP15_Reactive_4.append(FP)
                    else: 
                          FP15_Inductive_4.append(FP)
                    FDVoltage15_4.append(FDVoltage)
                    FDCurrent15_4.append(FDCurrent)
                    DAT15Voltage_4.append(DATVoltage)
                    DAT15Current_4.append(DATCurrent)      
    else:
        Volt15_4.append(Vrms)
        Current15_4.append(Irms)
        ActivePower15_4.append(ActivePower)
        ReactivePower15_4.append(ReactivePower)
        AparentPower15_4.append(AparentPower)
        if(FP>0.0):
              FP15_Reactive_4.append(FP)
        else: 
              FP15_Inductive_4.append(FP)
        FDVoltage15_4.append(FDVoltage)
        FDCurrent15_4.append(FDCurrent)
        DAT15Voltage_4.append(DATVoltage)
        DAT15Current_4.append(DATCurrent)
        Access_4 = 0
        
        if(len(Volt15_4)>4):
            Volt15_4.sort()
            indice=np.argmin(Volt15_4)
            Volt15_4.pop(indice+1)
            indice=np.argmax(Volt15_4)
            Volt15_4.pop(indice-1)
        if(len(Current15_4)>4):
            Current15_4.sort()
            indice=np.argmin(Current15_4)
            Current15_4.pop(indice+1)
            indice=np.argmax(Current15_4)
            Current15_4.pop(indice-1)
        if(len(ActivePower15_4)>4):
            ActivePower15_4.sort()
            indice=np.argmin(ActivePower15_4)
            ActivePower15_4.pop(indice+1)
            indice=np.argmax(ActivePower15_4)
            ActivePower15_4.pop(indice-1)
        if(len(ReactivePower15_4)>4):
            ReactivePower15_4.sort()
            indice=np.argmin(ReactivePower15_4)
            ReactivePower15_4.pop(indice+1)
            indice=np.argmax(ReactivePower15_4)
            ReactivePower15_4.pop(indice-1)
        if(len(AparentPower15_4)>4):
            AparentPower15_4.sort()
            indice=np.argmin(AparentPower15_4)
            AparentPower15_4.pop(indice+1)
            indice=np.argmax(AparentPower15_4)
            AparentPower15_4.pop(indice-1)
        if(len(FP15_Reactive_4)>4):
            FP15_Reactive_4.sort()
            indice=np.argmin(FP15_Reactive_4)
            FP15_Reactive_4.pop(indice+1)
            indice=np.argmax(FP15_Reactive_4)
            FP15_Reactive_4.pop(indice-1)
        if(len(FP15_Inductive_4)>4):
            FP15_Inductive_4.sort()
            indice=np.argmin(FP15_Inductive_4)
            FP15_Inductive_4.pop(indice+1)
            indice=np.argmax(FP15_Inductive_4)
            FP15_Inductive_4.pop(indice-1)
        if(len(FDVoltage15_4)>4):
            FDVoltage15_4.sort()
            indice=np.argmin(FDVoltage15_4)
            FDVoltage15_4.pop(indice+1)
            indice=np.argmax(FDVoltage15_4)
            FDVoltage15_4.pop(indice-1)
        if(len(FDCurrent15_4)>4):
            FDCurrent15_4.sort()
            indice=np.argmin(FDCurrent15_4)
            FDCurrent15_4.pop(indice+1)
            indice=np.argmax(FDCurrent15_4)
            FDCurrent15_4.pop(indice-1)
        if(len(DAT15Voltage_4)>4):
            DAT15Voltage_4.sort()
            indice=np.argmin(DAT15Voltage_4)
            DAT15Voltage_4.pop(indice+1)
            indice=np.argmax(DAT15Voltage_4)
            DAT15Voltage_4.pop(indice-1)
        if(len(DAT15Current_4)>4):
            DAT15Current_4.sort()
            indice=np.argmin(DAT15Current_4)
            DAT15Current_4.pop(indice+1)
            indice=np.argmax(DAT15Current_4)
            DAT15Current_4.pop(indice-1)

Access_5 = 0
MaxVoltage15_5=0.0
MeanVoltage15_5=0.0
MinVoltage15_5=0.0
MaxCurrent15_5=0.0
MeanCurrent15_5=0.0
MinCurrent15_5=0.0
MaxActivePower_5=0.0
MeanActivePower_5=0.0
MinActivePower_5=0.0
MaxReactivePower_5=0.0
MeanReactivePower_5=0.0
MinReactivePower_5=0.0
MaxAparentPower_5=0.0
MeanAparentPower_5=0.0
MinAparentPower_5=0.0
MaxFPInductive_5=-0.99
MeanFPInductive_5=-0.99
MinFPInductive_5=-0.99
MaxFPReactive_5=0.99
MeanFPReactive_5=0.99
MinFPReactive_5=0.99
MaxFD_5=0.0
MeanFD_5=0.0
MinFD_5=0.0
MaxDAT_5=0.0
MeanDAT_5=0.0
MinDAT_5=0.0
Volt15_5=[]
data15_5=[]
Current15_5=[]
ActivePower15_5=[]
ReactivePower15_5=[]
AparentPower15_5=[]
FP15_Reactive_5=[]
FP15_Inductive_5=[]
FDVoltage15_5=[]
FDCurrent15_5=[]
DAT15Voltage_5=[]
DAT15Current_5=[]
def Maximo15min_5(Vrms,Irms,ActivePower,ReactivePower,AparentPower,FP,FDVoltage,FDCurrent,DATVoltage,DATCurrent,OneHourEnergy,Energy,i,k,f):
    global data15_5
    global Volt15_5
    global data15_5
    global Current15_5
    global ActivePower15_5
    global ReactivePower15_5
    global AparentPower15_5
    global FP15_Reactive_5
    global FP15_Inductive_5
    global FDVoltage15_5
    global FDCurrent15_5
    global DAT15Voltage_5
    global DAT15Current_5
    global Access_5
    global OneHourEnergy_5
    global optionsave
    basea = datetime.datetime.now()
    if(basea.minute==0 or basea.minute==1 or basea.minute==2 or basea.minute==15 or basea.minute==16 or basea.minute==17 or basea.minute==30 or basea.minute==31 or basea.minute==32 or basea.minute==45 or basea.minute==46 or basea.minute==47):
               if(Access_5 == 0):
                    #graphVoltage(NoVoltageoffset2,ListaIrmsPeak2,samplings2,2)
                    Access_5 = 1
                    MaxVoltage15_5=max(Volt15_5)
                    MeanVoltage15_5=np.median(Volt15_5)
                    MinVoltage15_5=min(Volt15_5)
                    MaxCurrent15_5=max(Current15_5)
                    MeanCurrent15_5=np.median(Current15_5)
                    MinCurrent15_5=min(Current15_5)
                    MaxActivePower_5=max(ActivePower15_5)
                    MeanActivePower_5=np.median(ActivePower15_5)
                    MinActivePower_5=min(ActivePower15_5)
                    MaxReactivePower_5=max(ReactivePower15_5)
                    MeanReactivePower_5=np.median(ReactivePower15_5)
                    MinReactivePower_5=min(ReactivePower15_5)
                    MaxAparentPower_5=max(AparentPower15_5)
                    MeanAparentPower_5=np.median(AparentPower15_5)
                    MinAparentPower_5=min(AparentPower15_5)
                    if(len(FP15_Inductive_5)>0):
                           MaxFPInductive_5=max(FP15_Inductive_5)
                           MeanFPInductive_5=np.median(FP15_Inductive_5)
                           MinFPInductive_5=min(FP15_Inductive_5)
                    else:
                           MaxFPInductive_5=-0.99
                           MeanFPInductive_5=-0.99
                           MinFPInductive_5=-0.99
                    if(len(FP15_Reactive_5)>0):
                           MaxFPReactive_5=max(FP15_Reactive_5)
                           MeanFPReactive_5=np.median(FP15_Reactive_5)
                           MinFPReactive_5=min(FP15_Reactive_5)
                    else:
                           MaxFPReactive_5=0.99
                           MeanFPReactive_5=0.99
                           MinFPReactive_5=0.99
                    MaxFDVoltage_5=max(FDVoltage15_5)
                    MeanFDVoltage_5=np.median(FDVoltage15_5)
                    MinFDVoltage_5=min(FDVoltage15_5)
                    MaxFDCurrent_5=max(FDCurrent15_5)
                    MeanFDCurrent_5=np.median(FDCurrent15_5)
                    MinFDCurrent_5=min(FDCurrent15_5)
                    MaxDATVoltage_5=max(DAT15Voltage_5)
                    MeanDATVoltage_5=np.median(DAT15Voltage_5)
                    MinDATVoltage_5=min(DAT15Voltage_5)
                    MaxDATCurrent_5=max(DAT15Current_5)
                    MeanDATCurrent_5=np.median(DAT15Current_5)
                    MinDATCurrent_5=min(DAT15Current_5)
                    data15_5.insert(1,MaxVoltage15_5)
                    data15_5.insert(2,MeanVoltage15_5)
                    data15_5.insert(3,MinVoltage15_5)
                    data15_5.insert(4,MaxCurrent15_5)
                    data15_5.insert(5,MeanCurrent15_5)
                    data15_5.insert(6,MinCurrent15_5)
                    data15_5.insert(7,MaxActivePower_5)
                    data15_5.insert(8,MeanActivePower_5)
                    data15_5.insert(9,MinActivePower_5)
                    data15_5.insert(10,MaxReactivePower_5)
                    data15_5.insert(11,MeanReactivePower_5)
                    data15_5.insert(12,MinReactivePower_5)
                    data15_5.insert(13,MaxAparentPower_5)
                    data15_5.insert(14,MeanAparentPower_5)
                    data15_5.insert(15,MinAparentPower_5)
                    data15_5.insert(16,MaxFPInductive_5)
                    data15_5.insert(17,MeanFPInductive_5)
                    data15_5.insert(18,MinFPInductive_5)
                    data15_5.insert(19,MaxFPReactive_5)
                    data15_5.insert(20,MeanFPReactive_5)
                    data15_5.insert(21,MinFPReactive_5)
                    data15_5.insert(22,MaxFDVoltage_5)
                    data15_5.insert(23,MeanFDVoltage_5)
                    data15_5.insert(24,MinFDVoltage_5)
                    data15_5.insert(25,MaxFDCurrent_5)
                    data15_5.insert(26,MeanFDCurrent_5)
                    data15_5.insert(27,MinFDCurrent_5)
                    data15_5.insert(28,MaxDATVoltage_5)
                    data15_5.insert(29,MeanDATVoltage_5)
                    data15_5.insert(30,MinDATVoltage_5)
                    data15_5.insert(31,MaxDATCurrent_5)
                    data15_5.insert(32,MeanDATCurrent_5)
                    data15_5.insert(33,MinDATCurrent_5)
                    data15_5.insert(34,OneHourEnergy)
                    data15_5.insert(35,Energy)
                    data15_5.insert(0,datetime.datetime.now())
                    workbook=openpyxl.load_workbook(filename = dest_filename)
                    sheet6 = workbook[f"Mean-{k}-{f}"]
                    sheet6.append(list(data15_5))
                    print(f'Data 5: Guardando Promedios')
                    optionsave=1
                    SendDataToBroker(q=i,k=k,f=f,VoltajeMax=f'{MaxVoltage15_5}',VoltajePromedio=f'{MeanVoltage15_5}',VoltajeMin=f'{MinVoltage15_5}',MaxCorriente=f'{MaxCurrent15_5}',PromedioCorriente=f'{MeanCurrent15_5}',MinimoCorriente=f'{MinCurrent15_5}',PotenciaMax=f'{MaxAparentPower_5}',PromedioPotenciaAparente=f'{MeanAparentPower_5}',MinPotenciaAparente=f'{MinAparentPower_5}',OneHourEnergy=f'{OneHourEnergy}',Energia=f'{Energy}')
                    #print("Datos Insertados Correctamente!")
                    workbook.save(filename = dest_filename)
                    data15_5=[]
                    Volt15_5=[]
                    Current15_5=[]
                    ActivePower15_5=[]
                    ReactivePower15_5=[]
                    AparentPower15_5=[]
                    FP15_Reactive_5=[]
                    FP15_Inductive_5=[]
                    FDVoltage15_5=[]
                    FDCurrent15_5=[]
                    DAT15Voltage_5=[]
                    DAT15Current_5=[]
                    #OneHourEnergy_5=0
               elif(Access_5==1):
                    #print("paso elif 2")
                    Volt15_5.append(Vrms)
                    Current15_5.append(Irms)
                    ActivePower15_5.append(ActivePower)
                    ReactivePower15_5.append(ReactivePower)
                    AparentPower15_5.append(AparentPower)
                    if(FP>0.0):
                          FP15_Reactive_5.append(FP)
                    else: 
                          FP15_Inductive_5.append(FP)
                    FDVoltage15_5.append(FDVoltage)
                    FDCurrent15_5.append(FDCurrent)
                    DAT15Voltage_5.append(DATVoltage)
                    DAT15Current_5.append(DATCurrent)
              
    else:
        Volt15_5.append(Vrms)
        Current15_5.append(Irms)
        ActivePower15_5.append(ActivePower)
        ReactivePower15_5.append(ReactivePower)
        AparentPower15_5.append(AparentPower)
        if(FP>0.0):
              FP15_Reactive_5.append(FP)
        else: 
              FP15_Inductive_5.append(FP)
        FDVoltage15_5.append(FDVoltage)
        FDCurrent15_5.append(FDCurrent)
        DAT15Voltage_5.append(DATVoltage)
        DAT15Current_5.append(DATCurrent)
        Access_5 = 0
        
        if(len(Volt15_5)>4):
            Volt15_5.sort()
            indice=np.argmin(Volt15_5)
            Volt15_5.pop(indice+1)
            indice=np.argmax(Volt15_5)
            Volt15_5.pop(indice-1)
        if(len(Current15_5)>4):
            Current15_5.sort()
            indice=np.argmin(Current15_5)
            Current15_5.pop(indice+1)
            indice=np.argmax(Current15_5)
            Current15_5.pop(indice-1)
        if(len(ActivePower15_5)>4):
            ActivePower15_5.sort()
            indice=np.argmin(ActivePower15_5)
            ActivePower15_5.pop(indice+1)
            indice=np.argmax(ActivePower15_5)
            ActivePower15_5.pop(indice-1)
        if(len(ReactivePower15_5)>4):
            ReactivePower15_5.sort()
            indice=np.argmin(ReactivePower15_5)
            ReactivePower15_5.pop(indice+1)
            indice=np.argmax(ReactivePower15_5)
            ReactivePower15_5.pop(indice-1)
        if(len(AparentPower15_5)>4):
            AparentPower15_5.sort()
            indice=np.argmin(AparentPower15_5)
            AparentPower15_5.pop(indice+1)
            indice=np.argmax(AparentPower15_5)
            AparentPower15_5.pop(indice-1)
        if(len(FP15_Reactive_5)>4):
            FP15_Reactive_5.sort()
            indice=np.argmin(FP15_Reactive_5)
            FP15_Reactive_5.pop(indice+1)
            indice=np.argmax(FP15_Reactive_5)
            FP15_Reactive_5.pop(indice-1)
        if(len(FP15_Inductive_5)>4):
            FP15_Inductive_5.sort()
            indice=np.argmin(FP15_Inductive_5)
            FP15_Inductive_5.pop(indice+1)
            indice=np.argmax(FP15_Inductive_5)
            FP15_Inductive_5.pop(indice-1)
        if(len(FDVoltage15_5)>4):
            FDVoltage15_5.sort()
            indice=np.argmin(FDVoltage15_5)
            FDVoltage15_5.pop(indice+1)
            indice=np.argmax(FDVoltage15_5)
            FDVoltage15_5.pop(indice-1)
        if(len(FDCurrent15_5)>4):
            FDCurrent15_5.sort()
            indice=np.argmin(FDCurrent15_5)
            FDCurrent15_5.pop(indice+1)
            indice=np.argmax(FDCurrent15_5)
            FDCurrent15_5.pop(indice-1)
        if(len(DAT15Voltage_5)>4):
            DAT15Voltage_5.sort()
            indice=np.argmin(DAT15Voltage_5)
            DAT15Voltage_5.pop(indice+1)
            indice=np.argmax(DAT15Voltage_5)
            DAT15Voltage_5.pop(indice-1)
        if(len(DAT15Current_5)>4):
            DAT15Current_5.sort()
            indice=np.argmin(DAT15Current_5)
            DAT15Current_5.pop(indice+1)
            indice=np.argmax(DAT15Current_5)
            DAT15Current_5.pop(indice-1)
            
Access_7 = 0
MaxVoltage15_7=0.0
MeanVoltage15_7=0.0
MinVoltage15_7=0.0
MaxCurrent15_7=0.0
MeanCurrent15_7=0.0
MinCurrent15_7=0.0
MaxActivePower_7=0.0
MeanActivePower_7=0.0
MinActivePower_7=0.0
MaxReactivePower_7=0.0
MeanReactivePower_7=0.7
MinReactivePower_7=0.0
MaxAparentPower_7=0.0
MeanAparentPower_7=0.0
MinAparentPower_7=0.0
MaxFPInductive_7=-0.99
MeanFPInductive_7=-0.99
MinFPInductive_7=-0.99
MaxFPReactive_7=0.99
MeanFPReactive_7=0.99
MinFPReactive_7=0.99
MaxFD_7=0.0
MeanFD_7=0.0
MinFD_7=0.0
MaxDAT_7=0.0
MeanDAT_7=0.0
MinDAT_7=0.0
Volt15_7=[]
data15_7=[]
Current15_7=[]
ActivePower15_7=[]
ReactivePower15_7=[]
AparentPower15_7=[]
FP15_Reactive_7=[]
FP15_Inductive_7=[]
FDVoltage15_7=[]
FDCurrent15_7=[]
DAT15Voltage_7=[]
DAT15Current_7=[]
def Maximo15min_7(Vrms,Irms,ActivePower,ReactivePower,AparentPower,FP,FDVoltage,FDCurrent,DATVoltage,DATCurrent,OneHourEnergy,Energy,i,k,f):
    global data15_7
    global Volt15_7
    global data15_7
    global Current15_7
    global ActivePower15_7
    global ReactivePower15_7
    global AparentPower15_7
    global FP15_Reactive_7
    global FP15_Inductive_7
    global FDVoltage15_7
    global FDCurrent15_7
    global DAT15Voltage_7
    global DAT15Current_7
    global Access_7
    global OneHourEnergy_7
    global optionsave
    basea = datetime.datetime.now()
    if(basea.minute==0 or basea.minute==1 or basea.minute==2 or basea.minute==15 or basea.minute==16 or basea.minute==17 or basea.minute==30 or basea.minute==31 or basea.minute==32 or basea.minute==45 or basea.minute==46 or basea.minute==47):
               if(Access_7 == 0):
                    #graphVoltage(NoVoltageoffset2,ListaIrmsPeak2,samplings2,2)
                    Access_7 = 1
                    MaxVoltage15_7=max(Volt15_7)
                    MeanVoltage15_7=np.median(Volt15_7)
                    MinVoltage15_7=min(Volt15_7)
                    MaxCurrent15_7=max(Current15_7)
                    MeanCurrent15_7=np.median(Current15_7)
                    MinCurrent15_7=min(Current15_7)
                    MaxActivePower_7=max(ActivePower15_7)
                    MeanActivePower_7=np.median(ActivePower15_7)
                    MinActivePower_7=min(ActivePower15_7)
                    MaxReactivePower_7=max(ReactivePower15_7)
                    MeanReactivePower_7=np.median(ReactivePower15_7)
                    MinReactivePower_7=min(ReactivePower15_7)
                    MaxAparentPower_7=max(AparentPower15_7)
                    MeanAparentPower_7=np.median(AparentPower15_7)
                    MinAparentPower_7=min(AparentPower15_7)
                    if(len(FP15_Inductive_7)>0):
                           MaxFPInductive_7=max(FP15_Inductive_7)
                           MeanFPInductive_7=np.median(FP15_Inductive_7)
                           MinFPInductive_7=min(FP15_Inductive_7)
                    else:
                           MaxFPInductive_7=-0.99
                           MeanFPInductive_7=-0.99
                           MinFPInductive_7=-0.99
                    if(len(FP15_Reactive_7)>0):
                           MaxFPReactive_7=max(FP15_Reactive_7)
                           MeanFPReactive_7=np.median(FP15_Reactive_7)
                           MinFPReactive_7=min(FP15_Reactive_7)
                    else:
                           MaxFPReactive_7=0.99
                           MeanFPReactive_7=0.99
                           MinFPReactive_7=0.99
                    MaxFDVoltage_7=max(FDVoltage15_7)
                    MeanFDVoltage_7=np.median(FDVoltage15_7)
                    MinFDVoltage_7=min(FDVoltage15_7)
                    MaxFDCurrent_7=max(FDCurrent15_7)
                    MeanFDCurrent_7=np.median(FDCurrent15_7)
                    MinFDCurrent_7=min(FDCurrent15_7)
                    MaxDATVoltage_7=max(DAT15Voltage_7)
                    MeanDATVoltage_7=np.median(DAT15Voltage_7)
                    MinDATVoltage_7=min(DAT15Voltage_7)
                    MaxDATCurrent_7=max(DAT15Current_7)
                    MeanDATCurrent_7=np.median(DAT15Current_7)
                    MinDATCurrent_7=min(DAT15Current_7)
                    data15_7.insert(1,MaxVoltage15_7)
                    data15_7.insert(2,MeanVoltage15_7)
                    data15_7.insert(3,MinVoltage15_7)
                    data15_7.insert(4,MaxCurrent15_7)
                    data15_7.insert(5,MeanCurrent15_7)
                    data15_7.insert(6,MinCurrent15_7)
                    data15_7.insert(7,MaxActivePower_7)
                    data15_7.insert(8,MeanActivePower_7)
                    data15_7.insert(9,MinActivePower_7)
                    data15_7.insert(10,MaxReactivePower_7)
                    data15_7.insert(11,MeanReactivePower_7)
                    data15_7.insert(12,MinReactivePower_7)
                    data15_7.insert(13,MaxAparentPower_7)
                    data15_7.insert(14,MeanAparentPower_7)
                    data15_7.insert(15,MinAparentPower_7)
                    data15_7.insert(16,MaxFPInductive_7)
                    data15_7.insert(17,MeanFPInductive_7)
                    data15_7.insert(18,MinFPInductive_7)
                    data15_7.insert(19,MaxFPReactive_7)
                    data15_7.insert(20,MeanFPReactive_7)
                    data15_7.insert(21,MinFPReactive_7)
                    data15_7.insert(22,MaxFDVoltage_7)
                    data15_7.insert(23,MeanFDVoltage_7)
                    data15_7.insert(24,MinFDVoltage_7)
                    data15_7.insert(25,MaxFDCurrent_7)
                    data15_7.insert(26,MeanFDCurrent_7)
                    data15_7.insert(27,MinFDCurrent_7)
                    data15_7.insert(28,MaxDATVoltage_7)
                    data15_7.insert(29,MeanDATVoltage_7)
                    data15_7.insert(30,MinDATVoltage_7)
                    data15_7.insert(31,MaxDATCurrent_7)
                    data15_7.insert(32,MeanDATCurrent_7)
                    data15_7.insert(33,MinDATCurrent_7)
                    data15_7.insert(34,MinDATCurrent_7)
                    data15_7.insert(35,Energy)
                    data15_7.insert(0,datetime.datetime.now())
                    workbook=openpyxl.load_workbook(filename = dest_filename)
                    sheet8 = workbook[f"Mean-{k}-{f}"]
                    sheet8.append(list(data15_7))
                    print(f'Data 7: Guardando Promedios')
                    optionsave=1
                    SendDataToBroker(q=i,k=k,f=f,VoltajeMax=f'{MaxVoltage15_7}',VoltajePromedio=f'{MeanVoltage15_7}',VoltajeMin=f'{MinVoltage15_7}',MaxCorriente=f'{MaxCurrent15_7}',PromedioCorriente=f'{MeanCurrent15_7}',MinimoCorriente=f'{MinCurrent15_7}',PotenciaMax=f'{MaxAparentPower_7}',PromedioPotenciaAparente=f'{MeanAparentPower_7}',MinPotenciaAparente=f'{MinAparentPower_7}',OneHourEnergy=f'{OneHourEnergy}',Energia=f'{Energy}')
                    #print("Datos Insertados Correctamente!")
                    workbook.save(filename = dest_filename)
                    data15_7=[]
                    Volt15_7=[]
                    Current15_7=[]
                    ActivePower15_7=[]
                    ReactivePower15_7=[]
                    AparentPower15_7=[]
                    FP15_Reactive_7=[]
                    FP15_Inductive_7=[]
                    FDVoltage15_7=[]
                    FDCurrent15_7=[]
                    DAT15Voltage_7=[]
                    DAT15Current_7=[]
                    #OneHourEnergy_7=0
               elif(Access_7==1):
                    #print("paso elif 2")
                    Volt15_7.append(Vrms)
                    Current15_7.append(Irms)
                    ActivePower15_7.append(ActivePower)
                    ReactivePower15_7.append(ReactivePower)
                    AparentPower15_7.append(AparentPower)
                    if(FP>0.0):
                          FP15_Reactive_7.append(FP)
                    else: 
                          FP15_Inductive_7.append(FP)
                    FDVoltage15_7.append(FDVoltage)
                    FDCurrent15_7.append(FDCurrent)
                    DAT15Voltage_7.append(DATVoltage)
                    DAT15Current_7.append(DATCurrent)          
    else:
        Volt15_7.append(Vrms)
        Current15_7.append(Irms)
        ActivePower15_7.append(ActivePower)
        ReactivePower15_7.append(ReactivePower)
        AparentPower15_7.append(AparentPower)
        if(FP>0.0):
              FP15_Reactive_7.append(FP)
        else: 
              FP15_Inductive_7.append(FP)
        FDVoltage15_7.append(FDVoltage)
        FDCurrent15_7.append(FDCurrent)
        DAT15Voltage_7.append(DATVoltage)
        DAT15Current_7.append(DATCurrent)
        Access_7 = 0
        
        if(len(Volt15_7)>4):
            Volt15_7.sort()
            indice=np.argmin(Volt15_7)
            Volt15_7.pop(indice+1)
            indice=np.argmax(Volt15_7)          
            Volt15_7.pop(indice-1)
        if(len(Current15_7)>4):
            Current15_7.sort()
            indice=np.argmin(Current15_7)
            Current15_7.pop(indice+1)
            indice=np.argmax(Current15_7)
            Current15_7.pop(indice-1)
        if(len(ActivePower15_7)>4):
            ActivePower15_7.sort()
            indice=np.argmin(ActivePower15_7)
            ActivePower15_7.pop(indice+1)
            indice=np.argmax(ActivePower15_7)
            ActivePower15_7.pop(indice-1)
        if(len(ReactivePower15_7)>4):
            ReactivePower15_7.sort()
            indice=np.argmin(ReactivePower15_7)
            ReactivePower15_7.pop(indice+1)
            indice=np.argmax(ReactivePower15_7)
            ReactivePower15_7.pop(indice-1)
        if(len(AparentPower15_7)>4):
            AparentPower15_7.sort()
            indice=np.argmin(AparentPower15_7)
            AparentPower15_7.pop(indice+1)
            indice=np.argmax(AparentPower15_7)
            AparentPower15_7.pop(indice-1)
        if(len(FP15_Reactive_7)>4):
            FP15_Reactive_7.sort()
            indice=np.argmin(FP15_Reactive_7)
            FP15_Reactive_7.pop(indice+1)
            indice=np.argmax(FP15_Reactive_7)
            FP15_Reactive_7.pop(indice-1)
        if(len(FP15_Inductive_7)>4):
            FP15_Inductive_7.sort()
            indice=np.argmin(FP15_Inductive_7)
            FP15_Inductive_7.pop(indice+1)
            indice=np.argmax(FP15_Inductive_7)
            FP15_Inductive_7.pop(indice-1)
        if(len(FDVoltage15_7)>4):
            FDVoltage15_7.sort()
            indice=np.argmin(FDVoltage15_7)
            FDVoltage15_7.pop(indice+1)
            indice=np.argmax(FDVoltage15_7)
            FDVoltage15_7.pop(indice-1)
        if(len(FDCurrent15_7)>4):
            FDCurrent15_7.sort()
            indice=np.argmin(FDCurrent15_7)
            FDCurrent15_7.pop(indice+1)
            indice=np.argmax(FDCurrent15_7)
            FDCurrent15_7.pop(indice-1)
        if(len(DAT15Voltage_7)>4):
            DAT15Voltage_7.sort()
            indice=np.argmin(DAT15Voltage_7)
            DAT15Voltage_7.pop(indice+1)
            indice=np.argmax(DAT15Voltage_7)
            DAT15Voltage_7.pop(indice-1)
        if(len(DAT15Current_7)>4):
            DAT15Current_7.sort()
            indice=np.argmin(DAT15Current_7)
            DAT15Current_7.pop(indice+1)
            indice=np.argmax(DAT15Current_7)
            DAT15Current_7.pop(indice-1)

Access_8 = 0
MaxVoltage15_8=0.0
MeanVoltage15_8=0.0
MinVoltage15_8=0.0
MaxCurrent15_8=0.0
MeanCurrent15_8=0.0
MinCurrent15_8=0.0
MaxActivePower_8=0.0
MeanActivePower_8=0.0
MinActivePower_8=0.0
MaxReactivePower_8=0.0
MeanReactivePower_8=0.0
MinReactivePower_8=0.0
MaxAparentPower_8=0.0
MeanAparentPower_8=0.0
MinAparentPower_8=0.0
MaxFPInductive_8=-0.99
MeanFPInductive_8=-0.99
MinFPInductive_8=-0.99
MaxFPReactive_8=0.99
MeanFPReactive_8=0.99
MinFPReactive_8=0.99
MaxFD_8=0.0
MeanFD_8=0.0
MinFD_8=0.0
MaxDAT_8=0.0
MeanDAT_8=0.0
MinDAT_8=0.0
Volt15_8=[]
data15_8=[]
Current15_8=[]
ActivePower15_8=[]
ReactivePower15_8=[]
AparentPower15_8=[]
FP15_Reactive_8=[]
FP15_Inductive_8=[]
FDVoltage15_8=[]
FDCurrent15_8=[]
DAT15Voltage_8=[]
DAT15Current_8=[]
def Maximo15min_8(Vrms,Irms,ActivePower,ReactivePower,AparentPower,FP,FDVoltage,FDCurrent,DATVoltage,DATCurrent,OneHourEnergy,Energy,i,k,f):
    global data15_8
    global Volt15_8
    global data15_8
    global Current15_8
    global ActivePower15_8
    global ReactivePower15_8
    global AparentPower15_8
    global FP15_Reactive_8
    global FP15_Inductive_8
    global FDVoltage15_8
    global FDCurrent15_8
    global DAT15Voltage_8
    global DAT15Current_8
    global Access_8
    global OneHourEnergy_8
    global optionsave
    basea = datetime.datetime.now()
    if(basea.minute==0 or basea.minute==1 or basea.minute==2 or basea.minute==15 or basea.minute==16 or basea.minute==17 or basea.minute==30 or basea.minute==31 or basea.minute==32 or basea.minute==45 or basea.minute==46 or basea.minute==47):
               if(Access_8 == 0):
                    #graphVoltage(NoVoltageoffset2,ListaIrmsPeak2,samplings2,2)
                    Access_8 = 1
                    MaxVoltage15_8=max(Volt15_8)
                    MeanVoltage15_8=np.median(Volt15_8)
                    MinVoltage15_8=min(Volt15_8)
                    MaxCurrent15_8=max(Current15_8)
                    MeanCurrent15_8=np.median(Current15_8)
                    MinCurrent15_8=min(Current15_8)
                    MaxActivePower_8=max(ActivePower15_8)
                    MeanActivePower_8=np.median(ActivePower15_8)
                    MinActivePower_8=min(ActivePower15_8)
                    MaxReactivePower_8=max(ReactivePower15_8)
                    MeanReactivePower_8=np.median(ReactivePower15_8)
                    MinReactivePower_8=min(ReactivePower15_8)
                    MaxAparentPower_8=max(AparentPower15_8)
                    MeanAparentPower_8=np.median(AparentPower15_8)
                    MinAparentPower_8=min(AparentPower15_8)
                    if(len(FP15_Inductive_8)>0):
                           MaxFPInductive_8=max(FP15_Inductive_8)
                           MeanFPInductive_8=np.median(FP15_Inductive_8)
                           MinFPInductive_8=min(FP15_Inductive_8)
                    else:
                           MaxFPInductive_8=-0.59
                           MeanFPInductive_8=-0.99
                           MinFPInductive_8=-0.99
                    if(len(FP15_Reactive_8)>0):
                           MaxFPReactive_8=max(FP15_Reactive_8)
                           MeanFPReactive_8=np.median(FP15_Reactive_8)
                           MinFPReactive_8=min(FP15_Reactive_8)
                    else:
                           MaxFPReactive_8=0.99
                           MeanFPReactive_8=0.99
                           MinFPReactive_8=0.99
                    MaxFDVoltage_8=max(FDVoltage15_8)
                    MeanFDVoltage_8=np.median(FDVoltage15_8)
                    MinFDVoltage_8=min(FDVoltage15_8)
                    MaxFDCurrent_8=max(FDCurrent15_8)
                    MeanFDCurrent_8=np.median(FDCurrent15_8)
                    MinFDCurrent_8=min(FDCurrent15_8)
                    MaxDATVoltage_8=max(DAT15Voltage_8)
                    MeanDATVoltage_8=np.median(DAT15Voltage_8)
                    MinDATVoltage_8=min(DAT15Voltage_8)
                    MaxDATCurrent_8=max(DAT15Current_8)
                    MeanDATCurrent_8=np.median(DAT15Current_8)
                    MinDATCurrent_8=min(DAT15Current_8)
                    data15_8.insert(1,MaxVoltage15_8)
                    data15_8.insert(2,MeanVoltage15_8)
                    data15_8.insert(3,MinVoltage15_8)
                    data15_8.insert(4,MaxCurrent15_8)
                    data15_8.insert(5,MeanCurrent15_8)
                    data15_8.insert(6,MinCurrent15_8)
                    data15_8.insert(7,MaxActivePower_8)
                    data15_8.insert(8,MeanActivePower_8)
                    data15_8.insert(9,MinActivePower_8)
                    data15_8.insert(10,MaxReactivePower_8)
                    data15_8.insert(11,MeanReactivePower_8)
                    data15_8.insert(12,MinReactivePower_8)
                    data15_8.insert(13,MaxAparentPower_8)
                    data15_8.insert(14,MeanAparentPower_8)
                    data15_8.insert(15,MinAparentPower_8)
                    data15_8.insert(16,MaxFPInductive_8)
                    data15_8.insert(17,MeanFPInductive_8)
                    data15_8.insert(18,MinFPInductive_8)
                    data15_8.insert(19,MaxFPReactive_8)
                    data15_8.insert(20,MeanFPReactive_8)
                    data15_5.insert(21,MinFPReactive_8)
                    data15_8.insert(22,MaxFDVoltage_8)
                    data15_8.insert(23,MeanFDVoltage_8)
                    data15_8.insert(24,MinFDVoltage_8)
                    data15_8.insert(25,MaxFDCurrent_8)
                    data15_8.insert(26,MeanFDCurrent_8)
                    data15_8.insert(27,MinFDCurrent_8)
                    data15_8.insert(28,MaxDATVoltage_8)
                    data15_8.insert(29,MeanDATVoltage_8)
                    data15_8.insert(30,MinDATVoltage_8)
                    data15_8.insert(31,MaxDATCurrent_8)
                    data15_8.insert(32,MeanDATCurrent_8)
                    data15_8.insert(33,MinDATCurrent_8)
                    data15_8.insert(34,OneHourEnergy)
                    data15_8.insert(35,Energy)
                    data15_8.insert(0,datetime.datetime.now())
                    workbook=openpyxl.load_workbook(filename = dest_filename)
                    sheet9 = workbook[f"Mean-{k}-{f}"]
                    sheet9.append(list(data15_8))
                    print(f'Data 8: Guardando Promedios')
                    optionsave=1
                    SendDataToBroker(q=i,k=k,f=f,VoltajeMax=f'{MaxVoltage15_8}',VoltajePromedio=f'{MeanVoltage15_8}',VoltajeMin=f'{MinVoltage15_8}',MaxCorriente=f'{MaxCurrent15_8}',PromedioCorriente=f'{MeanCurrent15_8}',MinimoCorriente=f'{MinCurrent15_8}',PotenciaMax=f'{MaxAparentPower_8}',PromedioPotenciaAparente=f'{MeanAparentPower_8}',MinPotenciaAparente=f'{MinAparentPower_8}',OneHourEnergy=f'{OneHourEnergy}',Energia=f'{Energy}')
                    #print("Datos Insertados Correctamente!")
                    workbook.save(filename = dest_filename)
                    data15_8=[]
                    Volt15_8=[]
                    Current15_8=[]
                    ActivePower15_8=[]
                    ReactivePower15_8=[]
                    AparentPower15_8=[]
                    FP15_Reactive_8=[]
                    FP15_Inductive_8=[]
                    FDVoltage15_8=[]
                    FDCurrent15_8=[]
                    DAT15Voltage_8=[]
                    DAT15Current_8=[]
                    #OneHourEnergy_8=0
               elif(Access_8==1):
                    #print("paso elif 2")
                    Volt15_8.append(Vrms)
                    Current15_8.append(Irms)
                    ActivePower15_8.append(ActivePower)
                    ReactivePower15_8.append(ReactivePower)
                    AparentPower15_8.append(AparentPower)
                    if(FP>0.0):
                          FP15_Reactive_8.append(FP)
                    else: 
                          FP15_Inductive_8.append(FP)
                    FDVoltage15_8.append(FDVoltage)
                    FDCurrent15_8.append(FDCurrent)
                    DAT15Voltage_8.append(DATVoltage)
                    DAT15Current_8.append(DATCurrent)            
    else:
        Volt15_8.append(Vrms)
        Current15_8.append(Irms)
        ActivePower15_8.append(ActivePower)
        ReactivePower15_8.append(ReactivePower)
        AparentPower15_8.append(AparentPower)
        if(FP>0.0):
              FP15_Reactive_8.append(FP)
        else: 
              FP15_Inductive_8.append(FP)
        FDVoltage15_8.append(FDVoltage)
        FDCurrent15_8.append(FDCurrent)
        DAT15Voltage_8.append(DATVoltage)
        DAT15Current_8.append(DATCurrent)
        Access_8 = 0       
        if(len(Volt15_8)>4):
            Volt15_8.sort()
            indice=np.argmin(Volt15_8)
            Volt15_8.pop(indice+1)
            indice=np.argmax(Volt15_8)
            Volt15_8.pop(indice-1)
        if(len(Current15_8)>4):
            Current15_8.sort()
            indice=np.argmin(Current15_8)
            Current15_8.pop(indice+1)
            indice=np.argmax(Current15_8)
            Current15_8.pop(indice-1)
        if(len(ActivePower15_8)>4):
            ActivePower15_8.sort()
            indice=np.argmin(ActivePower15_8)
            ActivePower15_8.pop(indice+1)
            indice=np.argmax(ActivePower15_8)
            ActivePower15_8.pop(indice-1)
        if(len(ReactivePower15_8)>4):
            ReactivePower15_8.sort()
            indice=np.argmin(ReactivePower15_8)
            ReactivePower15_8.pop(indice+1)
            indice=np.argmax(ReactivePower15_8)
            ReactivePower15_8.pop(indice-1)
        if(len(AparentPower15_8)>4):
            AparentPower15_8.sort()
            indice=np.argmin(AparentPower15_8)
            AparentPower15_8.pop(indice+1)
            indice=np.argmax(AparentPower15_8)
            AparentPower15_8.pop(indice-1)
        if(len(FP15_Reactive_8)>4):
            FP15_Reactive_8.sort()
            indice=np.argmin(FP15_Reactive_8)
            FP15_Reactive_8.pop(indice+1)
            indice=np.argmax(FP15_Reactive_8)
            FP15_Reactive_8.pop(indice-1)
        if(len(FP15_Inductive_8)>4):
            FP15_Inductive_8.sort()
            indice=np.argmin(FP15_Inductive_8)
            FP15_Inductive_8.pop(indice+1)
            indice=np.argmax(FP15_Inductive_8)
            FP15_Inductive_8.pop(indice-1)
        if(len(FDVoltage15_8)>4):
            FDVoltage15_8.sort()
            indice=np.argmin(FDVoltage15_8)
            FDVoltage15_8.pop(indice+1)
            indice=np.argmax(FDVoltage15_8)
            FDVoltage15_8.pop(indice-1)
        if(len(FDCurrent15_8)>4):
            FDCurrent15_8.sort()
            indice=np.argmin(FDCurrent15_8)
            FDCurrent15_8.pop(indice+1)
            indice=np.argmax(FDCurrent15_8)
            FDCurrent15_8.pop(indice-1)
        if(len(DAT15Voltage_8)>4):
            DAT15Voltage_8.sort()
            indice=np.argmin(DAT15Voltage_8)
            DAT15Voltage_8.pop(indice+1)
            indice=np.argmax(DAT15Voltage_8)
            DAT15Voltage_8.pop(indice-1)
        if(len(DAT15Current_8)>4):
            DAT15Current_8.sort()
            indice=np.argmin(DAT15Current_8)
            DAT15Current_8.pop(indice+1)
            indice=np.argmax(DAT15Current_8)
            DAT15Current_8.pop(indice-1)

Access_6 = 0
MaxVoltage15_6=0.0
MeanVoltage15_6=0.0
MinVoltage15_6=0.0
MaxCurrent15_6=0.0
MeanCurrent15_6=0.0
MinCurrent15_6=0.0
MaxActivePower_6=0.0
MeanActivePower_6=0.0
MinActivePower_6=0.0
MaxReactivePower_6=0.0
MeanReactivePower_6=0.0
MinReactivePower_6=0.0
MaxAparentPower_6=0.0
MeanAparentPower_6=0.0
MinAparentPower_6=0.0
MaxFPInductive_6=-0.99
MeanFPInductive_6=-0.99
MinFPInductive_6=-0.99
MaxFPReactive_6=0.99
MeanFPReactive_6=0.99
MinFPReactive_6=0.99
MaxFD_6=0.0
MeanFD_6=0.0
MinFD_6=0.0
MaxDAT_6=0.0
MeanDAT_6=0.0
MinDAT_6=0.0
Volt15_6=[]
data15_6=[]
Current15_6=[]
ActivePower15_6=[]
ReactivePower15_6=[]
AparentPower15_6=[]
FP15_Reactive_6=[]
FP15_Inductive_6=[]
FDVoltage15_6=[]
FDCurrent15_6=[]
DAT15Voltage_6=[]
DAT15Current_6=[]
def Maximo15min_6(Vrms,Irms,ActivePower,ReactivePower,AparentPower,FP,FDVoltage,FDCurrent,DATVoltage,DATCurrent,OneHourEnergy,Energy,i,k,f):
    global data15_6
    global Volt15_6
    global data15_6
    global Current15_6
    global ActivePower15_6
    global ReactivePower15_6
    global AparentPower15_6
    global FP15_Reactive_6
    global FP15_Inductive_6
    global FDVoltage15_6
    global FDCurrent15_6
    global DAT15Voltage_6
    global DAT15Current_6
    global Access_6
    global OneHourEnergy_6
    global optionsave
    basea = datetime.datetime.now()
    if(basea.minute==0 or basea.minute==1 or basea.minute==2 or basea.minute==15 or basea.minute==16 or basea.minute==17 or basea.minute==30 or basea.minute==31 or basea.minute==32 or basea.minute==45 or basea.minute==46 or basea.minute==47):
               if(Access_6 == 0):
                    #graphVoltage(NoVoltageoffset2,ListaIrmsPeak2,samplings2,2)
                    Access_6 = 1
                    MaxVoltage15_6=max(Volt15_6)
                    MeanVoltage15_6=np.median(Volt15_6)
                    MinVoltage15_6=min(Volt15_6)
                    MaxCurrent15_6=max(Current15_6)
                    MeanCurrent15_6=np.median(Current15_6)
                    MinCurrent15_6=min(Current15_6)
                    MaxActivePower_6=max(ActivePower15_6)
                    MeanActivePower_6=np.median(ActivePower15_6)
                    MinActivePower_6=min(ActivePower15_6)
                    MaxReactivePower_6=max(ReactivePower15_6)
                    MeanReactivePower_6=np.median(ReactivePower15_6)
                    MinReactivePower_6=min(ReactivePower15_6)
                    MaxAparentPower_6=max(AparentPower15_6)
                    MeanAparentPower_6=np.median(AparentPower15_6)
                    MinAparentPower_6=min(AparentPower15_6)
                    if(len(FP15_Inductive_6)>0):
                           MaxFPInductive_6=max(FP15_Inductive_6)
                           MeanFPInductive_6=np.median(FP15_Inductive_6)
                           MinFPInductive_6=min(FP15_Inductive_6)
                    else:
                           MaxFPInductive_6=-0.59
                           MeanFPInductive_6=-0.99
                           MinFPInductive_6=-0.99
                    if(len(FP15_Reactive_6)>0):
                           MaxFPReactive_6=max(FP15_Reactive_6)
                           MeanFPReactive_6=np.median(FP15_Reactive_6)
                           MinFPReactive_6=min(FP15_Reactive_6)
                    else:
                           MaxFPReactive_6=0.99
                           MeanFPReactive_6=0.99
                           MinFPReactive_6=0.99
                    MaxFDVoltage_6=max(FDVoltage15_6)
                    MeanFDVoltage_6=np.median(FDVoltage15_6)
                    MinFDVoltage_6=min(FDVoltage15_6)
                    MaxFDCurrent_6=max(FDCurrent15_6)
                    MeanFDCurrent_6=np.median(FDCurrent15_6)
                    MinFDCurrent_6=min(FDCurrent15_6)
                    MaxDATVoltage_6=max(DAT15Voltage_6)
                    MeanDATVoltage_6=np.median(DAT15Voltage_6)
                    MinDATVoltage_6=min(DAT15Voltage_6)
                    MaxDATCurrent_6=max(DAT15Current_6)
                    MeanDATCurrent_6=np.median(DAT15Current_6)
                    MinDATCurrent_6=min(DAT15Current_6)
                    data15_6.insert(1,MaxVoltage15_6)
                    data15_6.insert(2,MeanVoltage15_6)
                    data15_6.insert(3,MinVoltage15_6)
                    data15_6.insert(4,MaxCurrent15_6)
                    data15_6.insert(5,MeanCurrent15_6)
                    data15_6.insert(6,MinCurrent15_6)
                    data15_6.insert(7,MaxActivePower_6)
                    data15_6.insert(8,MeanActivePower_6)
                    data15_6.insert(9,MinActivePower_6)
                    data15_6.insert(10,MaxReactivePower_6)
                    data15_6.insert(11,MeanReactivePower_6)
                    data15_6.insert(12,MinReactivePower_6)
                    data15_6.insert(13,MaxAparentPower_6)
                    data15_6.insert(14,MeanAparentPower_6)
                    data15_6.insert(15,MinAparentPower_6)
                    data15_6.insert(16,MaxFPInductive_6)
                    data15_6.insert(17,MeanFPInductive_6)
                    data15_6.insert(18,MinFPInductive_6)
                    data15_6.insert(19,MaxFPReactive_6)
                    data15_6.insert(20,MeanFPReactive_6)
                    data15_6.insert(21,MinFPReactive_6)
                    data15_6.insert(22,MaxFDVoltage_6)
                    data15_6.insert(23,MeanFDVoltage_6)
                    data15_6.insert(24,MinFDVoltage_6)
                    data15_6.insert(25,MaxFDCurrent_6)
                    data15_6.insert(26,MeanFDCurrent_6)
                    data15_6.insert(27,MinFDCurrent_6)
                    data15_6.insert(28,MaxDATVoltage_6)
                    data15_6.insert(29,MeanDATVoltage_6)
                    data15_6.insert(30,MinDATVoltage_6)
                    data15_6.insert(31,MaxDATCurrent_6)
                    data15_6.insert(32,MeanDATCurrent_6)
                    data15_6.insert(33,MinDATCurrent_6)
                    data15_6.insert(34,OneHourEnergy_6)
                    data15_6.insert(35,Energy)
                    data15_6.insert(0,datetime.datetime.now())
                    workbook=openpyxl.load_workbook(filename = dest_filename)
                    sheet7 = workbook[f"Mean-{k}-{f}"]
                    sheet7.append(list(data15_6))
                    print(f'Data 6: Guardando Promedios')
                    optionsave=1
                    SendDataToBroker(q=i,k=k,f=f,VoltajeMax=f'{MaxVoltage15_6}',VoltajePromedio=f'{MeanVoltage15_6}',VoltajeMin=f'{MinVoltage15_6}',MaxCorriente=f'{MaxCurrent15_6}',PromedioCorriente=f'{MeanCurrent15_6}',MinimoCorriente=f'{MinCurrent15_6}',PotenciaMax=f'{MaxAparentPower_6}',PromedioPotenciaAparente=f'{MeanAparentPower_6}',MinPotenciaAparente=f'{MinAparentPower_6}',OneHourEnergy=f'{OneHourEnergy}',Energia=f'{Energy}')
                    #print("Datos Insertados Correctamente!")
                    workbook.save(filename = dest_filename)
                    data15_6=[]
                    Volt15_6=[]
                    Current15_6=[]
                    ActivePower15_6=[]
                    ReactivePower15_6=[]
                    AparentPower15_6=[]
                    FP15_Reactive_6=[]
                    FP15_Inductive_6=[]
                    FDVoltage15_6=[]
                    FDCurrent15_6=[]
                    DAT15Voltage_6=[]
                    DAT15Current_6=[]
                    #OneHourEnergy_6=0
               elif(Access_6==1):
                    #print("paso elif 2")
                    Volt15_6.append(Vrms)
                    Current15_6.append(Irms)
                    ActivePower15_6.append(ActivePower)
                    ReactivePower15_6.append(ReactivePower)
                    AparentPower15_6.append(AparentPower)
                    if(FP>0.0):
                          FP15_Reactive_6.append(FP)
                    else: 
                          FP15_Inductive_6.append(FP)
                    FDVoltage15_6.append(FDVoltage)
                    FDCurrent15_6.append(FDCurrent)
                    DAT15Voltage_6.append(DATVoltage)
                    DAT15Current_6.append(DATCurrent)
              
    else:
        Volt15_6.append(Vrms)
        Current15_6.append(Irms)
        ActivePower15_6.append(ActivePower)
        ReactivePower15_6.append(ReactivePower)
        AparentPower15_6.append(AparentPower)
        if(FP>0.0):
              FP15_Reactive_6.append(FP)
        else: 
              FP15_Inductive_6.append(FP)
        FDVoltage15_6.append(FDVoltage)
        FDCurrent15_6.append(FDCurrent)
        DAT15Voltage_6.append(DATVoltage)
        DAT15Current_6.append(DATCurrent)
        Access_6 = 0
        
        if(len(Volt15_6)>4):
            Volt15_6.sort()
            indice=np.argmin(Volt15_6)
            Volt15_6.pop(indice+1)
            indice=np.argmax(Volt15_6)
            Volt15_6.pop(indice-1)
        if(len(Current15_6)>4):  
            Current15_6.sort()     
            indice=np.argmin(Current15_6)
            Current15_6.pop(indice+1)
            indice=np.argmax(Current15_6)
            Current15_6.pop(indice-1)
        if(len(ActivePower15_6)>4):
            ActivePower15_6.sort()
            indice=np.argmin(ActivePower15_6)
            ActivePower15_6.pop(indice+1)
            indice=np.argmax(ActivePower15_6)
            ActivePower15_6.pop(indice-1)
        if(len(ReactivePower15_6)>4):
            ReactivePower15_6.sort()
            indice=np.argmin(ReactivePower15_6)
            ReactivePower15_6.pop(indice+1)
            indice=np.argmax(ReactivePower15_6)
            ReactivePower15_6.pop(indice-1)
        if(len(AparentPower15_6)>4):
            AparentPower15_6.sort()
            indice=np.argmin(AparentPower15_6)
            AparentPower15_6.pop(indice+1)
            indice=np.argmax(AparentPower15_6)
            AparentPower15_6.pop(indice-1)
        if(len(FP15_Reactive_6)>4):
            FP15_Reactive_6.sort()
            indice=np.argmin(FP15_Reactive_6)
            FP15_Reactive_6.pop(indice+1)
            indice=np.argmax(FP15_Reactive_6)
            FP15_Reactive_6.pop(indice-1)
        if(len(FP15_Inductive_6)>4):
            FP15_Inductive_6.sort()
            indice=np.argmin(FP15_Inductive_6)
            FP15_Inductive_6.pop(indice+1)
            indice=np.argmax(FP15_Inductive_6)
            FP15_Inductive_6.pop(indice-1)
        if(len(FDVoltage15_6)>4):
            FDVoltage15_6.sort()
            indice=np.argmin(FDVoltage15_6)
            FDVoltage15_6.pop(indice+1)
            indice=np.argmax(FDVoltage15_6)
            FDVoltage15_6.pop(indice-1)
        if(len(FDCurrent15_6)>4):
            FDCurrent15_6.sort()
            indice=np.argmin(FDCurrent15_6)
            FDCurrent15_6.pop(indice+1)
            indice=np.argmax(FDCurrent15_6)
            FDCurrent15_6.pop(indice-1)
        if(len(DAT15Voltage_6)>4):
            DAT15Voltage_6.sort()
            indice=np.argmin(DAT15Voltage_6)
            DAT15Voltage_6.pop(indice+1)
            indice=np.argmax(DAT15Voltage_6)
            DAT15Voltage_6.pop(indice-1)
        if(len(DAT15Current_6)>4):
            DAT15Current_6.sort()
            indice=np.argmin(DAT15Current_6)
            DAT15Current_6.pop(indice+1)
            indice=np.argmax(DAT15Current_6)
            DAT15Current_6.pop(indice-1)

Access_9 = 0
MaxVoltage15_9=0.0
MeanVoltage15_9=0.0
MinVoltage15_9=0.0
MaxCurrent15_9=0.0
MeanCurrent15_9=0.0
MinCurrent15_9=0.0
MaxActivePower_9=0.0
MeanActivePower_9=0.0
MinActivePower_9=0.0
MaxReactivePower_9=0.0
MeanReactivePower_9=0.0
MinReactivePower_9=0.0
MaxAparentPower_9=0.0
MeanAparentPower_9=0.0
MinAparentPower_9=0.0
MaxFPInductive_9=-0.99
MeanFPInductive_9=-0.99
MinFPInductive_9=-0.99
MaxFPReactive_9=0.99
MeanFPReactive_9=0.99
MinFPReactive_9=0.99
MaxFD_9=0.0
MeanFD_9=0.0
MinFD_9=0.0
MaxDAT_9=0.0
MeanDAT_9=0.0
MinDAT_9=0.0
Volt15_9=[]
data15_9=[]
Current15_9=[]
ActivePower15_9=[]
ReactivePower15_9=[]
AparentPower15_9=[]
FP15_Reactive_9=[]
FP15_Inductive_9=[]
FDVoltage15_9=[]
FDCurrent15_9=[]
DAT15Voltage_9=[]
DAT15Current_9=[]
def Maximo15min_9(Vrms,Irms,ActivePower,ReactivePower,AparentPower,FP,FDVoltage,FDCurrent,DATVoltage,DATCurrent,OneHourEnergy,Energy,i,k,f):
    global data15_9
    global Volt15_9
    global data15_9
    global Current15_9
    global ActivePower15_9
    global ReactivePower15_9
    global AparentPower15_9
    global FP15_Reactive_9
    global FP15_Inductive_9
    global FDVoltage15_9
    global FDCurrent15_9
    global DAT15Voltage_9
    global DAT15Current_9
    global Access_9
    global OneHourEnergy_9
    global optionsave
    basea = datetime.datetime.now()
    if(basea.minute==0 or basea.minute==1 or basea.minute==2 or basea.minute==15 or basea.minute==16 or basea.minute==17 or basea.minute==30 or basea.minute==31 or basea.minute==32 or basea.minute==45 or basea.minute==46 or basea.minute==47):
               if(Access_9 == 0):
                    #graphVoltage(NoVoltageoffset2,ListaIrmsPeak2,samplings2,2)
                    Access_9 = 1
                    MaxVoltage15_9=max(Volt15_9)
                    MeanVoltage15_9=np.median(Volt15_9)
                    MinVoltage15_9=min(Volt15_9)
                    MaxCurrent15_9=max(Current15_9)
                    MeanCurrent15_9=np.median(Current15_9)
                    MinCurrent15_9=min(Current15_9)
                    MaxActivePower_9=max(ActivePower15_9)
                    MeanActivePower_9=np.median(ActivePower15_9)
                    MinActivePower_9=min(ActivePower15_9)
                    MaxReactivePower_9=max(ReactivePower15_9)
                    MeanReactivePower_9=np.median(ReactivePower15_9)
                    MinReactivePower_9=min(ReactivePower15_9)
                    MaxAparentPower_9=max(AparentPower15_9)
                    MeanAparentPower_9=np.median(AparentPower15_9)
                    MinAparentPower_9=min(AparentPower15_9)
                    if(len(FP15_Inductive_9)>0):
                           MaxFPInductive_9=max(FP15_Inductive_9)
                           MeanFPInductive_9=np.median(FP15_Inductive_9)
                           MinFPInductive_9=min(FP15_Inductive_9)
                    else:
                           MaxFPInductive_9=-0.99
                           MeanFPInductive_9=-0.99
                           MinFPInductive_9=-0.99
                    if(len(FP15_Reactive_9)>0):
                           MaxFPReactive_9=max(FP15_Reactive_9)
                           MeanFPReactive_9=np.median(FP15_Reactive_9)
                           MinFPReactive_9=min(FP15_Reactive_9)
                    else:
                           MaxFPReactive_9=0.99
                           MeanFPReactive_9=0.99
                           MinFPReactive_9=0.99
                    MaxFDVoltage_9=max(FDVoltage15_9)
                    MeanFDVoltage_9=np.median(FDVoltage15_9)
                    MinFDVoltage_9=min(FDVoltage15_9)
                    MaxFDCurrent_9=max(FDCurrent15_9)
                    MeanFDCurrent_9=np.median(FDCurrent15_9)
                    MinFDCurrent_9=min(FDCurrent15_9)
                    MaxDATVoltage_9=max(DAT15Voltage_9)
                    MeanDATVoltage_9=np.median(DAT15Voltage_9)
                    MinDATVoltage_9=min(DAT15Voltage_9)
                    MaxDATCurrent_9=max(DAT15Current_9)
                    MeanDATCurrent_9=np.median(DAT15Current_9)
                    MinDATCurrent_9=min(DAT15Current_9)
                    data15_9.insert(1,MaxVoltage15_9)
                    data15_9.insert(2,MeanVoltage15_9)
                    data15_9.insert(3,MinVoltage15_9)
                    data15_9.insert(4,MaxCurrent15_9)
                    data15_9.insert(5,MeanCurrent15_9)
                    data15_9.insert(6,MinCurrent15_9)
                    data15_9.insert(7,MaxActivePower_9)
                    data15_9.insert(8,MeanActivePower_9)
                    data15_9.insert(9,MinActivePower_9)
                    data15_9.insert(10,MaxReactivePower_9)
                    data15_9.insert(11,MeanReactivePower_9)
                    data15_9.insert(12,MinReactivePower_9)
                    data15_9.insert(13,MaxAparentPower_9)
                    data15_9.insert(14,MeanAparentPower_9)
                    data15_9.insert(15,MinAparentPower_9)
                    data15_9.insert(16,MaxFPInductive_9)
                    data15_9.insert(17,MeanFPInductive_9)
                    data15_9.insert(18,MinFPInductive_9)
                    data15_9.insert(19,MaxFPReactive_9)
                    data15_9.insert(20,MeanFPReactive_9)
                    data15_9.insert(21,MinFPReactive_9)
                    data15_9.insert(22,MaxFDVoltage_9)
                    data15_9.insert(23,MeanFDVoltage_9)
                    data15_9.insert(24,MinFDVoltage_9)
                    data15_9.insert(25,MaxFDCurrent_9)
                    data15_9.insert(26,MeanFDCurrent_9)
                    data15_9.insert(27,MinFDCurrent_9)
                    data15_9.insert(28,MaxDATVoltage_9)
                    data15_9.insert(29,MeanDATVoltage_9)
                    data15_9.insert(30,MinDATVoltage_9)
                    data15_9.insert(31,MaxDATCurrent_9)
                    data15_9.insert(32,MeanDATCurrent_9)
                    data15_9.insert(33,MinDATCurrent_9)
                    data15_9.insert(34,OneHourEnergy_9)
                    data15_9.insert(35,Energy)
                    data15_9.insert(0,datetime.datetime.now())
                    workbook=openpyxl.load_workbook(filename = dest_filename)
                    sheet10 = workbook[f"Mean-{k}-{f}"] 
                    sheet10.append(list(data15_9))
                    print(f'Data 9: Guardando Promedios')
                    optionsave=1
                    SendDataToBroker(q=i,k=k,f=f,VoltajeMax=f'{MaxVoltage15_9}',VoltajePromedio=f'{MeanVoltage15_9}',VoltajeMin=f'{MinVoltage15_9}',MaxCorriente=f'{MaxCurrent15_9}',PromedioCorriente=f'{MeanCurrent15_9}',MinimoCorriente=f'{MinCurrent15_9}',PotenciaMax=f'{MaxAparentPower_9}',PromedioPotenciaAparente=f'{MeanAparentPower_9}',MinPotenciaAparente=f'{MinAparentPower_9}',OneHourEnergy=f'{OneHourEnergy}',Energia=f'{Energy}')
                    #print("Datos Insertados Correctamente!")
                    workbook.save(filename = dest_filename)
                    data15_9=[]
                    Volt15_9=[]
                    Current15_9=[]
                    ActivePower15_9=[]
                    ReactivePower15_9=[]
                    AparentPower15_9=[]
                    FP15_Reactive_9=[]
                    FP15_Inductive_9=[]
                    FDVoltage15_9=[]
                    FDCurrent15_9=[]
                    DAT15Voltage_9=[]
                    DAT15Current_9=[]
                    #OneHourEnergy_9=0
               elif(Access_9==1):
                    #print("paso elif 2")
                    Volt15_9.append(Vrms)
                    Current15_9.append(Irms)
                    ActivePower15_9.append(ActivePower)
                    ReactivePower15_9.append(ReactivePower)
                    AparentPower15_9.append(AparentPower)
                    if(FP>0.0):
                          FP15_Reactive_9.append(FP)
                    else: 
                          FP15_Inductive_9.append(FP)
                    FDVoltage15_9.append(FDVoltage)
                    FDCurrent15_9.append(FDCurrent)
                    DAT15Voltage_9.append(DATVoltage)
                    DAT15Current_9.append(DATCurrent)
              
    else:
        Volt15_9.append(Vrms)
        Current15_9.append(Irms)
        ActivePower15_9.append(ActivePower)
        ReactivePower15_9.append(ReactivePower)
        AparentPower15_9.append(AparentPower)
        if(FP>0.0):
              FP15_Reactive_9.append(FP)
        else: 
              FP15_Inductive_9.append(FP)
        FDVoltage15_9.append(FDVoltage)
        FDCurrent15_9.append(FDCurrent)
        DAT15Voltage_9.append(DATVoltage)
        DAT15Current_9.append(DATCurrent)
        Access_9 = 0
        
        if(len(Volt15_9)>4):
            Volt15_9.sort()
            indice=np.argmin(Volt15_9)
            Volt15_9.pop(indice+1)
            indice=np.argmax(Volt15_9)
            Volt15_9.pop(indice-1)
        if(len(Current15_9)>4):
            Current15_9.sort()
            indice=np.argmin(Current15_9)
            Current15_9.pop(indice+1)
            indice=np.argmax(Current15_9)
            Current15_9.pop(indice-1)
        if(len(ActivePower15_9)>4):
            ActivePower15_9.sort()
            indice=np.argmin(ActivePower15_9)
            ActivePower15_9.pop(indice+1)
            indice=np.argmax(ActivePower15_9)
            ActivePower15_9.pop(indice-1)
        if(len(ReactivePower15_9)>4):
            ReactivePower15_9.sort()
            indice=np.argmin(ReactivePower15_9)
            ReactivePower15_9.pop(indice+1)
            indice=np.argmax(ReactivePower15_9)
            ReactivePower15_9.pop(indice-1)
        if(len(AparentPower15_9)>4):
            AparentPower15_9.sort()
            indice=np.argmin(AparentPower15_9)
            AparentPower15_9.pop(indice+1)
            indice=np.argmax(AparentPower15_9)
            AparentPower15_9.pop(indice-1)
        if(len(FP15_Reactive_9)>4):
            FP15_Reactive_9.sort()
            indice=np.argmin(FP15_Reactive_9)
            FP15_Reactive_9.pop(indice+1)
            indice=np.argmax(FP15_Reactive_9)
            FP15_Reactive_9.pop(indice-1)
        if(len(FP15_Inductive_9)>4):
            FP15_Inductive_9.sort()
            indice=np.argmin(FP15_Inductive_9)
            FP15_Inductive_9.pop(indice+1)
            indice=np.argmax(FP15_Inductive_9)
            FP15_Inductive_9.pop(indice-1)
        if(len(FDVoltage15_9)>4):
            FDVoltage15_9.sort()
            indice=np.argmin(FDVoltage15_9)
            FDVoltage15_9.pop(indice+1)
            indice=np.argmax(FDVoltage15_9)
            FDVoltage15_9.pop(indice-1)
        if(len(FDCurrent15_9)>4):
            FDCurrent15_9.sort()
            indice=np.argmin(FDCurrent15_9)
            FDCurrent15_9.pop(indice+1)
            indice=np.argmax(FDCurrent15_9)
            FDCurrent15_9.pop(indice-1)
        if(len(DAT15Voltage_9)>4):
            DAT15Voltage_9.sort()
            indice=np.argmin(DAT15Voltage_9)
            DAT15Voltage_9.pop(indice+1)
            indice=np.argmax(DAT15Voltage_9)
            DAT15Voltage_9.pop(indice-1)
        if(len(DAT15Current_9)>4):
            DAT15Current_9.sort()
            indice=np.argmin(DAT15Current_9)
            DAT15Current_9.pop(indice+1)
            indice=np.argmax(DAT15Current_9)
            DAT15Current_9.pop(indice-1)

def excelcreate():
    global dest_filename
    global sheet1
    global sheet2
    global sheet3
    global sheet4
    global sheet5
    global sheet6
    global sheet7
    global sheet8
    global sheet9
    global sheet10
    global sheet11
    global sheet12
    global sheet13
    global sheet14
    global sheet15
    global sheet16
    global sheet17
    global sheet18
    global sheet19
    global sheet20
    global sheet21
    global sheet22
    from openpyxl import Workbook
    exceltime=date.today()
    book = Workbook()
    dest_filename = f'{exceltime}.xlsx'
    #sheet1 = book.active
    sheet20 = book.create_sheet(f"MaxHora Fase 1 Diario") 
    sheet21 = book.create_sheet(f"MaxHora Fase 2 Diario") 
    sheet22 = book.create_sheet(f"MaxHora Fase 3 Diario") 
    sheet23 = book.create_sheet(f"MaxHora Fase 1 Mensual") 
    sheet24 = book.create_sheet(f"MaxHora Fase 2 Mensual") 
    sheet25 = book.create_sheet(f"MaxHora Fase 3 Mensual") 
    sheet2 = book.create_sheet(f"15Min-{k1}-{f1}")
    sheet3 = book.create_sheet(f"15Min-{k1}-{f2}")
    sheet4 = book.create_sheet(f"15Min-{k1}-{f3}")
    sheet5 = book.create_sheet(f"15Min-{k2}-{f1}")
    sheet6 = book.create_sheet(f"15Min-{k2}-{f2}")
    sheet7 = book.create_sheet(f"15Min-{k2}-{f3}")
    sheet8 = book.create_sheet(f"15Min-{k3}-{f1}")
    sheet9 = book.create_sheet(f"15Min-{k3}-{f2}")
    sheet10 = book.create_sheet(f"15Min-{k3}-{f3}")
    sheet11 = book.create_sheet(f"{k1}-{f1}") 
    sheet12 = book.create_sheet(f"{k1}-{f2}") 
    sheet13 = book.create_sheet(f"{k1}-{f3}") 
    sheet14 = book.create_sheet(f"{k2}-{f1}") 
    sheet15 = book.create_sheet(f"{k2}-{f2}") 
    sheet16 = book.create_sheet(f"{k2}-{f3}") 
    sheet17 = book.create_sheet(f"{k3}-{f1}") 
    sheet18 = book.create_sheet(f"{k3}-{f2}") 
    sheet19 = book.create_sheet(f"{k3}-{f3}") 
    sheet1  = book.create_sheet("Var Dispositivos")
    headings0 = ['Fecha y Hora'] + list(['T° Raspberry','Uso CPU %','RAM','T° ESP32'])
    headings=['Fecha y Hora'] + list(['Max Voltage','Mean Voltage','Min Voltage', 'Max Current','Mean Current','Min Current','Max Active Power','Mean Active Power','Mean Active Power','Max Reactive Power','Mean Reactive Power','Min Reactive Power','Max Aparent Power','Mean Aparent Power','Min Aparent Power','Max FPReact ','Mean FPReact','Min FPReact','Max FPInduct','Mean FPInduct','Min FPInduct','Max FDVoltage','Mean FDVoltage','Min FDVoltage','Max FDCurrent','Mean FDCurrent','Min FDCurrent','Max DATVoltage','Mean DATVoltage','Min DATVoltage','Max DATCurrent','Mean DATCurrent','Min DATCurrent','OnehourEnergy','Energy'])
    headings2=['Fecha y Hora'] + list(['Voltage', 'Current','Active Power','Reactive Power','Aparent Power','FP','FDVoltage','FDCurrent','DATVoltage','DATCurrent','cos(phi)','Energy','Hour Energy'])
    headings3=['Fecha y Hora'] + list(['Voltage', 'Current','Power','Energy','Hour Energy'])
    headings4=['Fecha y Hora'] + list(['Max Voltage', 'Mean Voltage', 'Min Voltage', 'Max Current','Mean Current', 'Min Current','Max Power','Power Mean', 'Power','Total Energy','Energy acumulada en 15'])
    headings5 = ['Fecha y Hora'] + list([f"{k1}-{f1}",f"{k1}-{f2}",f"{k1}-{f3}",f"{k2}-{f1}",f"{k2}-{f2}",f"{k2}-{f3}",f"{k3}-{f1}",f"{k3}-{f2}",f"{k3}-{f3}"])
    headingsFase1=list(['Hora','Energia-Fase-1-REDCompañia-Diario', 'Energia-Fase-1-CentralFotovoltaica-Diario','Energia-Fase-1-ConsumoCliente-Diario'])
    headingsFase2=list(['Hora','Energia-Fase-2-REDCompañia-Diario', 'Energia-Fase-2-CentralFotovoltaica-Diario','Energia-Fase-2-ConsumoCliente-Diario'])
    headingsFase3=list(['Hora','Energia-Fase-3-REDCompañia-Diario', 'Energia-Fase-3-CentralFotovoltaica-Diario','Energia-Fase-3-ConsumoCliente-Diario'])
    headingsFase1_Mensual=list(['Hora','Energia-Fase-1-REDCompañia-Mensual', 'Energia-Fase-1-CentralFotovoltaica-Mensual','Energia-Fase-1-ConsumoCliente-Mensual'])
    headingsFase2_Mensual=list(['Hora','Energia-Fase-2-REDCompañia-Mensual', 'Energia-Fase-2-CentralFotovoltaica-Mensual','Energia-Fase-2-ConsumoCliente-Mensual'])
    headingsFase3_Mensual=list(['Hora','Energia-Fase-3-REDCompañia-Mensual', 'Energia-Fase-3-CentralFotovoltaica-Mensual','Energia-Fase-3-ConsumoCliente-Mensual'])
    ceros=list([0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])
    sheet20.append(headingsFase1)
    sheet21.append(headingsFase2)
    sheet22.append(headingsFase3)
    sheet23.append(headingsFase1_Mensual)
    sheet24.append(headingsFase2_Mensual)
    sheet25.append(headingsFase3_Mensual)
    sheet1.append(headings0)
    sheet2.append(headings)
    sheet3.append(headings)
    sheet4.append(headings)
    sheet5.append(headings)
    sheet6.append(headings)
    sheet7.append(headings)
    sheet8.append(headings)
    sheet9.append(headings)
    sheet10.append(headings)
    sheet11.append(headings2)
    sheet12.append(headings2)
    sheet13.append(headings2)
    sheet14.append(headings2)
    sheet15.append(headings2)
    sheet16.append(headings2)
    sheet17.append(headings2)
    sheet18.append(headings2)
    sheet19.append(headings2)
    sheet1.append(list([0,0,0,0,0]))
    sheet2.append(ceros)
    sheet3.append(ceros)
    sheet4.append(ceros)
    sheet5.append(ceros)
    sheet6.append(ceros)
    sheet7.append(ceros)
    sheet8.append(ceros)
    sheet9.append(ceros)
    sheet10.append(ceros)
    sheet11.append(ceros)
    sheet12.append(ceros)
    sheet13.append(ceros)
    sheet14.append(ceros)
    sheet15.append(ceros)
    sheet16.append(ceros)
    sheet17.append(ceros)
    sheet18.append(ceros)
    sheet19.append(ceros)

    book.save(filename = dest_filename)



def AbrirExcel():
    global dest_filename
    global Energy_1
    global Energy_2
    global Energy_3
    global Energy_4
    global Energy_5
    global Energy_6
    global Energy_7
    global Energy_8
    global Energy_9
    
    dia=date.today()
    if(os.path.exists(f'{dia}.xlsx')):
            dest_filename = f'{dia}.xlsx'
            print("Existe")
            workbook=openpyxl.load_workbook(filename = dest_filename)
            sheet11 = workbook[f"{k1}-{f1}"]
            sheet12 = workbook[f"{k1}-{f2}"]
            sheet13 = workbook[f"{k1}-{f3}"]
            sheet14 = workbook[f"{k2}-{f1}"]
            sheet15 = workbook[f"{k2}-{f2}"]
            sheet16 = workbook[f"{k2}-{f3}"]
            sheet17 = workbook[f"{k3}-{f1}"]
            sheet18 = workbook[f"{k3}-{f2}"]
            sheet19 = workbook[f"{k3}-{f3}"]
            LargeSheet11=len(sheet11["FP"])
            LargeSheet12=len(sheet12["FP"])
            LargeSheet13=len(sheet13["FP"])
            LargeSheet14=len(sheet14["FP"])
            LargeSheet15=len(sheet15["FP"])
            LargeSheet16=len(sheet16["FP"])
            LargeSheet17=len(sheet17["FP"])
            LargeSheet18=len(sheet18["FP"])
            LargeSheet19=len(sheet19["FP"])
            print("Largo Excel Var 1: ",LargeSheet11)
            Energy_1 = float(sheet11[f'm{LargeSheet11}'].value)
            Energy_2 = float(sheet12[f'm{LargeSheet12}'].value)
            Energy_3 = float(sheet13[f'm{LargeSheet13}'].value)
            Energy_4 = float(sheet14[f'm{LargeSheet14}'].value)
            Energy_5 = float(sheet15[f'm{LargeSheet15}'].value)
            Energy_6 = float(sheet16[f'm{LargeSheet16}'].value)
            Energy_7 = float(sheet17[f'm{LargeSheet17}'].value)
            Energy_8 = float(sheet18[f'm{LargeSheet18}'].value)
            Energy_9 = float(sheet19[f'm{LargeSheet19}'].value)
            #energyBaterias = float(sheet8[f'k{largoexcelCGE-2}'].value)
            #energyPanelesDC = float(sheet10[f'k{largoexcelCGE-2}'].value)
            print(f'Valor Energia  {Energy_1} ')
    else:
            excelcreate()
            print("No Existe")

AbrirExcel()


def VariablesBasicas(Temp_Raspberry,cpu_uso,RAM,tempESP32):                   
       workbook=openpyxl.load_workbook(filename = dest_filename)
       sheet1 = workbook["Var Dispositivos"]
       Data=[datetime.datetime.now(),Temp_Raspberry,cpu_uso,RAM,tempESP32]
       sheet1.append(list(Data))
       workbook.save(filename = dest_filename)
       Data=[]

def SaveDataCsv(Vrms,Irms,ActivePower_1,ReactivePower_1,AparentPower_1,FP_1,CosPhi_1,FDVoltage_1,FDCurrent_1,DATVoltage_1,DATCurrent_1,Energy_1,OneHourEnergy_1,i,k,f):
       Data=[datetime.datetime.now(),round(Vrms,2), round(Irms,2), round(ActivePower_1,2), round(ReactivePower_1,2), round(AparentPower_1,2), round(FP_1,2), round(CosPhi_1,2), round(FDVoltage_1,2), round(FDCurrent_1,2), round(DATVoltage_1,2), round(DATCurrent_1,2), round(Energy_1,5), round(OneHourEnergy_1,5)]                    
       workbook=openpyxl.load_workbook(filename = dest_filename)
       if(i==1):
             sheet11 = workbook[f"{k}-{f}"]
             sheet11.append(list(Data))
       elif(i==2):
             sheet12 = workbook[f"{k}-{f}"]
             sheet12.append(list(Data))
       elif(i==3):
             sheet13 = workbook[f"{k}-{f}"]
             sheet13.append(list(Data))
       elif(i==4):
             sheet14 = workbook[f"{k}-{f}"]
             sheet14.append(list(Data))
       elif(i==5):
             sheet15 = workbook[f"{k}-{f}"]
             sheet15.append(list(Data))
       elif(i==6):
             sheet16 = workbook[f"{k}-{f}"]
             sheet16.append(list(Data))
       elif(i==7):
             sheet17 = workbook[f"{k}-{f}"]
             sheet17.append(list(Data))
       elif(i==8):
             sheet18 = workbook[f"{k}-{f}"]
             sheet18.append(list(Data))
       elif(i==9):
             sheet19 = workbook[f"{k}-{f}"]
             sheet19.append(list(Data))
       
       
       workbook.save(filename = dest_filename)
       Data=[]


import os
import base64
import requests
from azure.identity import InteractiveBrowserCredential
from msgraph.core import GraphClient

def SendEmail():
    
    
    def draft_attachment(file_path):
        if not os.path.exists(file_path):
            print('file is not found')
            return
        
        with open(file_path, 'rb') as upload:
            media_content = base64.b64encode(upload.read())
            
        data_body = {
            '@odata.type': '#microsoft.graph.fileAttachment',
            'contentBytes': media_content.decode('utf-8'),
            'name': os.path.basename(file_path)
        }
        return data_body
    
    APP_ID = '<app id>'
    SCOPES = ['Mail.Send', 'Mail.ReadWrite']
    
    access_token = generate_access_token(app_id=APP_ID, scopes=SCOPES)
    headers = {
        'Authorization': 'Bearer ' + access_token['access_token']
    }
    
    request_body = {
        'message': {
            # recipient list
            'toRecipients': [
                {
                    'emailAddress': {
                        'address': '<recipient email address>'
                    }
                }
            ],
            # email subject
            'subject': 'You got an email',
            'importance': 'normal',
            'body': {
                'contentType': 'HTML',
                'content': '<b>Be Awesome</b>'
            },
            # include attachments
            'attachments': [
                draft_attachment('hello.txt'),
                draft_attachment('image.png')
            ]
        }
    }

    GRAPH_ENDPOINT = 'https://graph.microsoft.com/v1.0'
    endpoint = GRAPH_ENDPOINT + '/me/sendMail'
    
    response = requests.post(endpoint, headers=headers, json=request_body)
    if response.status_code == 202:
        print('Email sent')
    else:
        print(response.reason)

def SendEmail2():
    #global dest_filename
    Lugar="Santa Cristina"
    username = "empresasserspa@gmail.com"
    password = "nbqpiwiwootrwffu" #"empresasserspa"
    destinatario = "ricardovera.93@hotmail.com"
    #destinatario2 = "ricardovera.93@hotmail.com"
    #destinatario2 = "demetrio.vera@serm.cl"
    
    mensaje = MIMEMultipart("Alternative")
    mensaje["Subject"] = "Reportes "+str(Lugar)+" "+str(datetime.date.today())
    mensaje["From"] = username
    mensaje["To"] = destinatario
    
    html = f"""
    <html>
    <body>
         <p> Hola <i>{destinatario}</i> <br>
         Reportes desde {Lugar} </b>
    </body>
    </html>
    """
    
    parte_html = MIMEText(html, "html")
    mensaje.attach(parte_html)
    
    archivo = dest_filename
    
    with open(archivo, "rb") as adjunto:
         contenido_adjunto = MIMEBase("application", "octet-stream")
         contenido_adjunto.set_payload(adjunto.read())
    
    encoders.encode_base64(contenido_adjunto)
    
    contenido_adjunto.add_header(
         "Content-Disposition",
         f"attachment; filename= {archivo}",
    )

    mensaje.attach(contenido_adjunto)
    text = mensaje.as_string()
    
    
    context = ssl.create_default_context()

    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
          server.login(username,password)
          print("Sesión Iniciada Correctamente !")
          #server.sendmail(username, destinatario, mensaje)
          server.sendmail(username, destinatario, text)
          #server.sendmail(username, destinatario2, text)
          print("Mensaje Enviado Correctamente !")


"""
while client.connected_flag: 
    #print("In loop")
    #received()
    publish(client)
    #time.sleep(5)
"""   


fecha=str(datetime.datetime.now())

f = open('mi_fichero2.txt', 'w')
try:
    f.write(fecha)
finally:
    f.close()

CurrentCal=0.91
Vrms=0.0
Irms=0.0
BufferVoltaje_1=[]
BufferVoltaje_2=[]
BufferVoltaje_3=[]
BufferVoltaje_4=[]
BufferVoltaje_5=[]
BufferVoltaje_6=[]
BufferVoltaje_7=[]
BufferVoltaje_8=[]
BufferVoltaje_9=[]
BufferCurrent_1=[]
BufferCurrent_2=[]
BufferCurrent_3=[]
BufferCurrent_4=[]
BufferCurrent_5=[]
BufferCurrent_6=[]
BufferCurrent_7=[]
BufferCurrent_8=[]
BufferCurrent_9=[]
global Access_1email
Access_1email=0
global Access_1excel
Access_1xcel=0
countbroker=0

def TomaDatos(list_Voltage,list_Current,samplings,i):
    global BufferVoltaje_1
    global BufferVoltaje_2
    global BufferVoltaje_3
    global BufferVoltaje_4
    global BufferVoltaje_5
    global BufferVoltaje_6
    global BufferVoltaje_7
    global BufferVoltaje_8
    global BufferVoltaje_9
    global BufferCurrent_1
    global BufferCurrent_2
    global BufferCurrent_3
    global BufferCurrent_4
    global BufferCurrent_5
    global BufferCurrent_6
    global BufferCurrent_7
    global BufferCurrent_8
    global BufferCurrent_9
    global Vrms
    global Irms
    global NoVoltageOffset
    global NoCurrentoffset
    global ListaIrmsPeak1
                           
    sos = signal.butter(10, 2500, 'low', fs=samplings, output='sos')
    list_VoltageFilterComplete = signal.sosfilt(sos, list_Voltage)
    list_CurrentFilterComplete = signal.sosfilt(sos, list_Current)
    list_FinalVoltage = list_VoltageFilterComplete[104:4200]
    list_FinalCurrent = list_CurrentFilterComplete [103:4200]
    #print(f'list_FinalCurrent {i} Max: {max(list_FinalCurrent)}')

    List_MaxVoltage=getMaxValues(list_FinalVoltage, 100)
    List_MinVoltage=getMinValues(list_FinalVoltage, 100)
    MaxVoltage = np.min(List_MaxVoltage)
    MinVoltage = np.max(List_MinVoltage)
    DC_VoltageMedian = (MaxVoltage+MinVoltage)/2
    NoVoltageOffset=(list_FinalVoltage-DC_VoltageMedian)             
    Vrms=VoltajeRms(NoVoltageOffset)*0.92
    if(i==1):
        if (len(BufferVoltaje_1)>=5):
            MediaBufferVoltaje=np.median(BufferVoltaje_1)
            Vrms=VoltRms(MediaBufferVoltaje)
            if(Vrms>240):
                print(f"Mayor a 240 {i}")
            else:
                print(f'Vrms {i}: {Vrms}')
                VoltageFFT(NoVoltageOffset,samplings,i)
            BufferVoltaje_1=[]
        else:
            BufferVoltaje_1.append(Vrms)
    elif(i==2):
        if (len(BufferVoltaje_2)>=5):
            MediaBufferVoltaje=np.median(BufferVoltaje_2)
            Vrms=VoltRms(MediaBufferVoltaje)
            if(Vrms>240):
                print(f"Mayor a 240 {i}")
            else:
                print(f'Vrms {i}: {Vrms}')
                VoltageFFT(NoVoltageOffset,samplings,i)
            BufferVoltaje_2=[]
        else:
            BufferVoltaje_2.append(Vrms)
    elif(i==3):
        if (len(BufferVoltaje_3)>=5):
            MediaBufferVoltaje=np.median(BufferVoltaje_3)
            Vrms=VoltRms(MediaBufferVoltaje)
            if(Vrms>240):
                print(f"Mayor a 240 {i}")
            else:
                print(f'Vrms {i}: {Vrms}')
                VoltageFFT(NoVoltageOffset,samplings,i)
            BufferVoltaje_3=[]
        else:
            BufferVoltaje_3.append(Vrms)
    elif(i==4):
        if (len(BufferVoltaje_4)>=5):
            MediaBufferVoltaje=np.median(BufferVoltaje_4)
            Vrms=VoltRms(MediaBufferVoltaje)
            if(Vrms>240):
                print(f"Mayor a 240 {i}")
            else:
                print(f'Vrms {i}: {Vrms}')
                VoltageFFT(NoVoltageOffset,samplings,i)
            BufferVoltaje_4=[]
        else:
            BufferVoltaje_4.append(Vrms)
    elif(i==5):
        if (len(BufferVoltaje_5)>=5):
            MediaBufferVoltaje=np.median(BufferVoltaje_5)
            Vrms=VoltRms(MediaBufferVoltaje)
            if(Vrms>240):
                print(f"Mayor a 240 {i}")
            else:
                print(f'Vrms {i}: {Vrms}')
                VoltageFFT(NoVoltageOffset,samplings,i)
            BufferVoltaje_5=[]
        else:
            BufferVoltaje_5.append(Vrms)
    elif(i==6):
        if (len(BufferVoltaje_6)>=5):
            MediaBufferVoltaje=np.median(BufferVoltaje_6)
            Vrms=VoltRms(MediaBufferVoltaje)
            if(Vrms>240):
                print(f"Mayor a 240 {i}")
            else:
                print(f'Vrms {i}: {Vrms}')
                VoltageFFT(NoVoltageOffset,samplings,i)
            BufferVoltaje_6=[]
        else:
            BufferVoltaje_6.append(Vrms)
    elif(i==7):
        if (len(BufferVoltaje_7)>=5):
            MediaBufferVoltaje=np.median(BufferVoltaje_7)
            Vrms=VoltRms(MediaBufferVoltaje)
            if(Vrms>240):
                print(f"Mayor a 240 {i}")
            else:
                print(f'Vrms {i}: {Vrms}')
                VoltageFFT(NoVoltageOffset,samplings,i)
            BufferVoltaje_7=[]
        else:
            BufferVoltaje_7.append(Vrms)
    elif(i==8):
        if (len(BufferVoltaje_8)>=5):
            MediaBufferVoltaje=np.median(BufferVoltaje_8)
            Vrms=VoltRms(MediaBufferVoltaje)
            if(Vrms>240):
                print(f"Mayor a 240 {i}")
            else:
                print(f'Vrms {i}: {Vrms}')
                VoltageFFT(NoVoltageOffset,samplings,i)
            BufferVoltaje_8=[]
        else:
            BufferVoltaje_8.append(Vrms)
    elif(i==9):
        if (len(BufferVoltaje_9)>=5):
            MediaBufferVoltaje=np.median(BufferVoltaje_9)
            Vrms=VoltRms(MediaBufferVoltaje)
            if(Vrms>240):
                print(f"Mayor a 240 {i}")
            else:
                print(f'Vrms {i}: {Vrms}')
                VoltageFFT(NoVoltageOffset,samplings,i)
            BufferVoltaje_9=[]
        else:
            BufferVoltaje_9.append(Vrms)
                        

    List_MaxCurrent=getMaxValues(list_FinalCurrent, 50)
    List_MinCurrent=getMinValues(list_FinalCurrent, 50)
    MaxCurrent = np.median(List_MaxCurrent)
    Mincurrent = np.median(List_MinCurrent)
    DC_CurrentMedian = (MaxCurrent+Mincurrent)/2
    NoCurrentoffset=list_FinalCurrent-DC_CurrentMedian
    Irms=CorrienteRms(NoCurrentoffset)
                               
    if(i==1):
        if(len(BufferCurrent_1)>=5 and Vrms<240):
            MediaBufferCurrent=np.median(BufferCurrent_1)
            Irms=CurrentRms(MediaBufferCurrent)*CurrentCal
            print(f'Irms {i}: {Irms}')
            CurrentFFT(NoCurrentoffset,samplings,i,Irms)
            potrmsCGE = PotenciaRms(NoCurrentoffset,NoVoltageOffset)
            Potencias(i,Irms,Vrms,potrmsCGE)     
            BufferCurrent_1=[]
        else:
            BufferCurrent_1.append(Irms)
    elif(i==2):
        if(len(BufferCurrent_2)>=5 and Vrms<240):
            MediaBufferCurrent=np.median(BufferCurrent_2)
            Irms=CurrentRms(MediaBufferCurrent)*CurrentCal
            print(f'Irms {i}: {Irms}')
            #print(f'Irms {i} Max: {max(NoCurrentoffset)}')
            CurrentFFT(NoCurrentoffset,samplings,i,Irms)
            potrmsCGE = PotenciaRms(NoCurrentoffset,NoVoltageOffset)
            Potencias(i,Irms,Vrms,potrmsCGE)     
            BufferCurrent_2=[]
        else:
            BufferCurrent_2.append(Irms)
    elif(i==3):
        if(len(BufferCurrent_3)>=5 and Vrms<240):
            MediaBufferCurrent=np.median(BufferCurrent_3)
            Irms=CurrentRms(MediaBufferCurrent)*CurrentCal
            #print(f'Current cal: {CurrentCal}')
            print(f'Irms {i}: {Irms}')
            #print(f'Irms {i} Max: {max(NoCurrentoffset)}')
            CurrentFFT(NoCurrentoffset,samplings,i,Irms)
            potrmsCGE = PotenciaRms(NoCurrentoffset,NoVoltageOffset)
            Potencias(i,Irms,Vrms,potrmsCGE)     
            BufferCurrent_3=[]
        else:
            BufferCurrent_3.append(Irms)
    elif(i==4):
        if(len(BufferCurrent_4)>=5 and Vrms<240):
            MediaBufferCurrent=np.median(BufferCurrent_4)
            Irms=CurrentRms(MediaBufferCurrent)*CurrentCal
            #print(f'Current cal: {CurrentCal}')
            print(f'Irms {i}: {Irms}')
            #print(f'Irms {i} Max: {max(NoCurrentoffset)}')
            CurrentFFT(NoCurrentoffset,samplings,i,Irms)
            potrmsCGE = PotenciaRms(NoCurrentoffset,NoVoltageOffset)
            Potencias(i,Irms,Vrms,potrmsCGE)     
            BufferCurrent_4=[]
        else:
            BufferCurrent_4.append(Irms)
    elif(i==5):
        if(len(BufferCurrent_5)>=5 and Vrms<240):
            MediaBufferCurrent=np.median(BufferCurrent_5)
            Irms=CurrentRms(MediaBufferCurrent)*CurrentCal
            print(f'Irms {i}: {Irms}')
            CurrentFFT(NoCurrentoffset,samplings,i,Irms)
            potrmsCGE = PotenciaRms(NoCurrentoffset,NoVoltageOffset)
            Potencias(i,Irms,Vrms,potrmsCGE)     
            BufferCurrent_5=[]
        else:
            BufferCurrent_5.append(Irms)
    elif(i==6):
        if(len(BufferCurrent_6)>=5 and Vrms<240):
            MediaBufferCurrent=np.median(BufferCurrent_6)
            Irms=CurrentRms(MediaBufferCurrent)*CurrentCal
            print(f'Irms {i}: {Irms}')
            CurrentFFT(NoCurrentoffset,samplings,i,Irms)
            potrmsCGE = PotenciaRms(NoCurrentoffset,NoVoltageOffset)
            Potencias(i,Irms,Vrms,potrmsCGE)     
            BufferCurrent_6=[]
        else:
            BufferCurrent_6.append(Irms)
    elif(i==7):
        if(len(BufferCurrent_7)>=5 and Vrms<240):
            MediaBufferCurrent=np.median(BufferCurrent_7)
            Irms=CurrentRms(MediaBufferCurrent)*CurrentCal
            print(f'Irms {i}: {Irms}')
            #print(f'Irms {i} Max: {max(NoCurrentoffset)}')
            CurrentFFT(NoCurrentoffset,samplings,i,Irms)
            potrmsCGE = PotenciaRms(NoCurrentoffset,NoVoltageOffset)
            Potencias(i,Irms,Vrms,potrmsCGE)     
            BufferCurrent_7=[]
        else:
            BufferCurrent_7.append(Irms)
    elif(i==8):
        if(len(BufferCurrent_8)>=5 and Vrms<240):
            MediaBufferCurrent=np.median(BufferCurrent_8)
            Irms=CurrentRms(MediaBufferCurrent)*CurrentCal
            print(f'Irms {i}: {Irms}')
            CurrentFFT(NoCurrentoffset,samplings,i,Irms)
            potrmsCGE = PotenciaRms(NoCurrentoffset,NoVoltageOffset)
            Potencias(i,Irms,Vrms,potrmsCGE)     
            BufferCurrent_8=[]
        else:
            BufferCurrent_8.append(Irms)
    elif(i==9):
        if(len(BufferCurrent_9)>=5 and Vrms<240):
            MediaBufferCurrent=np.median(BufferCurrent_9)
            Irms=CurrentRms(MediaBufferCurrent)*CurrentCal
            print(f'Irms {i}: {Irms}')
            CurrentFFT(NoCurrentoffset,samplings,i,Irms)
            potrmsCGE = PotenciaRms(NoCurrentoffset,NoVoltageOffset)
            Potencias(i,Irms,Vrms,potrmsCGE)     
            BufferCurrent_9=[]
        else:
            BufferCurrent_9.append(Irms)


def received():
    while True:
                 try:
                     esp32_bytes = esp32.readline()
                     decoded_bytes = str(esp32_bytes[0:len(esp32_bytes)-2].decode("utf-8"))#utf-8
                 except:
                     print("Error en la codificación")
                     continue
                 np_array = np.fromstring(decoded_bytes, dtype=float, sep=',')   
                 #print(f'Largo Array {len(np_array)}')
                 if (len(np_array) == 8402):
                       if (np_array[0] == 11 or np_array[0] == 22 or np_array[0] == 33 or np_array[0] == 44 or np_array[0] == 55 or np_array[0] == 66 or np_array[0] == 77 or np_array[0] == 88 or np_array[0] == 99):
                           if (np_array[0] == 11):
                               i = 1
                           elif (np_array[0] == 22):
                               i = 2
                           elif (np_array[0] == 33):
                               i = 3
                           elif (np_array[0] == 44):
                               i = 4
                           elif (np_array[0] == 55):
                               i = 5
                           elif (np_array[0] == 66):
                               i = 6
                           elif (np_array[0] == 77):
                               i = 7
                           elif (np_array[0] == 88):
                               i = 8
                           elif (np_array[0] == 99):
                               i = 9
                           samplings = np_array[-1]
                           list_Voltage = (np_array[0:4200])
                           list_Current = np_array[4201:8400]
                           try:
                                 TomaDatos(list_Voltage,list_Current,samplings,i)
                           except OSError as err:
                                 print("OS error: {0}".format(err))
                                 continue
                           except ValueError:
                                 print("Error: {ValueError}")
                                 continue
                 if (len(np_array)>0 and len(np_array)<=2):
                         Temp_Raspberry=cpu_temp()
                         cpu_uso=get_cpuload()
                         tempESP32 = np_array[0]
                         RAM = psutil.virtual_memory()[2]
                         VariablesBasicas(Temp_Raspberry,cpu_uso,RAM,tempESP32)
                         if (RAM > 93):
                              os.system("sudo reboot")
                         print(f"Temp_Raspberry: {Temp_Raspberry}")
                         #Temp_Raspberry_JSON = json.dumps(str_num)
                         #Ventilador()
                         #temphum()
                         #distance()
                         #str_num2 = {"value":tempESP320,"save":0}
                         #str_num = {"value":Temp_Raspberry0,"save":0}
                         #tempESP32 = json.dumps(str_num2)
                         

                 excel=datetime.datetime.now()
                 if(excel.hour==0 and excel.minute==4):
                          if(Access_1email==0):
                                 Access_1email=1
                                 print("Entro a SendEmail")
                                 #SendEmail()
                                 time.sleep(5)
                                 #os.remove(dest_filename)
                                 excelcreate()
                 else:
                          Access_1email=0
                          
                          
                 if(excel.minute==4 and excel.minute==19 or excel.minute==34 or excel.minute==49 ):
                     if(rcConnect > 5): 
                           print("Reiniciar")
                           os.system("sudo reboot")
                     else: 
                           print("Continue")
                        

        
if __name__ == '__main__':
    received()
    #t = threading.Thread(target=received)
    #t.daemon = True
    #t.start()

