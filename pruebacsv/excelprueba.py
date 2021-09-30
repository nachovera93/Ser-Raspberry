import requests
import datetime
import json
import random
import time
import paho.mqtt.client as mqtt
import os
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import openpyxl
import numpy as np
import smtplib, ssl
import getpass
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.mime.base import MIMEBase


broker = '18.228.175.193'    #mqtt server
port = 1883
dId = '123'
passw = 'SSf0SRnLkV'
#dId2 = '12344321'
#passw2 = 'yFJMESnzxl'
webhook_endpoint = 'http://18.228.175.193:3001/api/getdevicecredentials'


 
def get_mqtt_credentials():
    global usernamemqtt
    global passwordmqtt
    global mqttopic
    global str_client_id
    global topicmqtt
    global data
    print("Getting MQTT Credentials from WebHook")
    time.sleep(2)
    toSend = {"dId": dId, "password": passw}
    respuesta = requests.post(webhook_endpoint, data=toSend)

    if(respuesta.status_code < 0):
          print("Error Sending Post Request ", respuesta.status_code)
          respuesta.close()
          return False
    if(respuesta.status_code != 200):
          print("Error in response ", respuesta.status_code)
          respuesta.close()
          return False
    if(respuesta.status_code == 200):
          print("Mqtt Credentials Obtained Successfully :)   ")
          #print("json: " ,resp.content)
          #print('- ' * 20)
          my_bytes_value = respuesta.content      #Contenido entero del json
          my_new_string = my_bytes_value.decode("utf-8").replace("'", '"')
          data = json.loads(my_new_string)
          s = json.dumps(data, indent=4, sort_keys=True)
          #print(s)
          usernamemqtt = data["username"]
          #print("username:",usernamemqtt)
          passwordmqtt = data["password"]
          topicmqtt = data["topic"]   #topico al que nos vamos a suscribir
          mqttopic = f"{topicmqtt}+/actdata"
          str_client_id = f'device_{dId}_{random.randint(0, 9999)}'
          #print(mqttopic)
          respuesta.close()
          print("Ends mqtt credentials")
    return True


def get_mqtt_credentials2():
    global usernamemqtt2
    global passwordmqtt2
    global mqttopic2
    global str_client_id2
    global topicmqtt2
    global data2
    print("Getting MQTT Credentials from WebHook2")
    time.sleep(2)
    toSend2 = {"dId": dId2, "password": passw2}
    respuesta2 = requests.post(webhook_endpoint, data=toSend2)

    if(respuesta2.status_code < 0):
          print("Error Sending Post Request ", respuesta2.status_code)
          respuesta2.close()
          return False
    if(respuesta2.status_code != 200):
          print("Error in response ", respuesta2.status_code)
          respuesta2.close()
          return False
    if(respuesta2.status_code == 200):
          print("Mqtt Credentials 2 Obtained Successfully :)   ")
          #print("json: " ,resp.content)
          #print('- ' * 20)
          my_bytes_value2 = respuesta2.content      #Contenido entero del json
          my_new_string2 = my_bytes_value2.decode("utf-8").replace("'", '"')
          data2 = json.loads(my_new_string2)
          s2 = json.dumps(data2, indent=4, sort_keys=True)
          #print(s)
          usernamemqtt2 = data2["username"]
          #print("username:",usernamemqtt)
          passwordmqtt2 = data2["password"]
          topicmqtt2 = data2["topic"]   #topico al que nos vamos a suscribir
          mqttopic2 = f"{topicmqtt}+/actdata"
          str_client_id2 = f'device_{dId}_{random.randint(0, 9999)}'
          #print(mqttopic)
          respuesta2.close()
          print("Ends mqtt credentials 2")
    return True


#get_mqtt_credentials()
#get_mqtt_credentials2()
   
def on_disconnect(client, userdata, rc):
    if (rc != 0 and rc != 5):
        print("Unexpected disconnection, will auto-reconnect")
    elif(rc==5):
        print("Getting new credentials!")
        get_mqtt_credentials()
        client.username_pw_set(usernamemqtt, passwordmqtt)
                      
# The callback for when the client receives a CONNACK response from the server.
def on_connected(client, userdata, flags, rc):
    # Subscribing in on_connect() means that if we lose the connection and
    # reconnect then subscriptions will be renewed.
    if rc==0:
        client.connected_flag=True #set flag
        client.subscribe(mqttopic)
        print("connected OK")
        print("rc =",client.connected_flag)
    else:
        print("Bad connection Returned code=",rc)
        client.bad_connection_flag=False

def on_disconnect2(client, userdata, rc):
    if (rc != 0 and rc != 5):
        print("Unexpected disconnection, will auto-reconnect")
    elif(rc==5):
        print("Getting new credentials!")
        get_mqtt_credentials2()
        client2.username_pw_set(usernamemqtt2, passwordmqtt2)
                      
# The callback for when the client receives a CONNACK response from the server.
def on_connected2(client, userdata, flags, rc):
    # Subscribing in on_connect() means that if we lose the connection and
    # reconnect then subscriptions will be renewed.
    if rc==0:
        client2.connected_flag=True #set flag
        client2.subscribe(mqttopic2)
        print("connected OK")
        print("rc 2 =",client2.connected_flag)
    else:
        print("Bad connection Returned code=",rc)
        client2.bad_connection_flag=False
   
"""       
client = mqtt.Client(str_client_id)   #Creación cliente
client.connect(broker, port)     #Conexión al broker
client.on_disconnect = on_disconnect
client.username_pw_set(usernamemqtt, passwordmqtt)
client.on_connect = on_connected
client.loop_start()
time.sleep(5)
"""
#client2 = mqtt.Client(str_client_id2)   #Creación cliente
#client2.connect(broker, port)     #Conexión al broker
#client2.on_disconnect = on_disconnect2
#client2.username_pw_set(usernamemqtt2, passwordmqtt2)
#client2.on_connect = on_connected2
#client2.loop_start()



def Voltaje():
    global voltaje
    voltaje = random.randint(205, 225)
    Maximo15minVolt(voltaje)
    str_num = {"value":voltaje,"save":1}
    voltajejson = json.dumps(str_num)
    #return msg 


maximo15Volt=[]
data2=[]
data3=[]
acceso = 0
maximoVoltaje15=0
def Maximo15minVolt(voltaje):
    global maximoVoltaje15
    global base
    global maximo15Volt
    global acceso
    basea = datetime.datetime.now()
    print(f'Maximo Voltaje 15: {maximoVoltaje15}')
    if(basea.minute==0 or basea.minute==29 or basea.minute==34 or basea.minute==45): 
         print("paso if")
         if(acceso == 0):
              print("paso if 2")
              acceso = 1
              maximoVoltaje15=max(maximo15Volt)
              data2.insert(1,maximoVoltaje15)
              data3.insert(1,maximoVoltaje15)
              maximo15Volt=[]
         elif(acceso==1):
              print("paso elif ")
              maximo15Volt.append(voltaje)
 
    else:
        maximo15Volt.append(voltaje)
        acceso = 0
        if(len(maximo15Volt)>4):
            indice=np.argmin(maximo15Volt)
            maximo15Volt.pop(indice)
            print(f'maximo15Volt Despúes: {maximo15Volt}')
             
Voltaje()

def Corriente():
    global corriente
    corriente = random.randint(0, 50)
    str_num = {"value":corriente,"save":1}
    corrientejson = json.dumps(str_num)
    #return msg    
Corriente()
b=time.time()
c=time.time()
z=time.time()
x=time.time()

def publish(client):
    
        global b, c 
        a=time.time()
        for i in data["variables"]:

            #    if(data["variables"][i]["variableType"]=="output"):
            #        continue
            if(i["variableFullName"]=="Corriente-CGE"):
                freq = i["variableSendFreq"]
                if(a - b > freq):
                     b=time.time()
                     str_variable = i["variable"]
                     topic1 = topicmqtt + str_variable + "/sdata"
                     result = client.publish(topic1, humedad)
                     status = result[0]
                     
                     if status == 0:
                         print(f"Send humedad: `{humedad}` to topic `{topic1}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic1}")
        
                   
            if(i["variableFullName"]=="Voltaje-CGE"):
                freq = i["variableSendFreq"]
                if(a - c > freq):
                     #print("varlastsend 1: ",varsLastSend[i])
                     c=time.time()
                     str_variable2 = i["variable"]
                     topic2 = topicmqtt + str_variable2 + "/sdata"
                     result = client.publish(topic2, temperature)
                     status = result[0]
                     if status == 0:
                         print(f"Send temperatura: `{temperature}` to topic `{topic2}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic2}")
            
        
        """
        str_variable = data["variables"][0]["variable"]
        print("data:",str_variable)
        topic1 = topicmqtt + str_variable + "/sdata"
        #msg=randomnum2()
        result = client.publish(topic1, humedad)
        #result: [0, 1]
        status = result[0]
        if status == 0:
            print(f"Send `{humedad}` to topic `{topic1}`")
        else:
            print(f"Failed to send message to topic {topic1}")
        """ 


def publish2(client2):
    
        global z, x 
        h=time.time()
        for i in data2["variables"]:

            #    if(data["variables"][i]["variableType"]=="output"):
            #        continue
            if(i["variableFullName"]=="Corriente-CGE"):
                freq = i["variableSendFreq"]
                if(h - z > freq):
                     z=time.time()
                     str_variable = i["variable"]
                     topic = topicmqtt2 + str_variable + "/sdata"
                     result = client2.publish(topic, humedad)
                     status = result[0]
                     
                     if status == 0:
                         print(f"Send humedad: `{humedad}` to topic `{topic}` con freq: {freq}")
                     else:
                         print(f"Failed to send message to topic {topic}")

def excelcreate():
     global dest_filename
     global sheet2
     global sheet3
     global sheet4
     exceltime=datetime.datetime.now()
     book = Workbook()
     dest_filename = f'prueba_Medidor: {exceltime}.xlsx'
     sheet = book.active
     sheet.title = "Resumen Reportes"
     sheet2 = book.create_sheet("CGE")
     sheet3 = book.create_sheet("Carga")
     sheet4 = book.create_sheet("Paneles")
     headings=['Fecha y Hora'] + list(['Voltaje', 'Corriente']) #dict_keys(['math', 'science', 'english', 'gym'])
     sheet2.append(headings)
     sheet3.append(headings)
     sheet4.append(headings)
     book.save(filename = dest_filename)

excelcreate()

accesoexcel=0
def ExcelData():
       global data2
       global accesoexcel
       base=datetime.datetime.now()
       if(base.minute==1 or base.minute==30 or base.minute==35 or base.minute==46):
               if(accesoexcel==0):              
                       workbook=openpyxl.load_workbook(filename = dest_filename)
                       sheet2 = workbook["CGE"]
                       data2.insert(0,datetime.datetime.now())
                       sheet2.append(list(data2))
                       print(data2)
                       print("Datos Insertados Correctamente!")
                       workbook.save(filename = dest_filename)
                       data2=[]
                       accesoexcel=1
       else:
               accesoexcel=0

accesoexcel2=0
def ExcelData2():
       global data3
       global accesoexcel2
       base=datetime.datetime.now()
       if(base.minute==1 or base.minute==30 or base.minute==35 or base.minute==46):
               if(accesoexcel2==0):              
                       workbook=openpyxl.load_workbook(filename = dest_filename)
                       sheet3 = workbook["Carga"]
                       data3.insert(0,datetime.datetime.now())
                       sheet3.append(list(data3))
                       print(data3)
                       print("Datos Insertados Correctamente!")
                       workbook.save(filename = dest_filename)
                       data3=[]
                       accesoexcel2=1
       else:
               accesoexcel2=0



global accesoemail
accesoemail = 0
#exceltime = datetime.datetime.now()

while True:
          
          Voltaje()
          #Corriente()
          ExcelData()
          ExcelData2()
          #ExcelDataCarga()
          excel=datetime.datetime.now()
          print(excel)
          if(excel.hour==11 and excel.minute==36):
                if(accesoemail==0):
                          accesoemail=1
                          print("Entro a SendEmail")
                          #SendEmail()
                          time.sleep(5)
                          os.remove(dest_filename)
                          excelcreate()


          else:
              accesoemail=0
          time.sleep(20)
    
"""
while client.connected_flag:
        Humedad()
        Temperature()
        #publish(client)
        #publish2(client2)
   
    #publish(client2)
    #time.sleep(5)
"""



"""
{
    "password": "7dnkOzqUwL",
    "topic": "610a8d733d1a8852b4e6b775/12345678/",
    "username": "CRE6Gpub8q",
    "variables": [
        {
            "variable": "O1DM08hKvJ",
            "variableFullName": "Voltaje",
            "variableSendFreq": 10,
            "variableType": "input"
        },
        {
            "variable": "mmdKbcHfSA",
            "variableFullName": "Corriente",
            "variableSendFreq": 15,
            "variableType": "input"
        }
    ]
}
"""
