##Esta sirve

from connector import iot_ser_db
from connector import local_db
import json
import pymongo
import datetime
import os
import openpyxl
import pandas as pd

userId="63d87d13afeae04cb7827725"
dId=1212
db = iot_ser_db()

def list_data_db_insert_dispositive_historic(df):
    df_with_id=df.assign(userId=userId)
    list_filtered = df_with_id.to_dict('records')
    col_use = db.Dispositive_Variables
    col_use.insert_many(list_filtered)


def list_data_db_insert_max_day(df):
    df_with_id=df.assign(userId=userId)
    list_filtered = df_with_id.to_dict('records')
    col_use = db.monthmaxs
    col_use.insert_many(list_filtered)

def list_data_db_insert_max_month(df1,df2,df3):
    print("Paso2")
    #print(df1[0:3])
    #print(df2[0:3])
    #print(df3[0:3])
    df12=df1['time']
    print(df12[0:3])
    df11=df1['energia_fase_1_redcompañia_mensual']
    print("  ")
    print(df11[0:3])
    df22=df2['energia_fase_2_redcompañia_mensual']
    print(df22[0:3])
    df33=df3['energia_fase_3_redcompañia_mensual']
    print(df33[0:3])
    df_red_compañia=pd.concat([df12,df11,df22,df33], axis=1)
    df_red_compañia=df_red_compañia.assign(userId=userId,dId=dId)
    #df=pd.DataFrame(df_red_compañia,columns=['time',f'energia_fase_1_carga_mensual',f'energia_fase_2_carga_mensual',f'energia_fase_3_carga_mensual','userId','dId'])
    df_red_compañia = df_red_compañia.rename(columns={'energia_fase_1_redcompañia_mensual':'energia_fase_1_redcompañia_mensual','energia_fase_2_consumocliente_mensual':'energia_fase_2_redcompañia_mensual','energia_fase_3_redcompañia_mensual':'energia_fase_3_redcompañia_mensual'})
    print(df_red_compañia[0:3])
    col_use = db.monthmaxs 
    col_use.insert_many(df_red_compañia.to_dict('records'))
    
    

def main():
    #DiasExcel = [filecsv for filecsv in os.listdir('Reportes/') if filecsv[10:15]!="_save"]                   
    #DiasExcel=sorted(DiasExcel)
    #for archivo_csv in DiasExcel:
    archivo_csv=f'Reportes/2022-09-30.xlsx'
    try:
        #VariablesDispositivos(archivo_csv)
        MaximosDiariosMesFase(archivo_csv)  
    except Exception as e:
         print(f"El formato del archivo no es el correspondiente: {archivo_csv} -> {str(e)} ") 

def MaximosDiariosMesFase(archivo_csv):
    df=pd.read_excel(open(archivo_csv, 'rb'),sheet_name='Energia Fase 1 Mensual')
    df2=pd.read_excel(open(archivo_csv, 'rb'),sheet_name='Energia Fase 2 Mensual')
    df3=pd.read_excel(open(archivo_csv, 'rb'),sheet_name='Energia Fase 3 Mensual')
    try:
        df.columns = ['time','energia_fase_1_redcompañia_mensual','energia_fase_1_consumocliente_mensual','energia_fase_1_centralfotovoltaica_mensual','save','max1','max2','max3']
        df=TimetoTimestampdf(df)
     
        df2.columns = ['time','energia_fase_2_redcompañia_mensual','energia_fase_2_consumocliente_mensual','energia_fase_2_centralfotovoltaica_mensual','save','max1','max2','max3']
        df2=TimetoTimestampdf(df2)  
   
        df3.columns = ['time','energia_fase_3_redcompañia_mensual','energia_fase_3_consumocliente_mensual','energia_fase_3_centralfotovoltaica_mensual','save','max1','max2','max3']
        df3=TimetoTimestampdf(df3)
        list_data_db_insert_max_month(df,df2,df3)

        #workbook=openpyxl.load_workbook(filename = archivo_csv)
        #sheet = workbook[f"Energia Fase 1 Mensual"]
        ##if(sheet[f'E{LargeSheet-1}']==1):
        ##print(df)
        #LargeSheet=sheet.max_row
        #print("Insertando save = 1")
        #for row in range(1, LargeSheet):
        #  sheet[f'E{row+1}']=1
        #workbook.save(filename = archivo_csv)
            #os.rename(f'Reportes/{f}', f'Reportes/{f[0:10]}_save.xlsx')
            
    except Exception as e:
        print(f"Excepcion en Reporte: {archivo_csv} -> {str(e)} ")


def MaximosDiariosMesFase2(archivo_csv):
    df=pd.read_excel(open(archivo_csv, 'rb'),sheet_name='Energia Fase 2 Mensual')
    try:
        df.columns = ['time','energia_fase_2_redcompañia_mensual','energia_fase_2_consumocliente_mensual','energia_fase_2_centralfotovoltaica_mensual','save','max1','max2','max3']
        #print(df[0:1])
        df=TimetoTimestampdf(df)
        list_data_db_insert_max_month(df,2)
        workbook=openpyxl.load_workbook(filename = archivo_csv)
        sheet = workbook[f"Energia Fase 2 Mensual"]
        LargeSheet=sheet.max_row
        for row in range(1, LargeSheet):
          sheet[f'E{row+1}']=1
        workbook.save(filename = archivo_csv)
            #os.rename(f'Reportes/{f}', f'Reportes/{f[0:10]}_save.xlsx')
            
    except Exception as e:
        print(f"Excepcion en Reporte: {archivo_csv} -> {str(e)} ")



def MaximosDiariosMesFase3(archivo_csv):
    print("Paso")
    df=pd.read_excel(open(archivo_csv, 'rb'),sheet_name='Energia Fase 1 Mensual')
    try:
        df.columns = ['time','energia_fase_3_redcompañia_mensual','energia_fase_3_consumocliente_mensual','energia_fase_3_centralfotovoltaica_mensual','save','max1','max2','max3']
        #print(df[0:1])
        df=TimetoTimestampdf(df)
        list_data_db_insert_max_month(df,3)
        workbook=openpyxl.load_workbook(filename = archivo_csv)
        sheet = workbook[f"Energia Fase 3 Mensual"]
        LargeSheet=sheet.max_row
        print("Insertando save = 1")
        for row in range(1, LargeSheet):
          sheet[f'E{row+1}']=1
        workbook.save(filename = archivo_csv)
            #os.rename(f'Reportes/{f}', f'Reportes/{f[0:10]}_save.xlsx')
            
    except Exception as e:
        print(f"Excepcion en Reporte: {archivo_csv} -> {str(e)} ")

def MaximosDiarios(archivo_csv):
    df=pd.read_excel(open(f'Reportes/{archivo_csv}', 'rb'),sheet_name='MaxHora Fase 1 Diario')
    try:
        df.columns = ['Fecha_y_Hora','Energia_Fase_1_REDCompañia_Diario','Energia_Fase_1_ConsumoCliente_Diario','Energia_Fase_1_CentralFotovoltaica_Diario']
        #print(df[0:2])
        c=-1
        for i in df['Fecha_y_Hora'].values:
            c=c+1
            df['Fecha_y_Hora'].values[c]=f"{archivo_csv[0:10]} {df['Fecha_y_Hora'].values[c]}:00"
        print(df[0:1])
        df=TimetoTimestampdf(df)
        workbook=openpyxl.load_workbook(filename = f'Reportes/{archivo_csv}')
        sheet = workbook[f"MaxHora Fase 1 Diario"]
        #DA = workbook.get_sheet_by_name('MaxHora Fase 1 Diario')
        LargeSheet=sheet.max_row
        if(sheet[f'E{LargeSheet-1}']!=1):
            list_data_db_insert_max_day(df)
            print("Insertando save = 1")
            for row in range(1, LargeSheet):
              sheet[f'E{row+1}']=1
        workbook.save(filename = f'Reportes/{archivo_csv}')
            #os.rename(f'Reportes/{f}', f'Reportes/{f[0:10]}_save.xlsx')
            
    except Exception as e:
        print(f"Excepcion en Reporte: {archivo_csv} -> {str(e)} ")

def VariablesDispositivos(archivo_csv):
    df=pd.read_excel(open(f'Reportes/{archivo_csv}', 'rb'),sheet_name='Var Dispositivos')
    try:
        df.columns = ['Fecha_y_Hora','T_Raspberry','Uso_CPU','RAM','T_ESP32']
        df=TimetoTimestampdf(df)
        workbook=openpyxl.load_workbook(filename = f'Reportes/{archivo_csv}')
        sheet = workbook[f"Var Dispositivos"]
        LargeSheet=len(sheet["RAM"])
        if(sheet[f'F{LargeSheet-1}']!=1):
            list_data_db_insert_dispositive_historic(df)
            print("Insertando save = 1")
            for row in range(1, LargeSheet):
              sheet[f'F{row+1}']=1
            #os.rename(f'Reportes/{f}', f'Reportes/{f[0:10]}_save.xlsx')
        workbook.save(filename = f'Reportes/{archivo_csv}')
    except Exception as e:
                    print(f"Excepcion en Reporte: {archivo_csv} -> {str(e)} ")

def TimetoTimestampdf(df):
    try:
        print("TimetoTimestampdf")
        c=-1
        for i in df['time'].values:
           c=c+1
           date_time_str = str(df['time'].values[c])
           date_time_str=date_time_str[0:19]
           date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d')
           #date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d %H:%M:%S')
           #print('Time:', date_time_obj.time())##print('Date:', date_time_obj.date())#print('Date-time:', (date_time_obj))
           df['time'].values[c]=int(date_time_obj.timestamp())
        return df
    except Exception as e:
        print(f"Excepcion en TimetoTimestampdf -> {str(e)} ") 


##mycol = db["Maximos_Diarios"]
##mycol.drop()#
main()